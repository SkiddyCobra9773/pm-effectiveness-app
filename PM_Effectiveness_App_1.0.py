import io, re, datetime
import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# ── palette ──────────────────────────────────────────────────────────────────
DARK_BG   = "1F2D3D"; MID_BG  = "2E4057"; HDR_BG  = "3A5068"
LIGHT_ROW = "EAF1F8"; WHITE   = "FFFFFF"; YELLOW  = "FFF2CC"
RED_FLG   = "FCE4D6"; GREEN_FLG = "E2EFDA"; ORANGE  = "FCE9D6"
RAG_RED   = "C00000"; RAG_AMBER = "ED7D31"; RAG_GREEN = "375623"
RAG_RED_BG   = "FCE4D6"; RAG_AMBER_BG = "FCE9D6"; RAG_GREEN_BG = "E2EFDA"

PILLARS = [
    "1. Compliance & Scheduling",
    "2. Failure Prevention",
    "3. Failure Mode Coverage",
    "4. Work Order Quality",
    "5. Backlog Health",
    "6. Cost Visibility",
    "7. Equipment Reliability",
]

# ── style helpers ─────────────────────────────────────────────────────────────
def wfill(h):   return PatternFill("solid", start_color=h, fgColor=h)
def cal(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def title_row(ws, text, row, end_col="J"):
    ec = end_col
    ws.merge_cells(f"B{row}:{ec}{row}")
    c = ws[f"B{row}"]; c.value = text
    c.font = Font(name="Arial", bold=True, color=WHITE, size=13)
    c.fill = wfill(DARK_BG); c.alignment = cal(); ws.row_dimensions[row].height = 26

def sub_row(ws, text, row, end_col="J"):
    ws.merge_cells(f"B{row}:{end_col}{row}")
    c = ws[f"B{row}"]; c.value = text
    c.font = Font(name="Arial", italic=True, color="AAAAAA", size=9)
    c.fill = wfill(DARK_BG); c.alignment = cal(wrap=True); ws.row_dimensions[row].height = 16

def sec(ws, text, row, end_col="J", bg=MID_BG):
    ws.merge_cells(f"B{row}:{end_col}{row}")
    c = ws[f"B{row}"]; c.value = text
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = wfill(bg); c.alignment = cal("left"); ws.row_dimensions[row].height = 18

def hdr(ws, col, row, text, end_col=None):
    ref = f"{col}{row}"
    if end_col: ws.merge_cells(f"{ref}:{end_col}{row}")
    ws[ref].value = text
    ws[ref].font = Font(name="Arial", bold=True, color=WHITE, size=9)
    ws[ref].fill = wfill(HDR_BG); ws[ref].alignment = cal(wrap=True); ws[ref].border = bdr()

def dc(ws, col, row, val, bg=WHITE, bold=False, align="center", wrap=False, color="1F2D3D"):
    ws[f"{col}{row}"].value = val
    ws[f"{col}{row}"].font = Font(name="Arial", size=9, color=color, bold=bold)
    ws[f"{col}{row}"].fill = wfill(bg)
    ws[f"{col}{row}"].alignment = cal(align, wrap=wrap)
    ws[f"{col}{row}"].border = bdr()

def widths(ws, m):
    for c, w in m.items(): ws.column_dimensions[c].width = w

def rag_bg(status):
    return {"GREEN": RAG_GREEN_BG, "AMBER": RAG_AMBER_BG, "RED": RAG_RED_BG}.get(status, LIGHT_ROW)

def rag_fg(status):
    return {"GREEN": RAG_GREEN, "AMBER": RAG_AMBER, "RED": RAG_RED}.get(status, "1F2D3D")

def rag_label(status):
    return {"GREEN": "✓ GOOD", "AMBER": "⚠ WATCH", "RED": "✗ ACTION"}.get(status, "—")

# ── helpers ───────────────────────────────────────────────────────────────────
def hf(s, f): return f in str(s).split() if pd.notna(s) else False

def pfreq(d):
    if pd.isna(d): return None
    m = re.search(r'(\d+)\s*WK', d, re.IGNORECASE)
    if m: return int(m.group(1)) * 7
    if re.search(r'5W', d, re.IGNORECASE): return 35
    return None

# ── cluster definitions ───────────────────────────────────────────────────────
CLUSTERS = [
    ("Glue System",           ['glue','nozzle','nordson','versa blue']),
    ("Gearbox / Shaft",       ['gearbox','gear box','shaft']),
    ("Pusher / Movement",     ['pusher','layer pusher']),
    ("Fuses / Electrical",    ['fuse','vfd','not moving','electrical']),
    ("Belts / Chains / Lugs", ['belt','chain','lug']),
    ("Carton Handling",       ['carton','blow off']),
    ("Drive / Other",         ['drive','resolver','wadding']),
]

# ═════════════════════════════════════════════════════════════════════════════
# DATA PREPARATION
# ═════════════════════════════════════════════════════════════════════════════
def prep_data(file_obj):
    df = pd.read_excel(file_obj)
    df.rename(columns={
        'Work order classification for preventive and corrective processing metrics': 'WO_Classification'
    }, inplace=True)
    df['Basic Start Date'] = pd.to_datetime(df['Basic Start Date'], errors='coerce')
    df['Created On']       = pd.to_datetime(df['Created On'],       errors='coerce')
    df['is_TECO'] = df['System Status'].apply(lambda x: hf(x, 'TECO'))
    df['is_CLSD'] = df['System Status'].apply(lambda x: hf(x, 'CLSD'))
    df['is_REL']  = df['System Status'].apply(lambda x: hf(x, 'REL'))
    df['is_open'] = ~df['is_TECO'] & ~df['is_CLSD']

    # Planned hours column (may not exist)
    df['Planned Hours'] = df.get('Planned Hours', pd.Series([np.nan]*len(df)))

    pm  = df[df['OrderType'] == 'GM09'].copy().sort_values('Basic Start Date')
    bkd = df[df['OrderType'] == 'GM02'].copy().sort_values('Basic Start Date')
    fuc = df[df['OrderType'].isin(['GM01','PM01'])].copy().sort_values('Basic Start Date')

    # Reference date (latest Created On in dataset, or today)
    ref_date = df['Created On'].max()
    if pd.isna(ref_date): ref_date = pd.Timestamp.today()

    return df, pm, bkd, fuc, ref_date


# ═════════════════════════════════════════════════════════════════════════════
# PILLAR ANALYSIS FUNCTIONS
# ═════════════════════════════════════════════════════════════════════════════

def pillar1_compliance(pm, bkd, fuc, df):
    """Compliance & Scheduling"""
    teco_pct = pm['is_TECO'].mean() * 100 if len(pm) else 0

    # Schedule adherence
    pm_h1 = pm[pm['Basic Start Date'].between('2026-01-01','2026-06-30')].copy()
    sched_rows = []
    for desc, grp in pm_h1.groupby('Description'):
        grp = grp.sort_values('Basic Start Date'); interval = pfreq(desc)
        if interval is None or len(grp) < 2: continue
        dates = grp['Basic Start Date'].tolist()
        for i in range(1, len(dates)):
            actual = (dates[i]-dates[i-1]).days; var = actual - interval
            status = ('On time' if abs(var) <= max(1, interval*0.2)
                      else 'Early' if var < 0
                      else 'Late (≤50%)' if var <= interval*0.5
                      else 'Significantly late')
            sched_rows.append({'PM_Description': desc, 'Expected_Days': interval,
                'From': dates[i-1].strftime('%Y-%m-%d'), 'To': dates[i].strftime('%Y-%m-%d'),
                'Actual_Gap': actual, 'Variance': var, 'Status': status})
    sched = pd.DataFrame(sched_rows) if sched_rows else pd.DataFrame()
    if len(sched):
        vc       = sched['Status'].value_counts()
        on_time  = vc.get('On time', 0); early = vc.get('Early', 0)
        late_ct  = vc.get('Late (≤50%)', 0) + vc.get('Significantly late', 0)
        total_i  = len(sched); adh_pct = 100*on_time/total_i
        by_type  = sched.groupby(['PM_Description','Expected_Days']).agg(
            Intervals=('Actual_Gap','count'), Avg_Gap=('Actual_Gap','mean'),
            Avg_Var=('Variance','mean'), Max_Late=('Variance','max'),
            On_Time=('Status', lambda x: (x=='On time').sum())).reset_index()
        by_type['OT_Pct'] = (by_type['On_Time']/by_type['Intervals']*100).round(0).astype(int)
    else:
        on_time = early = late_ct = total_i = 0; adh_pct = 0; by_type = pd.DataFrame()

    # Cycle time: Created On → Basic Start Date (for completed), Created On → ref for open
    # Use Basic Start Date as proxy for work start
    pm_with_dates = pm.dropna(subset=['Created On','Basic Start Date']).copy()
    pm_with_dates['Cycle_Days'] = (pm_with_dates['Basic Start Date'] - pm_with_dates['Created On']).dt.days
    pm_with_dates = pm_with_dates[pm_with_dates['Cycle_Days'] >= 0]
    avg_cycle = pm_with_dates['Cycle_Days'].mean() if len(pm_with_dates) else 0
    ct_lt7  = (pm_with_dates['Cycle_Days'] <= 7).sum()
    ct_7_30 = ((pm_with_dates['Cycle_Days'] > 7) & (pm_with_dates['Cycle_Days'] <= 30)).sum()
    ct_30p  = (pm_with_dates['Cycle_Days'] > 30).sum()

    # RAG
    rag = "GREEN" if (teco_pct >= 90 and adh_pct >= 80) else ("AMBER" if (teco_pct >= 75 or adh_pct >= 65) else "RED")

    return {
        'teco_pct': teco_pct, 'adh_pct': adh_pct,
        'on_time': on_time, 'early': early, 'late_ct': late_ct, 'total_i': total_i,
        'avg_cycle': avg_cycle, 'ct_lt7': ct_lt7, 'ct_7_30': ct_7_30, 'ct_30p': ct_30p,
        'sched': sched, 'by_type': by_type, 'rag': rag,
        'key_metric': f"TECO {teco_pct:.0f}%  |  Adherence {adh_pct:.0f}%",
    }


def pillar2_failure_prevention(pm, bkd):
    """Failure Prevention"""
    months_str = ['Jan','Feb','Mar','Apr','May','Jun']
    month_keys = ['2026-01','2026-02','2026-03','2026-04','2026-05','2026-06']
    month_map  = {k: i for i, k in enumerate(month_keys)}

    bkd_mo = [0]*6
    bkd_h1 = bkd[bkd['Basic Start Date'].between('2026-01-01','2026-06-30')].dropna(subset=['Basic Start Date'])
    for _, row in bkd_h1.iterrows():
        k = row['Basic Start Date'].strftime('%Y-%m')
        if k in month_map: bkd_mo[month_map[k]] += 1

    # MTBF monthly (actual inter-event)
    bkd_h1c = bkd_h1.copy(); bkd_h1c['Month'] = bkd_h1c['Basic Start Date'].dt.to_period('M')
    mtbf_30     = [round(30/c, 1) if c > 0 else 0 for c in bkd_mo]
    mtbf_actual = []
    for mk in month_keys:
        grp = bkd_h1c[bkd_h1c['Month'].astype(str)==mk].sort_values('Basic Start Date')
        d   = grp['Basic Start Date'].tolist()
        if len(d) >= 2:
            mtbf_actual.append(round(np.mean([(d[i]-d[i-1]).days for i in range(1,len(d))]),1))
        else:
            mtbf_actual.append(0)

    # Rolling 3-month avg & trend direction
    rolling_3 = []
    for i in range(len(mtbf_30)):
        window = [m for m in mtbf_30[max(0,i-2):i+1] if m > 0]
        rolling_3.append(round(np.mean(window),1) if window else 0)

    h1_vals    = [m for m in mtbf_30 if m > 0]
    first3     = np.mean([m for m in mtbf_30[:3] if m > 0]) if any(m>0 for m in mtbf_30[:3]) else 0
    last3      = np.mean([m for m in mtbf_30[3:] if m > 0]) if any(m>0 for m in mtbf_30[3:]) else 0
    if first3 and last3:
        pct_chg = (last3-first3)/first3*100
        trend_dir = "Improving ↑" if pct_chg > 10 else ("Declining ↓" if pct_chg < -10 else "Stable →")
        trend_clr = "GREEN" if pct_chg > 10 else ("RED" if pct_chg < -10 else "AMBER")
    else:
        pct_chg = 0; trend_dir = "Insufficient data"; trend_clr = "AMBER"

    overall_mtbf = round(np.mean(h1_vals),1) if h1_vals else 0

    # PM→Breakdown pairs
    pm_bkd_rows = []
    for _, bd in bkd.iterrows():
        prior = pm[pm['Basic Start Date'] <= bd['Basic Start Date']]
        if not len(prior): continue
        last = prior.iloc[-1]; days = (bd['Basic Start Date'] - last['Basic Start Date']).days
        cat = ("Same day" if days == 0 else "1-3 days" if days <= 3
               else "4-7 days" if days <= 7 else "8+ days")
        pm_bkd_rows.append({'PM_WO': last['WorkOrderNumber'], 'PM_Desc': last['Description'],
            'PM_Date': last['Basic Start Date'].strftime('%Y-%m-%d'),
            'BKD_WO': bd['WorkOrderNumber'], 'BKD_Desc': bd['Description'],
            'BKD_Date': bd['Basic Start Date'].strftime('%Y-%m-%d') if pd.notna(bd['Basic Start Date']) else '—',
            'Days': days, 'Category': cat})
    pm_bkd = pd.DataFrame(pm_bkd_rows) if pm_bkd_rows else pd.DataFrame()

    rag = trend_clr  # trend is the main signal

    return {
        'months_str': months_str, 'bkd_mo': bkd_mo,
        'mtbf_30': mtbf_30, 'mtbf_actual': mtbf_actual, 'rolling_3': rolling_3,
        'overall_mtbf': overall_mtbf, 'trend_dir': trend_dir, 'trend_clr': trend_clr,
        'pct_chg': pct_chg, 'pm_bkd': pm_bkd,
        'rag': rag,
        'key_metric': f"MTBF {overall_mtbf}d  |  Trend {trend_dir}",
    }


def pillar3_coverage(pm, bkd):
    """Failure Mode Coverage"""
    pm_descs_lower = ' '.join(pm['Description'].fillna('').str.lower().tolist())

    coverage_rows = []
    for name, keywords in CLUSTERS:
        bkd_mask = bkd['Description'].str.lower().str.contains('|'.join(keywords), na=False)
        bkd_count = bkd_mask.sum()
        pm_covered = any(kw in pm_descs_lower for kw in keywords)
        gap = bkd_count > 0 and not pm_covered
        coverage_rows.append({
            'Cluster': name, 'Keywords': ', '.join(keywords),
            'BKD_Count': bkd_count, 'PM_Covered': pm_covered,
            'Gap': gap,
            'Status': "✗ GAP — no PM task" if gap else ("✓ Covered" if bkd_count > 0 else "No BKDs"),
            'Priority': "HIGH" if (gap and bkd_count >= 3) else ("MED" if gap else "OK"),
        })
    cov_df = pd.DataFrame(coverage_rows)
    gaps     = cov_df[cov_df['Gap']].shape[0]
    high_gaps= cov_df[cov_df['Priority']=='HIGH'].shape[0]

    rag = "GREEN" if gaps == 0 else ("AMBER" if high_gaps == 0 else "RED")

    return {
        'cov_df': cov_df, 'gaps': gaps, 'high_gaps': high_gaps,
        'rag': rag,
        'key_metric': f"{gaps} coverage gaps  |  {high_gaps} high-priority",
    }


def pillar4_wo_quality(pm, bkd, fuc):
    """Work Order Quality — zero-hour + planned vs actual"""
    pm_zero  = (pm['Actual Hours']  == 0).sum()
    bkd_zero = (bkd['Actual Hours'] == 0).sum()
    fuc_zero = (fuc['Actual Hours'] == 0).sum() if len(fuc) else 0

    pm_zero_pct  = 100*pm_zero/len(pm)   if len(pm)  else 0
    bkd_zero_pct = 100*bkd_zero/len(bkd) if len(bkd) else 0
    fuc_zero_pct = 100*fuc_zero/len(fuc)  if len(fuc) else 0

    # Planned vs Actual (only if planned hours column has data)
    has_planned = pm['Planned Hours'].notna().any()
    if has_planned:
        pva = pm.dropna(subset=['Planned Hours','Actual Hours']).copy()
        pva = pva[(pva['Planned Hours'] > 0) & (pva['Actual Hours'] > 0)]
        pva['Variance'] = pva['Actual Hours'] - pva['Planned Hours']
        pva['Var_Pct']  = pva['Variance'] / pva['Planned Hours'] * 100
        avg_var_pct = pva['Var_Pct'].mean() if len(pva) else 0
        over_cnt  = (pva['Variance'] > 0).sum()
        under_cnt = (pva['Variance'] < 0).sum()
        on_cnt    = (pva['Variance'] == 0).sum()
    else:
        pva = pd.DataFrame(); avg_var_pct = 0; over_cnt = under_cnt = on_cnt = 0

    rag = ("GREEN" if (bkd_zero_pct < 20 and pm_zero_pct == 0)
           else "AMBER" if bkd_zero_pct < 60 else "RED")

    return {
        'pm_zero': pm_zero, 'bkd_zero': bkd_zero, 'fuc_zero': fuc_zero,
        'pm_zero_pct': pm_zero_pct, 'bkd_zero_pct': bkd_zero_pct, 'fuc_zero_pct': fuc_zero_pct,
        'has_planned': has_planned, 'pva': pva,
        'avg_var_pct': avg_var_pct, 'over_cnt': over_cnt, 'under_cnt': under_cnt, 'on_cnt': on_cnt,
        'rag': rag,
        'key_metric': f"BKD zero-hr {bkd_zero_pct:.0f}%  |  PM zero-hr {pm_zero_pct:.0f}%",
    }


def pillar5_backlog(df, ref_date):
    """Backlog Health"""
    open_df = df[df['is_open']].copy()
    open_df['Age_Days'] = (ref_date - open_df['Created On']).dt.days.clip(lower=0)

    by_type = open_df.groupby('OrderType').agg(
        Count=('WorkOrderNumber','count'),
        Avg_Age=('Age_Days','mean'),
        Max_Age=('Age_Days','max')).reset_index()

    age_0_30  = (open_df['Age_Days'] <= 30).sum()
    age_31_90 = ((open_df['Age_Days'] > 30) & (open_df['Age_Days'] <= 90)).sum()
    age_91p   = (open_df['Age_Days'] > 90).sum()
    total_open = len(open_df)
    avg_age    = open_df['Age_Days'].mean() if total_open else 0
    oldest_age = open_df['Age_Days'].max() if total_open else 0

    # Monthly new WOs vs TECO'd (H1 2026) to show backlog trend
    months_str = ['Jan','Feb','Mar','Apr','May','Jun']
    month_keys = ['2026-01','2026-02','2026-03','2026-04','2026-05','2026-06']
    new_mo = [0]*6; closed_mo = [0]*6
    for _, row in df.iterrows():
        k = row['Created On'].strftime('%Y-%m') if pd.notna(row['Created On']) else None
        if k in month_keys: new_mo[month_keys.index(k)] += 1
        k2 = row['Basic Start Date'].strftime('%Y-%m') if pd.notna(row['Basic Start Date']) else None
        if k2 in month_keys and row['is_TECO']: closed_mo[month_keys.index(k2)] += 1
    net_mo = [n-c for n, c in zip(new_mo, closed_mo)]
    running_net = []
    running = 0
    for n in net_mo:
        running += n; running_net.append(running)

    rag = ("GREEN" if (age_91p == 0 and avg_age < 30)
           else "AMBER" if age_91p <= 5
           else "RED")

    return {
        'open_df': open_df, 'by_type': by_type,
        'age_0_30': age_0_30, 'age_31_90': age_31_90, 'age_91p': age_91p,
        'total_open': total_open, 'avg_age': avg_age, 'oldest_age': oldest_age,
        'months_str': months_str, 'new_mo': new_mo, 'closed_mo': closed_mo,
        'net_mo': net_mo, 'running_net': running_net,
        'rag': rag,
        'key_metric': f"{total_open} open  |  {age_91p} aged 90d+  |  avg {avg_age:.0f}d",
    }


def pillar6_cost(pm, bkd, fuc, df):
    """Cost Visibility"""
    all_hrs   = df['Actual Hours'].sum()
    pm_hrs    = pm['Actual Hours'].sum()
    bkd_hrs   = bkd['Actual Hours'].sum()
    fuc_hrs   = fuc['Actual Hours'].sum()
    unknown_hrs = all_hrs - pm_hrs - bkd_hrs - fuc_hrs

    bkd_zero  = (bkd['Actual Hours'] == 0).sum()
    bkd_zero_pct = 100*bkd_zero/len(bkd) if len(bkd) else 0
    fuc_zero  = (fuc['Actual Hours'] == 0).sum()
    fuc_zero_pct = 100*fuc_zero/len(fuc) if len(fuc) else 0

    # Estimated hidden hours: avg BKD hours (from those that have hours) * zero count
    bkd_with_hrs = bkd[bkd['Actual Hours'] > 0]
    avg_bkd_hr   = bkd_with_hrs['Actual Hours'].mean() if len(bkd_with_hrs) else 0
    est_hidden   = avg_bkd_hr * bkd_zero

    # PM vs Corrective labor split
    corr_hrs  = bkd_hrs + fuc_hrs
    planned_pct = 100*pm_hrs/all_hrs if all_hrs else 0
    corr_pct    = 100*corr_hrs/all_hrs if all_hrs else 0

    rag = "RED" if bkd_zero_pct > 50 else ("AMBER" if bkd_zero_pct > 20 else "GREEN")

    return {
        'all_hrs': all_hrs, 'pm_hrs': pm_hrs, 'bkd_hrs': bkd_hrs, 'fuc_hrs': fuc_hrs,
        'bkd_zero': bkd_zero, 'bkd_zero_pct': bkd_zero_pct,
        'fuc_zero': fuc_zero, 'fuc_zero_pct': fuc_zero_pct,
        'avg_bkd_hr': avg_bkd_hr, 'est_hidden': est_hidden,
        'planned_pct': planned_pct, 'corr_pct': corr_pct,
        'rag': rag,
        'key_metric': f"BKD hrs captured: {100-bkd_zero_pct:.0f}%  |  ~{est_hidden:.0f} hrs hidden",
    }


def pillar7_reliability(bkd, df, ref_date):
    """Equipment Reliability — MTBF + MTTR"""
    bkd_sorted = bkd.dropna(subset=['Basic Start Date']).sort_values('Basic Start Date')
    bd_dates   = bkd_sorted['Basic Start Date'].tolist()
    gaps       = [(bd_dates[i]-bd_dates[i-1]).days for i in range(1, len(bd_dates))]
    mtbf_overall = round(np.mean(gaps),1) if gaps else 0
    mtbf_min     = min(gaps) if gaps else 0
    mtbf_max     = max(gaps) if gaps else 0

    # MTTR: actual hours on BKDs that have hours
    bkd_with_hrs = bkd[bkd['Actual Hours'] > 0]
    mttr         = round(bkd_with_hrs['Actual Hours'].mean(),1) if len(bkd_with_hrs) else None
    mttr_n       = len(bkd_with_hrs)
    mttr_note    = f"Based on {mttr_n}/{len(bkd)} BKD WOs with actual hours recorded" if mttr else "No actual hours on BKD WOs — MTTR cannot be calculated"

    # Breakdown rate by month (H1 2026)
    months_str = ['Jan','Feb','Mar','Apr','May','Jun']
    month_keys = ['2026-01','2026-02','2026-03','2026-04','2026-05','2026-06']
    bkd_mo = [0]*6
    bkd_h1 = bkd[bkd['Basic Start Date'].between('2026-01-01','2026-06-30')].dropna(subset=['Basic Start Date'])
    for _, row in bkd_h1.iterrows():
        k = row['Basic Start Date'].strftime('%Y-%m')
        if k in month_keys: bkd_mo[month_keys.index(k)] += 1

    # MTTR by month
    mttr_mo = []
    bkd_h1c = bkd_h1.copy(); bkd_h1c['Month'] = bkd_h1c['Basic Start Date'].dt.to_period('M')
    for mk in month_keys:
        grp = bkd_h1c[(bkd_h1c['Month'].astype(str)==mk) & (bkd_h1c['Actual Hours']>0)]
        mttr_mo.append(round(grp['Actual Hours'].mean(),1) if len(grp) else 0)

    # Best/worst months
    best_idx  = bkd_mo.index(min(bkd_mo)) if bkd_mo else 0
    worst_idx = bkd_mo.index(max(bkd_mo)) if bkd_mo else 0

    rag = ("GREEN" if mtbf_overall > 7 else "AMBER" if mtbf_overall > 3 else "RED")

    return {
        'mtbf_overall': mtbf_overall, 'mtbf_min': mtbf_min, 'mtbf_max': mtbf_max,
        'mttr': mttr, 'mttr_n': mttr_n, 'mttr_note': mttr_note,
        'months_str': months_str, 'bkd_mo': bkd_mo, 'mttr_mo': mttr_mo,
        'best_idx': best_idx, 'worst_idx': worst_idx, 'gaps': gaps,
        'rag': rag,
        'key_metric': f"MTBF {mtbf_overall}d  |  MTTR {'—' if not mttr else str(mttr)+'h'}",
    }


# ═════════════════════════════════════════════════════════════════════════════
# WORKBOOK BUILDER
# ═════════════════════════════════════════════════════════════════════════════
def build_workbook(df, pm, bkd, fuc, ref_date, asset_label, analyses):
    p1, p2, p3, p4, p5, p6, p7 = analyses
    date_range = (f"{df['Created On'].min().strftime('%b %Y')}–"
                  f"{df['Created On'].max().strftime('%b %Y')}")
    wb = Workbook()

    # ── SHEET 0: Executive Scorecard ─────────────────────────────────────────
    ws0 = wb.active; ws0.title = "Executive Scorecard"; ws0.sheet_view.showGridLines = False
    title_row(ws0, f"PM EFFECTIVENESS  ·  {asset_label}  ·  EXECUTIVE SCORECARD", 1, "L")
    sub_row(ws0, f"Data: {date_range}  |  {len(df)} WOs  |  Ref date: {ref_date.strftime('%Y-%m-%d')}  |  GM09=Preventive  GM02=Breakdown  GM01/PM01=Follow-up Corrective", 2, "L")
    ws0.row_dimensions[3].height = 10

    # Scorecard header
    for col, txt, ec in [("B","Pillar","D"),("E","RAG",None),("F","Key Metric","J"),("K","Action Required","L")]:
        hdr(ws0, col, 4, txt, ec)
    ws0.row_dimensions[4].height = 18

    scorecard = [
        (PILLARS[0], p1['rag'], p1['key_metric'],
         "Review late PMs; investigate cycle time outliers" if p1['rag'] != "GREEN" else "Maintain current PM execution cadence"),
        (PILLARS[1], p2['rag'], p2['key_metric'],
         f"MTBF trending {p2['trend_dir']} — review PM frequency for top failure modes" if p2['rag'] != "GREEN" else "MTBF trend positive — continue monitoring"),
        (PILLARS[2], p3['rag'], p3['key_metric'],
         f"{p3['gaps']} failure clusters have no PM task — expand PM scope" if p3['gaps'] else "All failure clusters have PM coverage"),
        (PILLARS[3], p4['rag'], p4['key_metric'],
         f"Require technicians to book hours on GM02 BKD WOs — {p4['bkd_zero']} with 0hrs" if p4['rag'] != "GREEN" else "Hour capture is complete"),
        (PILLARS[4], p5['rag'], p5['key_metric'],
         f"Review {p5['age_91p']} WOs aged 90+ days; assess if still valid" if p5['rag'] != "GREEN" else "Backlog age is healthy"),
        (PILLARS[5], p6['rag'], p6['key_metric'],
         f"~{p6['est_hidden']:.0f} labor hours unaccounted on BKDs — cost of failure understated" if p6['rag'] != "GREEN" else "Labor cost capture is complete"),
        (PILLARS[6], p7['rag'], p7['key_metric'],
         f"MTBF {p7['mtbf_overall']}d — investigate top failure clusters to extend run intervals" if p7['rag'] != "GREEN" else f"MTBF is healthy at {p7['mtbf_overall']} days"),
    ]

    for i, (pillar, rag, metric, action) in enumerate(scorecard):
        r = i + 5; bg = rag_bg(rag)
        ws0.merge_cells(f"B{r}:D{r}"); ws0.merge_cells(f"F{r}:J{r}"); ws0.merge_cells(f"K{r}:L{r}")
        dc(ws0,"B",r,pillar,bg,bold=True,align="left")
        ws0[f"E{r}"].value = rag_label(rag)
        ws0[f"E{r}"].font = Font(name="Arial",bold=True,size=9,color=rag_fg(rag))
        ws0[f"E{r}"].fill = wfill(bg); ws0[f"E{r}"].alignment = cal(); ws0[f"E{r}"].border = bdr()
        dc(ws0,"F",r,metric,bg,align="left"); dc(ws0,"K",r,action,bg,align="left",wrap=True)
        ws0.row_dimensions[r].height = 26

    ws0.row_dimensions[13].height = 10
    sec(ws0,"OVERALL FINDINGS",13,"L")
    total_bkds = len(bkd)
    red_count  = sum(1 for _,rag,*_ in scorecard if rag=="RED")
    amb_count  = sum(1 for _,rag,*_ in scorecard if rag=="AMBER")
    findings = [
        f"{red_count} pillar(s) RED / {amb_count} AMBER — see detail tabs for corrective actions.",
        f"PM compliance: {p1['teco_pct']:.0f}% TECO rate across {len(pm)} PM WOs. Schedule adherence: {p1['adh_pct']:.0f}%.",
        f"MTBF: {p7['mtbf_overall']} days overall (H1 2026). Trend: {p2['trend_dir']}.",
        f"{p3['gaps']} of {len(CLUSTERS)} failure clusters have no corresponding PM task.",
        f"⚠  {p4['bkd_zero']}/{total_bkds} breakdown WOs carry zero actual hours — failure labor cost is largely invisible.",
        f"Backlog: {p5['total_open']} open WOs, avg age {p5['avg_age']:.0f} days, {p5['age_91p']} aged 90+ days.",
    ]
    for j, txt in enumerate(findings):
        r = 14 + j
        fg = rag_fg("RED") if "⚠" in txt else "1F2D3D"
        bg = RAG_RED_BG if "⚠" in txt else (RAG_RED_BG if j==0 and red_count>0 else LIGHT_ROW)
        ws0.merge_cells(f"B{r}:L{r}")
        dc(ws0,"B",r,txt,bg,align="left",wrap=True); ws0.row_dimensions[r].height = 22
    widths(ws0,{"A":2,"B":8,"C":8,"D":20,"E":12,"F":8,"G":8,"H":8,"I":8,"J":8,"K":24,"L":2})

    # ── SHEET 1: Pillar 1 — Compliance & Scheduling ──────────────────────────
    ws1 = wb.create_sheet("1 · Compliance"); ws1.sheet_view.showGridLines = False
    title_row(ws1, f"PILLAR 1: COMPLIANCE & SCHEDULING  ·  {asset_label}", 1, "L")
    sub_row(ws1, "PM TECO rate, schedule adherence by task, and WO cycle time (Created → Start)", 2, "L")
    ws1.row_dimensions[3].height = 10

    # KPI row
    kpis = [
        ("B","PM TECO Rate",f"{p1['teco_pct']:.1f}%",rag_bg("GREEN" if p1['teco_pct']>=90 else "AMBER" if p1['teco_pct']>=75 else "RED")),
        ("F","Schedule Adherence",f"{p1['adh_pct']:.1f}%",rag_bg("GREEN" if p1['adh_pct']>=80 else "AMBER" if p1['adh_pct']>=65 else "RED")),
        ("J","Avg Cycle Time",f"{p1['avg_cycle']:.1f} days",LIGHT_ROW),
    ]
    for col, lbl, val, bg in kpis:
        ec = chr(ord(col)+2)
        ws1.merge_cells(f"{col}4:{ec}4"); ws1[f"{col}4"].value = lbl
        ws1[f"{col}4"].font = Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws1[f"{col}4"].fill = wfill(WHITE); ws1[f"{col}4"].alignment = cal("left")
        ws1.merge_cells(f"{col}5:{ec}5"); ws1[f"{col}5"].value = val
        ws1[f"{col}5"].font = Font(name="Arial",bold=True,size=16,color="1F2D3D")
        ws1[f"{col}5"].fill = wfill(bg); ws1[f"{col}5"].alignment = cal(); ws1[f"{col}5"].border = bdr()
    ws1.row_dimensions[4].height=16; ws1.row_dimensions[5].height=30; ws1.row_dimensions[6].height=10

    sec(ws1,"SCHEDULE ADHERENCE BY PM TASK  (H1 2026)",7,"L")
    for col,txt,ec in [("B","PM Task","D"),("E","Expected","F"),("G","Intervals","H"),
                        ("I","Avg Gap","J"),("K","Avg Variance","L"),("M","On-Time %",None)]:
        hdr(ws1,col,8,txt,ec)
    ws1.row_dimensions[8].height = 18
    if len(p1['by_type']):
        for i, row in p1['by_type'].sort_values('OT_Pct',ascending=True).iterrows():
            r  = 9 + list(p1['by_type'].sort_values('OT_Pct',ascending=True).index).index(i)
            ot = row['OT_Pct']
            bg = GREEN_FLG if ot==100 else (LIGHT_ROW if ot>=90 else (ORANGE if ot>=67 else RED_FLG))
            var_s = f"+{row['Avg_Var']:.1f}d" if row['Avg_Var']>0 else f"{row['Avg_Var']:.1f}d"
            for mc in [f"B{r}:D{r}",f"E{r}:F{r}",f"G{r}:H{r}",f"I{r}:J{r}",f"K{r}:L{r}"]:
                ws1.merge_cells(mc)
            dc(ws1,"B",r,row['PM_Description'],bg,bold=True,align="left")
            dc(ws1,"E",r,f"{int(row['Expected_Days'])}d ({int(row['Expected_Days']//7)}WK)",bg)
            dc(ws1,"G",r,int(row['Intervals']),bg)
            dc(ws1,"I",r,f"{row['Avg_Gap']:.1f}d",bg)
            dc(ws1,"K",r,var_s,bg)
            dc(ws1,"M",r,f"{int(ot)}%",bg,bold=True); ws1.row_dimensions[r].height=18

    row_ct = 9 + (len(p1['by_type']) if len(p1['by_type']) else 1) + 1
    ws1.row_dimensions[row_ct].height = 10
    sec(ws1,"WO CYCLE TIME  (Created → Basic Start Date)",row_ct+1,"L")
    for col,txt,ec in [("B","Bucket",None),("D","Count",None),("F","% of PM WOs","H")]:
        hdr(ws1,col,row_ct+2,txt,ec)
    ws1.row_dimensions[row_ct+2].height=18
    ct_total = p1['ct_lt7'] + p1['ct_7_30'] + p1['ct_30p']
    for i,(lbl,ct) in enumerate([("≤ 7 days",p1['ct_lt7']),("8 – 30 days",p1['ct_7_30']),("31+ days",p1['ct_30p'])]):
        r = row_ct+3+i; bg = GREEN_FLG if i==0 else (LIGHT_ROW if i==1 else RED_FLG)
        ws1.merge_cells(f"B{r}:C{r}"); ws1.merge_cells(f"D{r}:E{r}"); ws1.merge_cells(f"F{r}:H{r}")
        dc(ws1,"B",r,lbl,bg,bold=True); dc(ws1,"D",r,ct,bg)
        dc(ws1,"F",r,f"{100*ct/ct_total:.0f}%" if ct_total else "—",bg)
        ws1.row_dimensions[r].height=18
    widths(ws1,{"A":2,"B":8,"C":8,"D":20,"E":4,"F":12,"G":4,"H":12,"I":12,"J":4,"K":16,"L":4,"M":12})

    # ── SHEET 2: Pillar 2 — Failure Prevention ───────────────────────────────
    ws2 = wb.create_sheet("2 · Failure Prevention"); ws2.sheet_view.showGridLines = False
    title_row(ws2,f"PILLAR 2: FAILURE PREVENTION  ·  {asset_label}",1,"L")
    sub_row(ws2,"PM→Breakdown timing, MTBF by month with 3-month rolling average and trend direction",2,"L")
    ws2.row_dimensions[3].height=10

    for col,lbl,val,bg in [
        ("B","Overall MTBF",f"{p2['overall_mtbf']} days",rag_bg(p7['rag'])),
        ("F","H1 2026 Trend",p2['trend_dir'],rag_bg(p2['trend_clr'])),
        ("J","PM→BKD Pairs",str(len(p2['pm_bkd'])),LIGHT_ROW),
    ]:
        ec = chr(ord(col)+2)
        ws2.merge_cells(f"{col}4:{ec}4"); ws2[f"{col}4"].value=lbl
        ws2[f"{col}4"].font=Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws2[f"{col}4"].fill=wfill(WHITE); ws2[f"{col}4"].alignment=cal("left")
        ws2.merge_cells(f"{col}5:{ec}5"); ws2[f"{col}5"].value=val
        ws2[f"{col}5"].font=Font(name="Arial",bold=True,size=14,color="1F2D3D")
        ws2[f"{col}5"].fill=wfill(bg); ws2[f"{col}5"].alignment=cal(); ws2[f"{col}5"].border=bdr()
    ws2.row_dimensions[4].height=16; ws2.row_dimensions[5].height=30; ws2.row_dimensions[6].height=10

    sec(ws2,"MTBF BY MONTH WITH ROLLING 3-MONTH AVERAGE",7,"L")
    for col,txt,ec in [("B","Month",None),("D","BKDs",None),("F","MTBF (30d/BKD)",None),
                        ("H","MTBF Actual",None),("J","Rolling 3mo",None),("L","Trend","L")]:
        hdr(ws2,col,8,txt,ec)
    ws2.row_dimensions[8].height=18
    prior_val = None
    for i,(mo,bkd_c,m30,ma,r3) in enumerate(zip(
            p2['months_str'],p2['bkd_mo'],p2['mtbf_30'],p2['mtbf_actual'],p2['rolling_3'])):
        r = i+9; bg = LIGHT_ROW if i%2==0 else WHITE
        if prior_val and ma:
            delta = ma-prior_val
            trend = f"↑ +{delta:.1f}d" if delta>1 else (f"↓ {delta:.1f}d" if delta<-1 else "→")
            tbg   = GREEN_FLG if delta>1 else (RED_FLG if delta<-1 else LIGHT_ROW)
        else:
            trend = "—"; tbg = bg
        ws2.merge_cells(f"B{r}:C{r}"); ws2.merge_cells(f"D{r}:E{r}"); ws2.merge_cells(f"F{r}:G{r}")
        ws2.merge_cells(f"H{r}:I{r}"); ws2.merge_cells(f"J{r}:K{r}"); ws2.merge_cells(f"L{r}:L{r}")
        dc(ws2,"B",r,mo+" 2026",bg,bold=True); dc(ws2,"D",r,bkd_c,bg)
        dc(ws2,"F",r,f"{m30}d" if m30 else "—",bg)
        dc(ws2,"H",r,f"{ma}d" if ma else "—",bg)
        dc(ws2,"J",r,f"{r3}d" if r3 else "—",bg)
        dc(ws2,"L",r,trend,tbg,bold=True); ws2.row_dimensions[r].height=18
        if ma: prior_val = ma

    # MTBF chart using hidden columns
    for i,(mo,m30,r3) in enumerate(zip(p2['months_str'],p2['mtbf_30'],p2['rolling_3'])):
        ws2[f"O{9+i}"].value = mo; ws2[f"P{9+i}"].value = m30; ws2[f"Q{9+i}"].value = r3
    ws2["P8"].value = "MTBF (30d)"; ws2["Q8"].value = "Rolling 3mo Avg"

    ws2.row_dimensions[17].height=10
    sec(ws2,"MTBF TREND  (H1 2026)",17,"L",HDR_BG)
    lc = LineChart(); lc.title="MTBF Trend with 3-Month Rolling Avg"; lc.style=10; lc.width=22; lc.height=12
    lc.y_axis.title="Days Between Breakdowns"
    lc.add_data(Reference(ws2,min_col=16,min_row=8,max_row=14),titles_from_data=True)
    lc.add_data(Reference(ws2,min_col=17,min_row=8,max_row=14),titles_from_data=True)
    lc.set_categories(Reference(ws2,min_col=15,min_row=9,max_row=14))
    lc.series[0].graphicalProperties.line.solidFill="2E75B6"; lc.series[0].graphicalProperties.line.width=28575
    lc.series[0].marker.symbol="circle"; lc.series[0].marker.size=7
    lc.series[1].graphicalProperties.line.solidFill="ED7D31"; lc.series[1].graphicalProperties.line.width=19050
    lc.series[1].graphicalProperties.line.dashDot="dash"
    ws2.add_chart(lc,"B18")
    for col in ["O","P","Q"]: ws2.column_dimensions[col].hidden = True

    # PM→BKD summary
    r_pb = 33; ws2.row_dimensions[r_pb].height=10
    sec(ws2,"PM → BREAKDOWN TIMING SUMMARY",r_pb+1,"L",HDR_BG)
    cat_clr = {"Same day": GREEN_FLG,"1-3 days": RED_FLG,"4-7 days": ORANGE,"8+ days": LIGHT_ROW}
    if len(p2['pm_bkd']):
        vc2 = p2['pm_bkd']['Category'].value_counts()
        for col,txt,ec in [("B","Timing Bucket",None),("E","Count",None),("G","% of BKDs",None),
                            ("I","Interpretation","L")]:
            hdr(ws2,col,r_pb+2,txt,ec)
        ws2.row_dimensions[r_pb+2].height=18
        buckets = [
            ("Same day","PM likely surfaced failure during inspection"),
            ("1-3 days","Immediate failure after PM — latent defect or component disturbed"),
            ("4-7 days","Short-interval failure — PM may not address root cause"),
            ("8+ days","Failure outside normal PM interval"),
        ]
        for i,(cat,interp) in enumerate(buckets):
            r = r_pb+3+i; ct = vc2.get(cat,0); bg = cat_clr.get(cat,LIGHT_ROW)
            ws2.merge_cells(f"B{r}:D{r}"); ws2.merge_cells(f"E{r}:F{r}"); ws2.merge_cells(f"G{r}:H{r}"); ws2.merge_cells(f"I{r}:L{r}")
            dc(ws2,"B",r,cat,bg,bold=True); dc(ws2,"E",r,ct,bg)
            dc(ws2,"G",r,f"{100*ct/len(p2['pm_bkd']):.0f}%",bg)
            dc(ws2,"I",r,interp,bg,align="left",wrap=True); ws2.row_dimensions[r].height=22
    widths(ws2,{"A":2,"B":8,"C":4,"D":8,"E":8,"F":4,"G":14,"H":4,"I":14,"J":10,"K":4,"L":12})

    # ── SHEET 3: Pillar 3 — Failure Mode Coverage ─────────────────────────────
    ws3 = wb.create_sheet("3 · Coverage Gap"); ws3.sheet_view.showGridLines = False
    title_row(ws3,f"PILLAR 3: FAILURE MODE COVERAGE  ·  {asset_label}",1,"L")
    sub_row(ws3,"Each breakdown cluster checked against PM task descriptions. GAP = breakdown cluster exists but no PM task addresses it.",2,"L")
    ws3.row_dimensions[3].height=10

    for col,lbl,val,bg in [
        ("B","Coverage Gaps",str(p3['gaps']),rag_bg("GREEN" if p3['gaps']==0 else "RED")),
        ("F","High Priority Gaps",str(p3['high_gaps']),rag_bg("GREEN" if p3['high_gaps']==0 else "RED")),
        ("J","Clusters Analyzed",str(len(CLUSTERS)),LIGHT_ROW),
    ]:
        ec = chr(ord(col)+2)
        ws3.merge_cells(f"{col}4:{ec}4"); ws3[f"{col}4"].value=lbl
        ws3[f"{col}4"].font=Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws3[f"{col}4"].fill=wfill(WHITE); ws3[f"{col}4"].alignment=cal("left")
        ws3.merge_cells(f"{col}5:{ec}5"); ws3[f"{col}5"].value=val
        ws3[f"{col}5"].font=Font(name="Arial",bold=True,size=16,color="1F2D3D")
        ws3[f"{col}5"].fill=wfill(bg); ws3[f"{col}5"].alignment=cal(); ws3[f"{col}5"].border=bdr()
    ws3.row_dimensions[4].height=16; ws3.row_dimensions[5].height=30; ws3.row_dimensions[6].height=10

    sec(ws3,"FAILURE CLUSTER — PM COVERAGE CHECK",7,"L")
    for col,txt,ec in [("B","Failure Cluster","C"),("D","BKD WOs","E"),("F","PM Covered?","G"),
                        ("H","Status","I"),("J","Priority","K"),("L","Recommended Action","L")]:
        hdr(ws3,col,8,txt,ec)
    ws3.row_dimensions[8].height=18
    cov_sorted = p3['cov_df'].sort_values('BKD_Count',ascending=False)
    for i,(_,row) in enumerate(cov_sorted.iterrows()):
        r = i+9
        bg = RED_FLG if row['Gap'] else (GREEN_FLG if row['BKD_Count']>0 else LIGHT_ROW)
        action = (f"Add PM task targeting {row['Cluster'].lower()} components" if row['Gap']
                  else ("PM coverage confirmed" if row['BKD_Count']>0 else "No breakdowns — continue monitoring"))
        for mc in [f"B{r}:C{r}",f"D{r}:E{r}",f"F{r}:G{r}",f"H{r}:I{r}",f"J{r}:K{r}"]:
            ws3.merge_cells(mc)
        dc(ws3,"B",r,row['Cluster'],bg,bold=True,align="left")
        dc(ws3,"D",r,row['BKD_Count'],bg)
        dc(ws3,"F",r,"✓ Yes" if row['PM_Covered'] else "✗ No",
           GREEN_FLG if row['PM_Covered'] else RED_FLG,bold=True)
        dc(ws3,"H",r,row['Status'],bg)
        dc(ws3,"J",r,row['Priority'],RED_FLG if row['Priority']=="HIGH" else (ORANGE if row['Priority']=="MED" else GREEN_FLG),bold=True)
        dc(ws3,"L",r,action,bg,align="left",wrap=True); ws3.row_dimensions[r].height=22
    widths(ws3,{"A":2,"B":8,"C":18,"D":8,"E":6,"F":8,"G":6,"H":18,"I":4,"J":8,"K":6,"L":28})

    # ── SHEET 4: Pillar 4 — Work Order Quality ─────────────────────────────
    ws4 = wb.create_sheet("4 · WO Quality"); ws4.sheet_view.showGridLines = False
    title_row(ws4,f"PILLAR 4: WORK ORDER QUALITY  ·  {asset_label}",1,"L")
    sub_row(ws4,"Zero-hour analysis by WO type. Planned vs. actual hours variance (if planned hours available in export).",2,"L")
    ws4.row_dimensions[3].height=10

    for col,lbl,val,bg in [
        ("B","PM Zero-Hour %",f"{p4['pm_zero_pct']:.0f}%",rag_bg("GREEN" if p4['pm_zero_pct']==0 else "AMBER")),
        ("F","BKD Zero-Hour %",f"{p4['bkd_zero_pct']:.0f}%",rag_bg("GREEN" if p4['bkd_zero_pct']<20 else "AMBER" if p4['bkd_zero_pct']<60 else "RED")),
        ("J","Follow-up Corr Zero-Hr %",f"{p4['fuc_zero_pct']:.0f}%",ORANGE),
    ]:
        ec = chr(ord(col)+2)
        ws4.merge_cells(f"{col}4:{ec}4"); ws4[f"{col}4"].value=lbl
        ws4[f"{col}4"].font=Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws4[f"{col}4"].fill=wfill(WHITE); ws4[f"{col}4"].alignment=cal("left")
        ws4.merge_cells(f"{col}5:{ec}5"); ws4[f"{col}5"].value=val
        ws4[f"{col}5"].font=Font(name="Arial",bold=True,size=16,color="1F2D3D")
        ws4[f"{col}5"].fill=wfill(bg); ws4[f"{col}5"].alignment=cal(); ws4[f"{col}5"].border=bdr()
    ws4.row_dimensions[4].height=16; ws4.row_dimensions[5].height=30; ws4.row_dimensions[6].height=10

    sec(ws4,"ZERO-HOUR WO SUMMARY BY TYPE",7,"L")
    for col,txt,ec in [("B","WO Type","D"),("E","Total WOs","F"),("G","Zero-Hour","H"),("I","Zero-Hr %","J"),
                        ("K","Assessment","L")]:
        hdr(ws4,col,8,txt,ec)
    ws4.row_dimensions[8].height=18
    zh_rows = [
        ("PM (GM09)",      len(pm),  p4['pm_zero'],  p4['pm_zero_pct'],
         "✓ Good — all PMs have hours" if p4['pm_zero']==0 else f"⚠ {p4['pm_zero']} PMs missing hours",
         GREEN_FLG if p4['pm_zero']==0 else YELLOW),
        ("Breakdown (GM02)",len(bkd), p4['bkd_zero'], p4['bkd_zero_pct'],
         "⚠ Labor not captured — true failure cost unknown", YELLOW),
        ("Follow-up (GM01/PM01)",len(fuc),p4['fuc_zero'],p4['fuc_zero_pct'],
         "⚠ Separate parts-only from labor WOs; audit non-parts zeros", ORANGE),
    ]
    for i,(wt,tot,zh,zhp,assess,bg) in enumerate(zh_rows):
        r = i+9
        for mc in [f"B{r}:D{r}",f"E{r}:F{r}",f"G{r}:H{r}",f"I{r}:J{r}",f"K{r}:L{r}"]:
            ws4.merge_cells(mc)
        dc(ws4,"B",r,wt,bg,bold=True,align="left"); dc(ws4,"E",r,tot,bg)
        dc(ws4,"G",r,zh,bg); dc(ws4,"I",r,f"{zhp:.1f}%",bg,bold=True)
        dc(ws4,"K",r,assess,bg,align="left",wrap=True); ws4.row_dimensions[r].height=28

    if p4['has_planned']:
        ws4.row_dimensions[13].height=10
        sec(ws4,"PLANNED vs. ACTUAL HOURS — PM WOs",14,"L")
        for col,txt,ec in [("B","Metric","D"),("E","Value","G")]:
            hdr(ws4,col,15,txt,ec)
        ws4.row_dimensions[15].height=18
        pva_rows = [
            ("PM WOs compared",       len(p4['pva']),  LIGHT_ROW),
            ("Avg variance %",         f"{p4['avg_var_pct']:+.1f}%", ORANGE if abs(p4['avg_var_pct'])>20 else LIGHT_ROW),
            ("Over estimate (actual > planned)", p4['over_cnt'],  RED_FLG if p4['over_cnt']>3 else LIGHT_ROW),
            ("Under estimate (actual < planned)",p4['under_cnt'], LIGHT_ROW),
            ("On estimate (actual = planned)",   p4['on_cnt'],    GREEN_FLG),
        ]
        for i,(lbl,val,bg) in enumerate(pva_rows):
            r = 16+i
            ws4.merge_cells(f"B{r}:D{r}"); ws4.merge_cells(f"E{r}:G{r}")
            dc(ws4,"B",r,lbl,bg,align="left"); dc(ws4,"E",r,str(val),bg,bold=True)
            ws4.row_dimensions[r].height=18
    else:
        ws4.row_dimensions[13].height=10
        sec(ws4,"PLANNED vs. ACTUAL HOURS",14,"L",HDR_BG)
        ws4.merge_cells("B15:L15")
        dc(ws4,"B",15,
           "Planned Hours column not found in Celonis export. Add 'Planned Hours' to the export to enable this analysis.",
           YELLOW,align="left",wrap=True); ws4.row_dimensions[15].height=28
    widths(ws4,{"A":2,"B":8,"C":8,"D":16,"E":8,"F":8,"G":10,"H":4,"I":10,"J":4,"K":26,"L":2})

    # ── SHEET 5: Pillar 5 — Backlog Health ───────────────────────────────────
    ws5 = wb.create_sheet("5 · Backlog Health"); ws5.sheet_view.showGridLines = False
    title_row(ws5,f"PILLAR 5: BACKLOG HEALTH  ·  {asset_label}",1,"L")
    sub_row(ws5,f"Open WOs (no TECO/CLSD) aged from Created On to ref date {ref_date.strftime('%Y-%m-%d')}. Monthly new vs. closed trend.",2,"L")
    ws5.row_dimensions[3].height=10

    for col,lbl,val,bg in [
        ("B","Total Open WOs",str(p5['total_open']),rag_bg("GREEN" if p5['total_open']<15 else "AMBER" if p5['total_open']<30 else "RED")),
        ("F","Aged 90+ Days",str(p5['age_91p']),rag_bg("GREEN" if p5['age_91p']==0 else "AMBER" if p5['age_91p']<=5 else "RED")),
        ("J","Avg Age",f"{p5['avg_age']:.0f} days",LIGHT_ROW),
    ]:
        ec = chr(ord(col)+2)
        ws5.merge_cells(f"{col}4:{ec}4"); ws5[f"{col}4"].value=lbl
        ws5[f"{col}4"].font=Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws5[f"{col}4"].fill=wfill(WHITE); ws5[f"{col}4"].alignment=cal("left")
        ws5.merge_cells(f"{col}5:{ec}5"); ws5[f"{col}5"].value=val
        ws5[f"{col}5"].font=Font(name="Arial",bold=True,size=16,color="1F2D3D")
        ws5[f"{col}5"].fill=wfill(bg); ws5[f"{col}5"].alignment=cal(); ws5[f"{col}5"].border=bdr()
    ws5.row_dimensions[4].height=16; ws5.row_dimensions[5].height=30; ws5.row_dimensions[6].height=10

    sec(ws5,"BACKLOG AGING BUCKETS",7,"L")
    for col,txt,ec in [("B","Age Bucket","D"),("E","Count","F"),("G","% of Open","H"),("I","Assessment","L")]:
        hdr(ws5,col,8,txt,ec)
    ws5.row_dimensions[8].height=18
    aging = [
        ("0 – 30 days",  p5['age_0_30'],  GREEN_FLG, "Current — monitor"),
        ("31 – 90 days", p5['age_31_90'], ORANGE,    "Aging — review priority and assign completion date"),
        ("91+ days",     p5['age_91p'],   RED_FLG,   "Stale — assess validity; close or escalate"),
    ]
    for i,(lbl,ct,bg,assess) in enumerate(aging):
        r = i+9
        ws5.merge_cells(f"B{r}:D{r}"); ws5.merge_cells(f"E{r}:F{r}"); ws5.merge_cells(f"G{r}:H{r}"); ws5.merge_cells(f"I{r}:L{r}")
        dc(ws5,"B",r,lbl,bg,bold=True); dc(ws5,"E",r,ct,bg)
        dc(ws5,"G",r,f"{100*ct/p5['total_open']:.0f}%" if p5['total_open'] else "—",bg)
        dc(ws5,"I",r,assess,bg,align="left",wrap=True); ws5.row_dimensions[r].height=22

    sec(ws5,"OPEN WOs BY TYPE",13,"L",HDR_BG)
    for col,txt,ec in [("B","Order Type","D"),("E","Count","F"),("G","Avg Age (d)","H"),("I","Max Age (d)","J")]:
        hdr(ws5,col,14,txt,ec)
    ws5.row_dimensions[14].height=18
    for i,(_,row) in enumerate(p5['by_type'].sort_values('Count',ascending=False).iterrows()):
        r = 15+i; bg = LIGHT_ROW if i%2==0 else WHITE
        ws5.merge_cells(f"B{r}:D{r}"); ws5.merge_cells(f"E{r}:F{r}"); ws5.merge_cells(f"G{r}:H{r}"); ws5.merge_cells(f"I{r}:J{r}")
        dc(ws5,"B",r,row['OrderType'],bg,bold=True); dc(ws5,"E",r,int(row['Count']),bg)
        dc(ws5,"G",r,f"{row['Avg_Age']:.0f}",bg); dc(ws5,"I",r,int(row['Max_Age']),bg)
        ws5.row_dimensions[r].height=18

    r_ch = 15 + len(p5['by_type']) + 2
    ws5.row_dimensions[r_ch].height=10
    sec(ws5,"MONTHLY NEW vs. CLOSED WOs (H1 2026)",r_ch+1,"L",HDR_BG)
    for i,(mo,nw,cl,nt) in enumerate(zip(p5['months_str'],p5['new_mo'],p5['closed_mo'],p5['net_mo'])):
        ws5[f"O{r_ch+2+i}"].value = mo
        ws5[f"P{r_ch+2+i}"].value = nw
        ws5[f"Q{r_ch+2+i}"].value = cl
    ws5[f"O{r_ch+1}"].value = "Month"
    ws5[f"P{r_ch+1}"].value = "New WOs"
    ws5[f"Q{r_ch+1}"].value = "Closed (TECO)"
    bc5 = BarChart(); bc5.type="col"; bc5.grouping="clustered"
    bc5.title="Monthly New vs. TECO'd WOs (H1 2026)"; bc5.style=10; bc5.width=22; bc5.height=12
    bc5.add_data(Reference(ws5,min_col=16,min_row=r_ch+1,max_row=r_ch+7),titles_from_data=True)
    bc5.add_data(Reference(ws5,min_col=17,min_row=r_ch+1,max_row=r_ch+7),titles_from_data=True)
    bc5.set_categories(Reference(ws5,min_col=15,min_row=r_ch+2,max_row=r_ch+7))
    bc5.series[0].graphicalProperties.solidFill="2E75B6"
    bc5.series[1].graphicalProperties.solidFill="375623"
    ws5.add_chart(bc5,f"B{r_ch+2}")
    for col in ["O","P","Q"]: ws5.column_dimensions[col].hidden = True
    widths(ws5,{"A":2,"B":8,"C":8,"D":16,"E":8,"F":6,"G":12,"H":4,"I":12,"J":4,"K":8,"L":2})

    # ── SHEET 6: Pillar 6 — Cost Visibility ──────────────────────────────────
    ws6 = wb.create_sheet("6 · Cost Visibility"); ws6.sheet_view.showGridLines = False
    title_row(ws6,f"PILLAR 6: COST VISIBILITY  ·  {asset_label}",1,"L")
    sub_row(ws6,"Labor hour capture by WO type. Estimated hidden cost from unbooked breakdown hours. PM vs corrective labor split.",2,"L")
    ws6.row_dimensions[3].height=10

    for col,lbl,val,bg in [
        ("B","Total Captured Hours",f"{p6['all_hrs']:.1f}",LIGHT_ROW),
        ("F","BKD Hours Captured",f"{100-p6['bkd_zero_pct']:.0f}%",rag_bg("GREEN" if p6['bkd_zero_pct']<20 else "AMBER" if p6['bkd_zero_pct']<60 else "RED")),
        ("J","Est. Hidden BKD Hours",f"~{p6['est_hidden']:.0f}",rag_bg("RED" if p6['est_hidden']>20 else "AMBER")),
    ]:
        ec = chr(ord(col)+2)
        ws6.merge_cells(f"{col}4:{ec}4"); ws6[f"{col}4"].value=lbl
        ws6[f"{col}4"].font=Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws6[f"{col}4"].fill=wfill(WHITE); ws6[f"{col}4"].alignment=cal("left")
        ws6.merge_cells(f"{col}5:{ec}5"); ws6[f"{col}5"].value=val
        ws6[f"{col}5"].font=Font(name="Arial",bold=True,size=16,color="1F2D3D")
        ws6[f"{col}5"].fill=wfill(bg); ws6[f"{col}5"].alignment=cal(); ws6[f"{col}5"].border=bdr()
    ws6.row_dimensions[4].height=16; ws6.row_dimensions[5].height=30; ws6.row_dimensions[6].height=10

    sec(ws6,"LABOR HOUR CAPTURE BY WO TYPE",7,"L")
    for col,txt,ec in [("B","WO Type","D"),("E","WO Count","F"),("G","Actual Hours","H"),
                        ("I","Zero-Hr WOs","J"),("K","Capture Rate","L")]:
        hdr(ws6,col,8,txt,ec)
    ws6.row_dimensions[8].height=18
    cost_rows = [
        ("PM (GM09)",             len(pm),  p6['pm_hrs'],  p4['pm_zero'],  100-p4['pm_zero_pct'],  GREEN_FLG),
        ("Breakdown (GM02)",      len(bkd), p6['bkd_hrs'], p4['bkd_zero'], 100-p4['bkd_zero_pct'], YELLOW if p4['bkd_zero_pct']>50 else ORANGE),
        ("Follow-up (GM01/PM01)", len(fuc), p6['fuc_hrs'], p4['fuc_zero'], 100-p4['fuc_zero_pct'], LIGHT_ROW),
    ]
    for i,(wt,cnt,hrs,zh,cap,bg) in enumerate(cost_rows):
        r = i+9
        for mc in [f"B{r}:D{r}",f"E{r}:F{r}",f"G{r}:H{r}",f"I{r}:J{r}",f"K{r}:L{r}"]:
            ws6.merge_cells(mc)
        dc(ws6,"B",r,wt,bg,bold=True,align="left"); dc(ws6,"E",r,cnt,bg)
        dc(ws6,"G",r,f"{hrs:.1f}",bg); dc(ws6,"I",r,zh,bg)
        dc(ws6,"K",r,f"{cap:.0f}%",bg,bold=True); ws6.row_dimensions[r].height=22

    ws6.row_dimensions[13].height=10
    sec(ws6,"COST VISIBILITY GAP ANALYSIS",14,"L",HDR_BG)
    gap_rows = [
        ("BKD WOs with zero hours",          f"{p4['bkd_zero']} of {len(bkd)}",   RED_FLG),
        ("Avg hours per captured BKD WO",    f"{p6['avg_bkd_hr']:.2f} hrs",        LIGHT_ROW),
        ("Estimated hidden breakdown hours", f"~{p6['est_hidden']:.0f} hrs",       YELLOW),
        ("⚠ Failure cost cannot be assessed until BKD WOs carry actual hours",
         "Corrective action required", RED_FLG),
    ]
    for col,txt,ec in [("B","Item","H"),("I","Value","L")]:
        hdr(ws6,col,15,txt,ec)
    ws6.row_dimensions[15].height=18
    for i,(lbl,val,bg) in enumerate(gap_rows):
        r = 16+i
        ws6.merge_cells(f"B{r}:H{r}"); ws6.merge_cells(f"I{r}:L{r}")
        dc(ws6,"B",r,lbl,bg,align="left",wrap=True,bold=("⚠" in lbl))
        dc(ws6,"I",r,val,bg,bold=True); ws6.row_dimensions[r].height=22

    ws6.row_dimensions[21].height=10
    sec(ws6,"LABOR MIX: PLANNED vs. CORRECTIVE",22,"L",HDR_BG)
    for col,txt,ec in [("B","Category","D"),("E","Hours","G"),("H","% of Total","J")]:
        hdr(ws6,col,23,txt,ec)
    ws6.row_dimensions[23].height=18
    mix_rows = [
        ("Planned (PM)",      p6['pm_hrs'],   p6['planned_pct'], GREEN_FLG),
        ("Corrective (BKD+FUC)",p6['bkd_hrs']+p6['fuc_hrs'], p6['corr_pct'], RED_FLG if p6['corr_pct']>60 else ORANGE),
        ("Total",             p6['all_hrs'],  100.0, MID_BG),
    ]
    for i,(lbl,hrs,pct,bg) in enumerate(mix_rows):
        r = 24+i
        ws6.merge_cells(f"B{r}:D{r}"); ws6.merge_cells(f"E{r}:G{r}"); ws6.merge_cells(f"H{r}:J{r}")
        dc(ws6,"B",r,lbl,bg,bold=True,align="left",color=WHITE if bg==MID_BG else "1F2D3D")
        dc(ws6,"E",r,f"{hrs:.1f}",bg,color=WHITE if bg==MID_BG else "1F2D3D")
        dc(ws6,"H",r,f"{pct:.0f}%",bg,bold=True,color=WHITE if bg==MID_BG else "1F2D3D")
        ws6.row_dimensions[r].height=20
    widths(ws6,{"A":2,"B":8,"C":8,"D":18,"E":8,"F":8,"G":8,"H":10,"I":4,"J":4,"K":14,"L":2})

    # ── SHEET 7: Pillar 7 — Equipment Reliability ─────────────────────────────
    ws7 = wb.create_sheet("7 · Reliability"); ws7.sheet_view.showGridLines = False
    title_row(ws7,f"PILLAR 7: EQUIPMENT RELIABILITY  ·  {asset_label}",1,"L")
    sub_row(ws7,"MTBF (Mean Time Between Failures) and MTTR (Mean Time to Repair) by month. H1 2026.",2,"L")
    ws7.row_dimensions[3].height=10

    for col,lbl,val,bg in [
        ("B","MTBF (Overall)",f"{p7['mtbf_overall']} days",rag_bg("GREEN" if p7['mtbf_overall']>7 else "AMBER" if p7['mtbf_overall']>3 else "RED")),
        ("F","MTTR",f"{p7['mttr']} hrs" if p7['mttr'] else "N/A",rag_bg("AMBER") if not p7['mttr'] else LIGHT_ROW),
        ("J","Total Breakdowns (H1)",str(sum(p7['bkd_mo'])),LIGHT_ROW),
    ]:
        ec = chr(ord(col)+2)
        ws7.merge_cells(f"{col}4:{ec}4"); ws7[f"{col}4"].value=lbl
        ws7[f"{col}4"].font=Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws7[f"{col}4"].fill=wfill(WHITE); ws7[f"{col}4"].alignment=cal("left")
        ws7.merge_cells(f"{col}5:{ec}5"); ws7[f"{col}5"].value=val
        ws7[f"{col}5"].font=Font(name="Arial",bold=True,size=16,color="1F2D3D")
        ws7[f"{col}5"].fill=wfill(bg); ws7[f"{col}5"].alignment=cal(); ws7[f"{col}5"].border=bdr()
    ws7.row_dimensions[4].height=16; ws7.row_dimensions[5].height=30; ws7.row_dimensions[6].height=10

    sec(ws7,"MONTHLY RELIABILITY METRICS  (H1 2026)",7,"L")
    for col,txt,ec in [("B","Month","C"),("D","BKDs","E"),("F","MTBF (30d/BKD)","G"),
                        ("H","MTTR (hrs)","I"),("J","Best/Worst","L")]:
        hdr(ws7,col,8,txt,ec)
    ws7.row_dimensions[8].height=18
    for i,(mo,bkd_c,m30,mt) in enumerate(zip(p7['months_str'],p7['bkd_mo'],
                                               [round(30/c,1) if c>0 else 0 for c in p7['bkd_mo']],
                                               p7['mttr_mo'])):
        r = i+9; bg = LIGHT_ROW if i%2==0 else WHITE
        flag = ""
        if bkd_c == min(c for c in p7['bkd_mo'] if c>0): flag = "Best month ↑"; fbg = GREEN_FLG
        elif bkd_c == max(p7['bkd_mo']): flag = "Worst month ↓"; fbg = RED_FLG
        else: fbg = bg
        for mc in [f"B{r}:C{r}",f"D{r}:E{r}",f"F{r}:G{r}",f"H{r}:I{r}",f"J{r}:L{r}"]:
            ws7.merge_cells(mc)
        dc(ws7,"B",r,mo+" 2026",bg,bold=True); dc(ws7,"D",r,bkd_c,bg)
        dc(ws7,"F",r,f"{m30}d" if m30 else "—",bg)
        dc(ws7,"H",r,f"{mt}h" if mt else "—",bg)
        dc(ws7,"J",r,flag,fbg,align="left"); ws7.row_dimensions[r].height=18

    ws7.row_dimensions[16].height=10
    sec(ws7,f"MTTR NOTE: {p7['mttr_note']}",16,"L",ORANGE if not p7['mttr'] else HDR_BG)

    # MTBF chart (hidden cols)
    for i,(mo,m30) in enumerate(zip(p7['months_str'],[round(30/c,1) if c>0 else 0 for c in p7['bkd_mo']])):
        ws7[f"O{9+i}"].value = mo; ws7[f"P{9+i}"].value = m30
    ws7["P8"].value = "MTBF (days)"
    ws7.row_dimensions[18].height=10
    sec(ws7,"MONTHLY MTBF CHART",18,"L",HDR_BG)
    lc7 = LineChart(); lc7.title="MTBF by Month (H1 2026)"; lc7.style=10; lc7.width=22; lc7.height=12
    lc7.y_axis.title="Days Between Failures"
    lc7.add_data(Reference(ws7,min_col=16,min_row=8,max_row=14),titles_from_data=True)
    lc7.set_categories(Reference(ws7,min_col=15,min_row=9,max_row=14))
    lc7.series[0].graphicalProperties.line.solidFill="2E75B6"; lc7.series[0].graphicalProperties.line.width=28575
    lc7.series[0].marker.symbol="circle"; lc7.series[0].marker.size=8
    ws7.add_chart(lc7,"B19")
    for col in ["O","P"]: ws7.column_dimensions[col].hidden = True
    widths(ws7,{"A":2,"B":8,"C":8,"D":10,"E":4,"F":14,"G":4,"H":12,"I":4,"J":14,"K":8,"L":4})

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ═════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ═════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="PM Effectiveness App 1.0", page_icon="🔧", layout="centered")

st.title("PM Effectiveness App 1.0")
st.caption("Upload a Celonis export. Get a 7-pillar analysis workbook.")

asset_label = st.text_input("Asset label", value="NEW MILFORD  ·  MBF01 | CARTONER")
uploaded    = st.file_uploader("Upload Celonis Export (.xlsx)", type=["xlsx"])

if uploaded:
    with st.spinner("Running analysis across 7 pillars…"):
        try:
            df, pm, bkd, fuc, ref_date = prep_data(uploaded)

            p1 = pillar1_compliance(pm, bkd, fuc, df)
            p2 = pillar2_failure_prevention(pm, bkd)
            p3 = pillar3_coverage(pm, bkd)
            p4 = pillar4_wo_quality(pm, bkd, fuc)
            p5 = pillar5_backlog(df, ref_date)
            p6 = pillar6_cost(pm, bkd, fuc, df)
            p7 = pillar7_reliability(bkd, df, ref_date)

            # Quick scorecard preview
            st.success("Analysis complete.")
            cols = st.columns(7)
            for col_ui, pname, p_data in zip(cols, PILLARS, [p1,p2,p3,p4,p5,p6,p7]):
                rag   = p_data['rag']
                emoji = "🟢" if rag=="GREEN" else ("🟡" if rag=="AMBER" else "🔴")
                col_ui.metric(pname.split(". ",1)[1], emoji)

            output = build_workbook(df, pm, bkd, fuc, ref_date, asset_label, [p1,p2,p3,p4,p5,p6,p7])
            from datetime import date
            fname  = f"PM_Effectiveness_{date.today().strftime('%Y%m%d')}.xlsx"
            st.download_button(
                label="⬇️ Download Analysis Workbook",
                data=output, file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Analysis failed: {e}")
            import traceback; st.code(traceback.format_exc())

with st.expander("Expected Celonis export columns"):
    st.markdown("""
- `WorkOrderNumber`, `Description`, `OrderType`
- `Work order classification for preventive and corrective processing metrics`
- `System Status`, `Created On`, `Basic Start Date`
- `Actual Hours`, `SystemCondition`
- `Planned Hours` *(optional — enables planned vs. actual variance analysis)*
""")
