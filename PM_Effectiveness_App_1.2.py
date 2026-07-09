import io, re, datetime
import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# ── palette ──────────────────────────────────────────────────────────────────
DARK_BG   = "1F2D3D"; MID_BG  = "2E4057"; HDR_BG  = "3A5068"
LIGHT_ROW = "EAF1F8"; WHITE   = "FFFFFF"; YELLOW  = "FFF2CC"
RED_FLG   = "FCE4D6"; GREEN_FLG = "E2EFDA"; ORANGE  = "FCE9D6"
RAG_RED   = "C00000"; RAG_AMBER = "ED7D31"; RAG_GREEN = "375623"
RAG_RED_BG   = "FCE4D6"; RAG_AMBER_BG = "FCE9D6"; RAG_GREEN_BG = "E2EFDA"

PILLARS = [
    "1. Compliance & Scheduling",
    "2. Failure Prevention",
    "3. Failure Mode Coverage",
    "4. Work Order Quality",
    "5. Backlog Health",
    "6. Cost Visibility",
    "7. Equipment Reliability",
]

# ── style helpers ─────────────────────────────────────────────────────────────
def wfill(h):   return PatternFill("solid", start_color=h, fgColor=h)
def cal(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def title_row(ws, text, row, end_col="J"):
    ec = end_col
    ws.merge_cells(f"B{row}:{ec}{row}")
    c = ws[f"B{row}"]; c.value = text
    c.font = Font(name="Arial", bold=True, color=WHITE, size=13)
    c.fill = wfill(DARK_BG); c.alignment = cal(); ws.row_dimensions[row].height = 26

def sub_row(ws, text, row, end_col="J", height=20):
    ws.merge_cells(f"B{row}:{end_col}{row}")
    c = ws[f"B{row}"]; c.value = text
    c.font = Font(name="Arial", italic=True, color="AAAAAA", size=9)
    c.fill = wfill(DARK_BG); c.alignment = cal(wrap=True); ws.row_dimensions[row].height = height

def sec(ws, text, row, end_col="J", bg=MID_BG):
    ws.merge_cells(f"B{row}:{end_col}{row}")
    c = ws[f"B{row}"]; c.value = text
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = wfill(bg); c.alignment = cal("left"); ws.row_dimensions[row].height = 18

def hdr(ws, col, row, text, end_col=None):
    ref = f"{col}{row}"
    if end_col: ws.merge_cells(f"{ref}:{end_col}{row}")
    ws[ref].value = text
    ws[ref].font = Font(name="Arial", bold=True, color=WHITE, size=9)
    ws[ref].fill = wfill(HDR_BG); ws[ref].alignment = cal(wrap=True); ws[ref].border = bdr()

def dc(ws, col, row, val, bg=WHITE, bold=False, align="center", wrap=False, color="1F2D3D"):
    ws[f"{col}{row}"].value = val
    ws[f"{col}{row}"].font = Font(name="Arial", size=9, color=color, bold=bold)
    ws[f"{col}{row}"].fill = wfill(bg)
    ws[f"{col}{row}"].alignment = cal(align, wrap=wrap)
    ws[f"{col}{row}"].border = bdr()

def widths(ws, m):
    for c, w in m.items(): ws.column_dimensions[c].width = w

def rag_bg(status):
    return {"GREEN": RAG_GREEN_BG, "AMBER": RAG_AMBER_BG, "RED": RAG_RED_BG}.get(status, LIGHT_ROW)

def rag_fg(status):
    return {"GREEN": RAG_GREEN, "AMBER": RAG_AMBER, "RED": RAG_RED}.get(status, "1F2D3D")

def rag_label(status):
    return {"GREEN": "✓ GOOD", "AMBER": "⚠ WATCH", "RED": "✗ ACTION"}.get(status, "—")

# ── helpers ───────────────────────────────────────────────────────────────────
def hf(s, f): return f in str(s).split() if pd.notna(s) else False

def pfreq(d):
    if pd.isna(d): return None
    m = re.search(r'(\d+)\s*WK', d, re.IGNORECASE)
    if m: return int(m.group(1)) * 7
    if re.search(r'5W', d, re.IGNORECASE): return 35
    return None

# ── dynamic month window ─────────────────────────────────────────────────────
def build_month_window(df):
    """Derive month buckets from Basic Start Date range (the execution window)."""
    all_dates = df['Basic Start Date'].dropna()
    if all_dates.empty:
        end   = pd.Timestamp.today()
        start = end - pd.DateOffset(months=5)
    else:
        start = all_dates.min()
        end   = all_dates.max()

    periods    = pd.period_range(start=start, end=end, freq='M')
    month_keys = [str(p) for p in periods]           # ['2025-01', ...]
    n          = len(month_keys)
    multi_year = periods[0].year != periods[-1].year

    # Display labels: include year suffix when data spans >1 year
    if multi_year or n > 12:
        months_str = [p.strftime("%b '%y") for p in periods]
    else:
        months_str = [p.strftime('%b') for p in periods]

    # Compact period label for chart/section titles
    if n <= 6 and not multi_year:
        h            = 'H1' if periods[-1].month <= 6 else 'H2'
        period_label = f"{h} {periods[0].year}"
    elif n <= 12 and not multi_year:
        period_label = str(periods[0].year)
    else:
        period_label = f"{periods[0].strftime('%b %Y')} – {periods[-1].strftime('%b %Y')}"

    last_period  = periods[-1]
    window_start = f"{month_keys[0]}-01"
    window_end   = f"{month_keys[-1]}-{last_period.days_in_month:02d}"

    return {
        'month_keys':   month_keys,
        'months_str':   months_str,
        'month_map':    {k: i for i, k in enumerate(month_keys)},
        'n':            n,
        'period_label': period_label,
        'window_start': window_start,
        'window_end':   window_end,
    }


# ── auto-cluster failure modes from breakdown descriptions ───────────────────
def build_clusters_from_bkd(bkd, n_clusters=10):
    """Derive failure clusters from BKD WO descriptions via term co-occurrence."""
    from collections import Counter
    NOISE = {
        'and','the','of','in','on','for','to','a','an','is','not','no','with',
        'or','replace','replacement','inspection','inspect','insp','service',
        'check','pm','wr','wk','repair','clean','cleaning','preventive',
        'maintenance','work','order','corrective','via','per','due','has',
        'was','are','were','been','does','did','have','had','will','at','by',
        'from','that','this','its','into','run','running','out','off','set',
        'up','down','over','under','left','right','new','old','found',
        'broken','failed','failure','issue','problem','faulty','worn',
        'unit','time','date','area',
    }
    desc_word_sets = []
    all_words      = []
    for desc in bkd['Description'].dropna():
        words = re.findall(r'[a-zA-Z]{3,}', desc.lower())
        words = [w for w in words if w not in NOISE]
        all_words.extend(words)
        desc_word_sets.append(set(words))

    if not all_words:
        return [("General Breakdown", ['breakdown', 'failure'])]

    freq       = Counter(all_words)
    candidates = [w for w, _ in freq.most_common(80) if freq[w] >= 2]

    clusters  = []
    used      = set()
    for seed in candidates:
        if seed in used:
            continue
        seed_docs = [ws for ws in desc_word_sets if seed in ws]
        if not seed_docs:
            continue
        co = [seed]
        for other in candidates:
            if other == seed or other in used:
                continue
            both = sum(1 for ws in desc_word_sets if seed in ws and other in ws)
            if both >= max(2, len(seed_docs) * 0.4):
                co.append(other)
        clusters.append((seed.replace('-',' ').title(), co))
        used.update(co)
        if len(clusters) >= n_clusters:
            break

    return clusters if clusters else [("General", ['breakdown'])]

# ═════════════════════════════════════════════════════════════════════════════
# DATA PREPARATION
# ═════════════════════════════════════════════════════════════════════════════
def prep_data(file_obj):
    df = pd.read_excel(file_obj)
    new_fmt = 'Maintenance Activity Type' in df.columns

    if new_fmt:
        # ── New Celonis format (Maintenance Activity Type + Priority) ─────────
        df.columns = df.columns.str.strip()   # remove trailing spaces in col names
        df.rename(columns={
            'Work Order Number':          'WorkOrderNumber',
            'Basic Start date':           'Basic Start Date',
            'Earliest Basic Start date':  'Created On',      # best proxy for creation date
            'Actual Work Done':           'Actual Hours',
            'Planned Duration':           'Planned Hours',
            'ActualFinish':               'Actual Finish',
            '# Schedule date changes':    'Schedule_Changes',
            '# days to Due Date':         'Days_To_Due',
        }, inplace=True)

        df['Basic Start Date'] = pd.to_datetime(df['Basic Start Date'], errors='coerce')
        df['Created On']       = pd.to_datetime(df['Created On'],       errors='coerce')
        df['Actual Finish']    = pd.to_datetime(df.get('Actual Finish'), errors='coerce')

        # Status flags from Overall Work Order Status (clean enum in new format)
        status = df['Overall Work Order Status']
        df['is_TECO'] = status.isin(['Technically completed'])
        df['is_CLSD'] = status.isin(['Closed'])
        df['is_REL']  = status.isin(['Released', 'Confirmed', 'Partially confirmed'])
        df['is_open'] = ~df['is_TECO'] & ~df['is_CLSD']

        # Derive SAP-compatible OrderType from Maintenance Activity Type + Priority
        mat = df['Maintenance Activity Type'].fillna('')
        pri = df['Priority'].fillna('')
        df['OrderType'] = np.select(
            [
                mat.str.startswith('003'),
                mat.str.startswith('002') & pri.str.startswith('3 | Breakdown'),
                mat.str.startswith('002'),
                mat.str.startswith('021'),
                mat.str.startswith('008'),
            ],
            ['GM09', 'GM02', 'GM01', 'GM01', 'PdM'],
            default='OTHER'
        )

        # Schedule slip: days from original planned date to rescheduled planned date
        df['Schedule_Slip_Days'] = (df['Basic Start Date'] - df['Created On']).dt.days.clip(lower=0)

    else:
        # ── Legacy Celonis format (GM09/GM02/GM01 OrderType) ─────────────────
        df.rename(columns={
            'Work order classification for preventive and corrective processing metrics': 'WO_Classification'
        }, inplace=True)
        df['Basic Start Date'] = pd.to_datetime(df['Basic Start Date'], errors='coerce')
        df['Created On']       = pd.to_datetime(df['Created On'],       errors='coerce')
        df['is_TECO'] = df['System Status'].apply(lambda x: hf(x, 'TECO'))
        df['is_CLSD'] = df['System Status'].apply(lambda x: hf(x, 'CLSD'))
        df['is_REL']  = df['System Status'].apply(lambda x: hf(x, 'REL'))
        df['is_open'] = ~df['is_TECO'] & ~df['is_CLSD']
        df['Schedule_Slip_Days'] = 0
        df['Days_To_Due']        = np.nan
        df['Actual Finish']      = pd.NaT
        df['Schedule_Changes']   = 0

    # ── Common ────────────────────────────────────────────────────────────────
    df['Planned Hours'] = df.get('Planned Hours', pd.Series([np.nan]*len(df)))

    pm  = df[df['OrderType'] == 'GM09'].copy().sort_values('Basic Start Date')
    bkd = df[df['OrderType'] == 'GM02'].copy().sort_values('Basic Start Date')
    fuc = df[df['OrderType'].isin(['GM01','PM01'])].copy().sort_values('Basic Start Date')

    # Reference date: latest actual finish, else latest created on, else today
    ref_date = df.get('Actual Finish', pd.Series(dtype='datetime64[ns]')).max()
    if pd.isna(ref_date):
        ref_date = df['Created On'].max()
    if pd.isna(ref_date):
        ref_date = pd.Timestamp.today()

    return df, pm, bkd, fuc, ref_date


# ═════════════════════════════════════════════════════════════════════════════
# PILLAR ANALYSIS FUNCTIONS
# ═════════════════════════════════════════════════════════════════════════════

def pillar1_compliance(pm, bkd, fuc, df, mw):
    """Compliance & Scheduling"""
    # Exclude future-scheduled WOs from TECO rate (they logically can't be complete yet)
    today    = pd.Timestamp.today().normalize()
    pm_hist  = pm[pm['Basic Start Date'] <= today]
    teco_pct = pm_hist['is_TECO'].mean() * 100 if len(pm_hist) else 0

    # Schedule adherence — window-filtered
    pm_h1 = pm[pm['Basic Start Date'].between(mw['window_start'], mw['window_end'])].copy()
    sched_rows = []
    for desc, grp in pm_h1.groupby('Description'):
        grp = grp.sort_values('Basic Start Date'); interval = pfreq(desc)
        if interval is None or len(grp) < 2: continue
        dates = grp['Basic Start Date'].tolist()
        for i in range(1, len(dates)):
            actual = (dates[i]-dates[i-1]).days; var = actual - interval
            status = ('On time' if abs(var) <= max(1, interval*0.2)
                      else 'Early' if var < 0
                      else 'Late (≤50%)' if var <= interval*0.5
                      else 'Significantly late')
            sched_rows.append({'PM_Description': desc, 'Expected_Days': interval,
                'From': dates[i-1].strftime('%Y-%m-%d'), 'To': dates[i].strftime('%Y-%m-%d'),
                'Actual_Gap': actual, 'Variance': var, 'Status': status})
    sched = pd.DataFrame(sched_rows) if sched_rows else pd.DataFrame()
    if len(sched):
        vc       = sched['Status'].value_counts()
        on_time  = vc.get('On time', 0); early = vc.get('Early', 0)
        late_ct  = vc.get('Late (≤50%)', 0) + vc.get('Significantly late', 0)
        total_i  = len(sched); adh_pct = 100*on_time/total_i
        by_type  = sched.groupby(['PM_Description','Expected_Days']).agg(
            Intervals=('Actual_Gap','count'),
            Avg_Gap=('Actual_Gap','mean'),
            Avg_AbsVar=('Variance', lambda x: x.abs().mean()),  # avg absolute deviation — never misleadingly zero
            Max_Dev=('Variance',   lambda x: x.abs().max()),    # worst single interval
            On_Time=('Status', lambda x: (x=='On time').sum())).reset_index()
        by_type['OT_Pct'] = (by_type['On_Time']/by_type['Intervals']*100).round(0).astype(int)
    else:
        on_time = early = late_ct = total_i = 0; adh_pct = 0; by_type = pd.DataFrame()

    # Cycle time: Created On → Basic Start Date
    pm_with_dates = pm.dropna(subset=['Created On','Basic Start Date']).copy()
    pm_with_dates['Cycle_Days'] = (pm_with_dates['Basic Start Date'] - pm_with_dates['Created On']).dt.days
    pm_with_dates = pm_with_dates[pm_with_dates['Cycle_Days'] >= 0]
    avg_cycle = pm_with_dates['Cycle_Days'].mean() if len(pm_with_dates) else 0
    ct_lt7  = (pm_with_dates['Cycle_Days'] <= 7).sum()
    ct_7_30 = ((pm_with_dates['Cycle_Days'] > 7) & (pm_with_dates['Cycle_Days'] <= 30)).sum()
    ct_30p  = (pm_with_dates['Cycle_Days'] > 30).sum()

    # Actual hours by PM task
    pm_task_hrs = pm.groupby('Description').agg(
        Count=('WorkOrderNumber','count'),
        Total_Hrs=('Actual Hours','sum'),
        Avg_Hrs=('Actual Hours','mean'),
        Zero_Hr=('Actual Hours', lambda x: (x==0).sum())
    ).reset_index()
    pm_task_hrs['Zero_Pct'] = (pm_task_hrs['Zero_Hr'] / pm_task_hrs['Count'] * 100).round(0).astype(int)
    pm_task_hrs = pm_task_hrs.sort_values('Total_Hrs', ascending=False).reset_index(drop=True)

    # Schedule stability (new format only — uses Earliest Basic Start date vs Basic Start Date)
    has_slip = 'Schedule_Slip_Days' in pm.columns and pm['Schedule_Slip_Days'].notna().any()
    if has_slip and len(pm_h1):
        stable_pct = 100.0 * (pm_h1['Schedule_Slip_Days'] == 0).mean()
        rescheduled_df = pm_h1[pm_h1['Schedule_Slip_Days'] > 0][
            ['Description', 'Basic Start Date', 'Schedule_Slip_Days', 'Schedule_Changes']
        ].copy().sort_values('Schedule_Slip_Days', ascending=False).reset_index(drop=True)
        avg_slip       = rescheduled_df['Schedule_Slip_Days'].mean() if len(rescheduled_df) else 0
        multi_reschedule = int((pm_h1.get('Schedule_Changes', pd.Series(dtype=int)) > 1).sum())
    else:
        stable_pct = None; rescheduled_df = pd.DataFrame(); avg_slip = 0; multi_reschedule = 0

    # RAG
    rag = "GREEN" if (teco_pct >= 90 and adh_pct >= 80) else ("AMBER" if (teco_pct >= 75 or adh_pct >= 65) else "RED")

    return {
        'teco_pct': teco_pct, 'adh_pct': adh_pct,
        'on_time': on_time, 'early': early, 'late_ct': late_ct, 'total_i': total_i,
        'avg_cycle': avg_cycle, 'ct_lt7': ct_lt7, 'ct_7_30': ct_7_30, 'ct_30p': ct_30p,
        'sched': sched, 'by_type': by_type, 'pm_task_hrs': pm_task_hrs,
        'stable_pct': stable_pct, 'rescheduled_df': rescheduled_df,
        'avg_slip': avg_slip, 'multi_reschedule': multi_reschedule,
        'rag': rag,
        'key_metric': f"TECO {teco_pct:.0f}%  |  Adherence {adh_pct:.0f}%",
    }


def pillar2_failure_prevention(pm, bkd, mw):
    """Failure Prevention"""
    months_str = mw['months_str']
    month_keys = mw['month_keys']
    month_map  = mw['month_map']
    n_mo       = mw['n']

    bkd_mo = [0]*n_mo
    bkd_h1 = bkd[bkd['Basic Start Date'].between(mw['window_start'], mw['window_end'])].dropna(subset=['Basic Start Date'])
    for _, row in bkd_h1.iterrows():
        k = row['Basic Start Date'].strftime('%Y-%m')
        if k in month_map: bkd_mo[month_map[k]] += 1

    # MTBF monthly (actual inter-event)
    bkd_h1c = bkd_h1.copy(); bkd_h1c['Month'] = bkd_h1c['Basic Start Date'].dt.to_period('M')
    mtbf_30     = [round(30/c, 1) if c > 0 else 0 for c in bkd_mo]
    mtbf_actual = []
    for mk in month_keys:
        grp = bkd_h1c[bkd_h1c['Month'].astype(str)==mk].sort_values('Basic Start Date')
        d   = grp['Basic Start Date'].tolist()
        if len(d) >= 2:
            mtbf_actual.append(round(np.mean([(d[i]-d[i-1]).days for i in range(1,len(d))]),1))
        else:
            mtbf_actual.append(0)

    # Rolling 3-month avg & trend direction
    rolling_3 = []
    for i in range(len(mtbf_30)):
        window = [m for m in mtbf_30[max(0,i-2):i+1] if m > 0]
        rolling_3.append(round(np.mean(window),1) if window else 0)

    h1_vals    = [m for m in mtbf_30 if m > 0]
    mid        = max(1, n_mo // 2)
    first_half = [m for m in mtbf_30[:mid]  if m > 0]
    last_half  = [m for m in mtbf_30[mid:]  if m > 0]
    first3     = np.mean(first_half) if first_half else 0
    last3      = np.mean(last_half)  if last_half  else 0
    if first3 and last3:
        pct_chg = (last3-first3)/first3*100
        trend_dir = "Improving ↑" if pct_chg > 10 else ("Declining ↓" if pct_chg < -10 else "Stable →")
        trend_clr = "GREEN" if pct_chg > 10 else ("RED" if pct_chg < -10 else "AMBER")
    else:
        pct_chg = 0; trend_dir = "Insufficient data"; trend_clr = "AMBER"

    overall_mtbf = round(np.mean(h1_vals),1) if h1_vals else 0

    # Monthly PM count — window
    pm_mo = [0]*n_mo
    pm_h1 = pm[pm['Basic Start Date'].between(mw['window_start'], mw['window_end'])].dropna(subset=['Basic Start Date'])
    for _, row in pm_h1.iterrows():
        k = row['Basic Start Date'].strftime('%Y-%m')
        if k in month_map: pm_mo[month_map[k]] += 1

    # PM:BKD correlation (Pearson) — higher PM count, fewer BKDs?
    if sum(pm_mo) > 0 and sum(bkd_mo) > 0:
        pm_arr  = np.array(pm_mo,  dtype=float)
        bkd_arr = np.array(bkd_mo, dtype=float)
        corr_coef = float(np.corrcoef(pm_arr, bkd_arr)[0,1])
    else:
        corr_coef = 0.0
    corr_interp = ("Negative — more PMs associated with fewer breakdowns ✓" if corr_coef < -0.3
                   else "Positive — PMs and breakdowns trending together (review PM effectiveness)" if corr_coef > 0.3
                   else "No clear linear relationship")

    # Per-task: how often does a BKD occur within 7 days after this PM task?
    task_corr_rows = []
    for desc, grp in pm_h1.groupby('Description'):
        grp = grp.sort_values('Basic Start Date')
        bkd_after = 0
        for _, pm_row in grp.iterrows():
            win_end = pm_row['Basic Start Date'] + pd.Timedelta(days=7)
            nearby  = bkd[(bkd['Basic Start Date'] > pm_row['Basic Start Date']) &
                          (bkd['Basic Start Date'] <= win_end)]
            if len(nearby) > 0: bkd_after += 1
        task_corr_rows.append({
            'PM_Task': desc, 'PM_Count': len(grp),
            'BKD_Within_7d': bkd_after,
            'BKD_Rate_Pct': round(100*bkd_after/len(grp), 0) if len(grp) else 0,
        })
    task_corr = pd.DataFrame(task_corr_rows).sort_values('BKD_Rate_Pct', ascending=False).reset_index(drop=True)

    # PM→Breakdown pairs
    pm_bkd_rows = []
    for _, bd in bkd.iterrows():
        prior = pm[pm['Basic Start Date'] <= bd['Basic Start Date']]
        if not len(prior): continue
        last = prior.iloc[-1]; days = (bd['Basic Start Date'] - last['Basic Start Date']).days
        cat = ("Same day" if days == 0 else "1-3 days" if days <= 3
               else "4-7 days" if days <= 7 else "8+ days")
        pm_bkd_rows.append({'PM_WO': last['WorkOrderNumber'], 'PM_Desc': last['Description'],
            'PM_Date': last['Basic Start Date'].strftime('%Y-%m-%d'),
            'BKD_WO': bd['WorkOrderNumber'], 'BKD_Desc': bd['Description'],
            'BKD_Date': bd['Basic Start Date'].strftime('%Y-%m-%d') if pd.notna(bd['Basic Start Date']) else '—',
            'Days': days, 'Category': cat})
    pm_bkd = pd.DataFrame(pm_bkd_rows) if pm_bkd_rows else pd.DataFrame()

    rag = trend_clr

    return {
        'months_str': months_str, 'bkd_mo': bkd_mo, 'pm_mo': pm_mo,
        'mtbf_30': mtbf_30, 'mtbf_actual': mtbf_actual, 'rolling_3': rolling_3,
        'overall_mtbf': overall_mtbf, 'trend_dir': trend_dir, 'trend_clr': trend_clr,
        'pct_chg': pct_chg, 'pm_bkd': pm_bkd,
        'corr_coef': corr_coef, 'corr_interp': corr_interp, 'task_corr': task_corr,
        'rag': rag,
        'key_metric': f"MTBF {overall_mtbf}d  |  Trend {trend_dir}",
    }


def pillar3_coverage(pm, bkd):
    """Failure Mode Coverage"""
    pm_descs_lower = ' '.join(pm['Description'].fillna('').str.lower().tolist())
    clusters       = build_clusters_from_bkd(bkd)

    coverage_rows = []
    for name, keywords in clusters:
        bkd_mask = bkd['Description'].str.lower().str.contains('|'.join(keywords), na=False)
        bkd_count = bkd_mask.sum()
        pm_covered = any(kw in pm_descs_lower for kw in keywords)
        gap = bkd_count > 0 and not pm_covered
        coverage_rows.append({
            'Cluster': name, 'Keywords': ', '.join(keywords),
            'BKD_Count': bkd_count, 'PM_Covered': pm_covered,
            'Gap': gap,
            'Status': "✗ GAP — no PM task" if gap else ("✓ Covered" if bkd_count > 0 else "No BKDs"),
            'Priority': "HIGH" if (gap and bkd_count >= 3) else ("MED" if gap else "OK"),
        })
    cov_df = pd.DataFrame(coverage_rows)
    gaps     = cov_df[cov_df['Gap']].shape[0]
    high_gaps= cov_df[cov_df['Priority']=='HIGH'].shape[0]

    rag = "GREEN" if gaps == 0 else ("AMBER" if high_gaps == 0 else "RED")

    return {
        'cov_df': cov_df, 'gaps': gaps, 'high_gaps': high_gaps,
        'rag': rag,
        'key_metric': f"{gaps} coverage gaps  |  {high_gaps} high-priority",
    }


def pillar4_wo_quality(pm, bkd, fuc):
    """Work Order Quality — zero-hour + planned vs actual"""
    pm_zero  = (pm['Actual Hours']  == 0).sum()
    bkd_zero = (bkd['Actual Hours'] == 0).sum()
    fuc_zero = (fuc['Actual Hours'] == 0).sum() if len(fuc) else 0

    pm_zero_pct  = 100*pm_zero/len(pm)   if len(pm)  else 0
    bkd_zero_pct = 100*bkd_zero/len(bkd) if len(bkd) else 0
    fuc_zero_pct = 100*fuc_zero/len(fuc)  if len(fuc) else 0

    # Planned vs Actual (only if planned hours column has data)
    has_planned = pm['Planned Hours'].notna().any()
    if has_planned:
        pva = pm.dropna(subset=['Planned Hours','Actual Hours']).copy()
        pva = pva[(pva['Planned Hours'] > 0) & (pva['Actual Hours'] > 0)]
        pva['Variance'] = pva['Actual Hours'] - pva['Planned Hours']
        pva['Var_Pct']  = pva['Variance'] / pva['Planned Hours'] * 100
        avg_var_pct = pva['Var_Pct'].mean() if len(pva) else 0
        over_cnt  = (pva['Variance'] > 0).sum()
        under_cnt = (pva['Variance'] < 0).sum()
        on_cnt    = (pva['Variance'] == 0).sum()
    else:
        pva = pd.DataFrame(); avg_var_pct = 0; over_cnt = under_cnt = on_cnt = 0

    rag = ("GREEN" if (bkd_zero_pct < 20 and pm_zero_pct == 0)
           else "AMBER" if bkd_zero_pct < 60 else "RED")

    return {
        'pm_zero': pm_zero, 'bkd_zero': bkd_zero, 'fuc_zero': fuc_zero,
        'pm_zero_pct': pm_zero_pct, 'bkd_zero_pct': bkd_zero_pct, 'fuc_zero_pct': fuc_zero_pct,
        'has_planned': has_planned, 'pva': pva,
        'avg_var_pct': avg_var_pct, 'over_cnt': over_cnt, 'under_cnt': under_cnt, 'on_cnt': on_cnt,
        'rag': rag,
        'key_metric': f"BKD zero-hr {bkd_zero_pct:.0f}%  |  PM zero-hr {pm_zero_pct:.0f}%",
    }


def pillar5_backlog(df, ref_date, mw):
    """Backlog Health"""
    open_df = df[df['is_open']].copy()
    open_df['Age_Days'] = (ref_date - open_df['Created On']).dt.days.clip(lower=0)

    by_type = open_df.groupby('OrderType').agg(
        Count=('WorkOrderNumber','count'),
        Avg_Age=('Age_Days','mean'),
        Max_Age=('Age_Days','max')).reset_index()

    age_0_30  = (open_df['Age_Days'] <= 30).sum()
    age_31_90 = ((open_df['Age_Days'] > 30) & (open_df['Age_Days'] <= 90)).sum()
    age_91p   = (open_df['Age_Days'] > 90).sum()
    total_open = len(open_df)
    avg_age    = open_df['Age_Days'].mean() if total_open else 0
    oldest_age = open_df['Age_Days'].max() if total_open else 0

    # Monthly new WOs vs TECO'd — window
    months_str = mw['months_str']
    month_keys = mw['month_keys']
    new_mo = [0]*mw['n']; closed_mo = [0]*mw['n']
    for _, row in df.iterrows():
        k = row['Created On'].strftime('%Y-%m') if pd.notna(row['Created On']) else None
        if k in month_keys: new_mo[month_keys.index(k)] += 1
        k2 = row['Basic Start Date'].strftime('%Y-%m') if pd.notna(row['Basic Start Date']) else None
        if k2 in month_keys and row['is_TECO']: closed_mo[month_keys.index(k2)] += 1
    net_mo = [n-c for n, c in zip(new_mo, closed_mo)]
    running_net = []
    running = 0
    for n in net_mo:
        running += n; running_net.append(running)

    # Due date urgency (new format — Days_To_Due field)
    due_df = df[df['Days_To_Due'].notna()].copy() if 'Days_To_Due' in df.columns else pd.DataFrame()
    if len(due_df):
        overdue_ct = int((due_df['Days_To_Due'] < 0).sum())
        due_0_7    = int(((due_df['Days_To_Due'] >= 0) & (due_df['Days_To_Due'] <= 7)).sum())
        due_8_30   = int(((due_df['Days_To_Due'] > 7) & (due_df['Days_To_Due'] <= 30)).sum())
        overdue_df = (due_df[due_df['Days_To_Due'] < 0]
                      .sort_values('Days_To_Due')[['WorkOrderNumber','Description','Days_To_Due','Overall Work Order Status'
                                                    if 'Overall Work Order Status' in due_df.columns else 'OrderType']]
                      .head(15).reset_index(drop=True))
    else:
        overdue_ct = due_0_7 = due_8_30 = 0; overdue_df = pd.DataFrame()

    rag = ("GREEN" if (age_91p == 0 and avg_age < 30)
           else "AMBER" if age_91p <= 5
           else "RED")
    # Escalate if overdue WOs exist
    if overdue_ct > 10 and rag == "GREEN": rag = "AMBER"

    return {
        'open_df': open_df, 'by_type': by_type,
        'age_0_30': age_0_30, 'age_31_90': age_31_90, 'age_91p': age_91p,
        'total_open': total_open, 'avg_age': avg_age, 'oldest_age': oldest_age,
        'months_str': months_str, 'new_mo': new_mo, 'closed_mo': closed_mo,
        'net_mo': net_mo, 'running_net': running_net,
        'overdue_ct': overdue_ct, 'due_0_7': due_0_7, 'due_8_30': due_8_30,
        'overdue_df': overdue_df,
        'rag': rag,
        'key_metric': f"{total_open} open  |  {age_91p} aged 90d+  |  {overdue_ct} overdue",
    }


def pillar6_cost(pm, bkd, fuc, df, cost_df=None, mw=None):
    """Cost Visibility — enhanced with cost export when available"""
    # ── labor hour analysis (always) ──
    bkd_zero     = (bkd['Actual Hours'] == 0).sum()
    bkd_zero_pct = 100*bkd_zero/len(bkd) if len(bkd) else 0
    fuc_zero     = (fuc['Actual Hours'] == 0).sum()
    fuc_zero_pct = 100*fuc_zero/len(fuc) if len(fuc) else 0
    pm_zero      = (pm['Actual Hours'] == 0).sum()
    bkd_with_hrs = bkd[bkd['Actual Hours'] > 0]
    avg_bkd_hr   = bkd_with_hrs['Actual Hours'].mean() if len(bkd_with_hrs) else 0

    has_cost = cost_df is not None and len(cost_df) > 0

    if has_cost:
        # Join cost to main df on WO number
        cost_df = cost_df.copy()
        cost_df['WO_Number'] = cost_df['Work Order Number'].astype(str).str.strip()
        df2 = df.copy()
        df2['WO_Number'] = df2['WorkOrderNumber'].astype(str).str.strip()
        merged = df2.merge(
            cost_df[['WO_Number','Cost per Work Order','Actual Cost Classification']],
            on='WO_Number', how='left'
        )
        pm_m  = merged[merged['OrderType']=='GM09']
        bkd_m = merged[merged['OrderType']=='GM02']
        fuc_m = merged[merged['OrderType'].isin(['GM01','PM01'])]

        total_cost   = merged['Cost per Work Order'].sum()
        pm_cost      = pm_m['Cost per Work Order'].sum()
        bkd_cost     = bkd_m['Cost per Work Order'].sum()
        fuc_cost     = fuc_m['Cost per Work Order'].sum()
        avg_pm_cost  = pm_m['Cost per Work Order'].mean()
        avg_bkd_cost = bkd_m['Cost per Work Order'].mean()
        cost_ratio   = avg_bkd_cost / avg_pm_cost if avg_pm_cost else 0

        # Cost classification counts
        vc = cost_df['Actual Cost Classification'].str.strip().value_counts()
        exceeds  = vc.get('Actual Cost Exceeds Planned Cost', 0)
        within   = vc.get('Within Boundaries', 0)
        below    = vc.get('Actual Cost is Lower than Planned Cost', 0)
        exceed_pct = 100*exceeds/len(cost_df) if len(cost_df) else 0

        # Monthly cost trend — window
        _mw = mw if mw else build_month_window(df)
        month_keys = _mw['month_keys']
        months_str = _mw['months_str']
        merged['Month'] = pd.to_datetime(merged['Created On'], errors='coerce').dt.to_period('M').astype(str)
        cost_mo = []
        for mk in month_keys:
            cost_mo.append(merged[merged['Month']==mk]['Cost per Work Order'].sum())

        # Top 10 WOs by cost
        top_wos = cost_df.nlargest(10,'Cost per Work Order')[
            ['Work Order Number','Description','Cost per Work Order','Actual Cost Classification']
        ].copy()

        # WOs with cost > 0 but no hours (critical flag)
        merged['has_cost']  = merged['Cost per Work Order'] > 0
        merged['zero_hrs']  = merged['Actual Hours'] == 0
        cost_no_hrs = merged[(merged['has_cost']) & (merged['zero_hrs'])]
        cost_no_hrs_bkd = cost_no_hrs[cost_no_hrs['OrderType']=='GM02']

        rag_cost = "GREEN" if exceed_pct < 15 else ("AMBER" if exceed_pct < 30 else "RED")
        # If breakdowns have cost but no hours — escalate to at least AMBER
        if len(cost_no_hrs_bkd) > 5 and rag_cost == "GREEN": rag_cost = "AMBER"

        rag = rag_cost
        key_metric = (f"Total ${total_cost:,.0f}  |  BKD avg ${avg_bkd_cost:,.0f} "
                      f"vs PM avg ${avg_pm_cost:,.0f}  |  {exceed_pct:.0f}% over plan")
    else:
        # No cost file — fall back to labor-only analysis
        total_cost = pm_cost = bkd_cost = fuc_cost = 0
        avg_pm_cost = avg_bkd_cost = cost_ratio = 0
        exceeds = within = below = exceed_pct = 0
        _mw = mw if mw else build_month_window(df)
        cost_mo = [0]*_mw['n']; months_str = _mw['months_str']
        top_wos = pd.DataFrame(); merged = df.copy()
        cost_no_hrs_bkd = pd.DataFrame()
        rag = "RED" if bkd_zero_pct > 50 else ("AMBER" if bkd_zero_pct > 20 else "GREEN")
        key_metric = f"No cost file — BKD hrs captured: {100-bkd_zero_pct:.0f}%"

    return {
        'has_cost': has_cost,
        'bkd_zero': bkd_zero, 'bkd_zero_pct': bkd_zero_pct,
        'fuc_zero': fuc_zero, 'fuc_zero_pct': fuc_zero_pct,
        'pm_zero': pm_zero, 'avg_bkd_hr': avg_bkd_hr,
        'total_cost': total_cost, 'pm_cost': pm_cost,
        'bkd_cost': bkd_cost, 'fuc_cost': fuc_cost,
        'avg_pm_cost': avg_pm_cost, 'avg_bkd_cost': avg_bkd_cost,
        'cost_ratio': cost_ratio,
        'exceeds': exceeds, 'within': within, 'below': below, 'exceed_pct': exceed_pct,
        'cost_mo': cost_mo, 'months_str': months_str,
        'top_wos': top_wos, 'cost_no_hrs_bkd': cost_no_hrs_bkd,
        'rag': rag, 'key_metric': key_metric,
    }


def pillar7_reliability(bkd, df, ref_date, mw):
    """Equipment Reliability — MTBF + MTTR"""
    bkd_sorted = bkd.dropna(subset=['Basic Start Date']).sort_values('Basic Start Date')
    bd_dates   = bkd_sorted['Basic Start Date'].tolist()
    gaps       = [(bd_dates[i]-bd_dates[i-1]).days for i in range(1, len(bd_dates))]
    mtbf_overall = round(np.mean(gaps),1) if gaps else 0
    mtbf_min     = min(gaps) if gaps else 0
    mtbf_max     = max(gaps) if gaps else 0

    # MTTR from labor hours (booked on BKD WOs)
    bkd_with_hrs = bkd[bkd['Actual Hours'] > 0]
    mttr         = round(bkd_with_hrs['Actual Hours'].mean(),1) if len(bkd_with_hrs) else None
    mttr_n       = len(bkd_with_hrs)
    mttr_note    = f"Based on {mttr_n}/{len(bkd)} BKD WOs with actual hours recorded" if mttr else "No actual hours on BKD WOs — MTTR cannot be calculated"

    # MTTR from ActualFinish − Basic Start Date (elapsed response time, new format only)
    if 'Actual Finish' in bkd.columns:
        bkd_fin = bkd.dropna(subset=['Actual Finish', 'Basic Start Date']).copy()
        bkd_fin = bkd_fin[bkd_fin['Actual Finish'] > bkd_fin['Basic Start Date']]
        bkd_fin['elapsed_hrs'] = (bkd_fin['Actual Finish'] - bkd_fin['Basic Start Date']).dt.total_seconds() / 3600
        # Cap at 240h to exclude parked/stale WOs
        bkd_fin = bkd_fin[bkd_fin['elapsed_hrs'] <= 240]
        mttr_elapsed   = round(bkd_fin['elapsed_hrs'].mean(), 1) if len(bkd_fin) else None
        mttr_elapsed_n = len(bkd_fin)
    else:
        mttr_elapsed = None; mttr_elapsed_n = 0

    # Breakdown rate by month — window
    months_str = mw['months_str']
    month_keys = mw['month_keys']
    bkd_mo     = [0]*mw['n']
    bkd_h1     = bkd[bkd['Basic Start Date'].between(mw['window_start'], mw['window_end'])].dropna(subset=['Basic Start Date'])
    for _, row in bkd_h1.iterrows():
        k = row['Basic Start Date'].strftime('%Y-%m')
        if k in month_keys: bkd_mo[month_keys.index(k)] += 1

    # MTTR by month
    mttr_mo = []
    bkd_h1c = bkd_h1.copy(); bkd_h1c['Month'] = bkd_h1c['Basic Start Date'].dt.to_period('M')
    for mk in month_keys:
        grp = bkd_h1c[(bkd_h1c['Month'].astype(str)==mk) & (bkd_h1c['Actual Hours']>0)]
        mttr_mo.append(round(grp['Actual Hours'].mean(),1) if len(grp) else 0)

    # Best/worst months
    best_idx  = bkd_mo.index(min(bkd_mo)) if bkd_mo else 0
    worst_idx = bkd_mo.index(max(bkd_mo)) if bkd_mo else 0

    rag = ("GREEN" if mtbf_overall > 7 else "AMBER" if mtbf_overall > 3 else "RED")

    return {
        'mtbf_overall': mtbf_overall, 'mtbf_min': mtbf_min, 'mtbf_max': mtbf_max,
        'mttr': mttr, 'mttr_n': mttr_n, 'mttr_note': mttr_note,
        'mttr_elapsed': mttr_elapsed, 'mttr_elapsed_n': mttr_elapsed_n,
        'months_str': months_str, 'bkd_mo': bkd_mo, 'mttr_mo': mttr_mo,
        'best_idx': best_idx, 'worst_idx': worst_idx, 'gaps': gaps,
        'rag': rag,
        'key_metric': f"MTBF {mtbf_overall}d  |  MTTR {'—' if not mttr else str(mttr)+'h'}",
    }


# ═════════════════════════════════════════════════════════════════════════════
# DYNAMIC RECOMMENDATIONS ENGINE
# ═════════════════════════════════════════════════════════════════════════════
def generate_recommendations(p1, p2, p3, p4, p5, p6, p7, pm, bkd, mw):
    """
    Evaluate pillar metrics against threshold rules and return ranked priority actions.
    Each recommendation only fires when its specific threshold is crossed — nothing
    is hardcoded. Specific task names, counts, and dollar figures come from the data.
    Returns top 5 by severity score.
    """
    recs = []

    # ── 1. POST-PM DISTURBANCE FLAG ───────────────────────────────────────────
    # Fires when any PM task has ≥50% post-execution breakdown rate within 7 days
    tc = p2.get('task_corr', pd.DataFrame())
    if len(tc):
        worst_tc = tc.iloc[0]   # already sorted desc by BKD_Rate_Pct
        if worst_tc['BKD_Rate_Pct'] >= 50:
            rate = int(worst_tc['BKD_Rate_Pct'])
            name = str(worst_tc['PM_Task'])[:60]
            cnt  = int(worst_tc['PM_Count'])
            recs.append({
                'severity': 'HIGH',
                'category': '2 · Failure Prevention',
                'finding':  f'"{name}" — {rate}% of {cnt} executions followed by a breakdown within 7 days',
                'action':   'Review task steps for component disturbance; consider revised procedure or split inspection approach',
                'score':    rate * cnt,
            })

    # ── 2. COVERAGE GAP — highest-risk uncovered failure cluster ──────────────
    # Fires when any failure cluster has breakdowns but no PM task addresses it
    cov = p3.get('cov_df', pd.DataFrame())
    gap_rows = cov[cov['Gap']].sort_values('BKD_Count', ascending=False) if len(cov) else pd.DataFrame()
    if len(gap_rows):
        top_gap  = gap_rows.iloc[0]
        bkd_n    = int(top_gap['BKD_Count'])
        recs.append({
            'severity': 'HIGH' if bkd_n >= 3 else 'MED',
            'category': '3 · Failure Mode Coverage',
            'finding':  f'"{top_gap["Cluster"]}" — {bkd_n} BKD WOs, no PM task covers this failure mode',
            'action':   f'Add PM task targeting {top_gap["Cluster"].lower()} components; review {bkd_n} BKD WOs for root cause patterns',
            'score':    bkd_n * 10,
        })

    # ── 3. LABOR HOUR COMPLIANCE ──────────────────────────────────────────────
    # Fires when >20% of breakdown WOs have zero actual hours recorded
    bkd_total  = len(bkd)
    bkd_zh_pct = p4.get('bkd_zero_pct', 0)
    if bkd_zh_pct > 20 and bkd_total > 0:
        recs.append({
            'severity': 'HIGH' if bkd_zh_pct > 60 else 'MED',
            'category': '4 · WO Quality',
            'finding':  f'{p4["bkd_zero"]}/{bkd_total} breakdown WOs ({bkd_zh_pct:.0f}%) have zero actual hours — MTTR and failure cost invisible',
            'action':   'Require technicians to book hours on all GM02 WOs; add labor entry to supervisor close-out checklist',
            'score':    bkd_zh_pct * 0.8,
        })

    # ── 4. MTBF DECLINING TRAJECTORY ─────────────────────────────────────────
    # Fits a linear trend to monthly MTBF values; fires when slope < -1 day/month
    mtbf_vals = [m for m in p2.get('mtbf_30', []) if m > 0]
    if len(mtbf_vals) >= 3:
        x     = np.arange(len(mtbf_vals), dtype=float)
        slope = float(np.polyfit(x, mtbf_vals, 1)[0])
        if slope < -1.0:
            proj_3mo = max(0.0, mtbf_vals[-1] + slope * 3)
            recs.append({
                'severity': 'HIGH' if slope < -3 else 'MED',
                'category': '7 · Equipment Reliability',
                'finding':  f'MTBF declining {abs(slope):.1f} days/month — projected {proj_3mo:.0f}d in 3 months (current {mtbf_vals[-1]:.0f}d)',
                'action':   'Investigate top failure cluster; consider increasing PM frequency or adding condition-based monitoring',
                'score':    abs(slope) * 15,
            })

    # ── 5. WORST COMPLIANCE TASK ──────────────────────────────────────────────
    # Fires when a PM task with ≥3 measured intervals has <70% on-time rate
    by_type = p1.get('by_type', pd.DataFrame())
    if len(by_type):
        meaningful = by_type[by_type['Intervals'] >= 3]
        if len(meaningful):
            worst_ct = meaningful.sort_values('OT_Pct').iloc[0]
            if worst_ct['OT_Pct'] < 70:
                name = str(worst_ct['PM_Description'])[:55]
                ot   = int(worst_ct['OT_Pct'])
                recs.append({
                    'severity': 'MED',
                    'category': '1 · Compliance & Scheduling',
                    'finding':  f'"{name}" — {ot}% on-time ({int(worst_ct["Intervals"])} intervals, avg |dev| ±{worst_ct["Avg_AbsVar"]:.0f}d)',
                    'action':   'Review scheduling window and execution capacity; check if frequency aligns with available labor hours',
                    'score':    (100 - ot) * 0.6,
                })

    # ── 6. COST LEVERAGE RATIO ────────────────────────────────────────────────
    # Fires when BKD avg cost > 2× PM avg cost (requires cost export)
    if p6.get('has_cost') and p6.get('cost_ratio', 0) > 2.0 and p6.get('avg_pm_cost', 0) > 0:
        ratio    = p6['cost_ratio']
        est_save = p6['bkd_cost'] * 0.30   # conservative 30% prevention improvement
        recs.append({
            'severity': 'MED',
            'category': '6 · Cost Visibility',
            'finding':  f'BKD avg ${p6["avg_bkd_cost"]:,.0f} vs PM avg ${p6["avg_pm_cost"]:,.0f} — {ratio:.1f}× cost leverage ratio',
            'action':   f'Improving prevention rate 30% → est. ${est_save:,.0f} reduction in reactive spend',
            'score':    ratio * 8,
        })

    # ── 7. BACKLOG GROWING ────────────────────────────────────────────────────
    # Fires when 3-month trailing avg net WOs/month > 2
    net = p5.get('net_mo', [])
    if len(net) >= 3:
        avg_net = float(np.mean(net[-3:]))
        if avg_net > 2:
            months_to_double = (p5['total_open'] / avg_net) if avg_net > 0 else 999
            recs.append({
                'severity': 'MED',
                'category': '5 · Backlog Health',
                'finding':  f'Backlog growing avg {avg_net:.0f} WOs/month (last 3 months) — {p5["total_open"]} open WOs currently',
                'action':   f'At this rate backlog doubles in {months_to_double:.0f} months — review close-out rate and WO validity',
                'score':    avg_net * 5,
            })

    # ── 8. PM INTERVAL vs BKD INTERVAL MISMATCH ──────────────────────────────
    # Fires when overall BKD inter-arrival < 70% of the shortest PM interval
    # (equipment failing faster than PM cadence was designed for)
    if p7.get('mtbf_overall', 0) > 0 and len(by_type) and 'Expected_Days' in by_type.columns:
        exp_valid = by_type['Expected_Days'][by_type['Expected_Days'] > 0]
        if len(exp_valid):
            min_pm  = float(exp_valid.min())
            avg_bkd = p7['mtbf_overall']
            if avg_bkd < min_pm * 0.70:
                recs.append({
                    'severity': 'HIGH',
                    'category': '2 · Failure Prevention',
                    'finding':  f'Avg BKD interval {avg_bkd:.0f}d vs shortest PM interval {min_pm:.0f}d — equipment failing faster than PM cadence',
                    'action':   'Increase PM frequency on highest-failure tasks; evaluate condition-based triggers between scheduled PMs',
                    'score':    (min_pm / max(avg_bkd, 1)) * 12,
                })

    recs.sort(key=lambda x: x['score'], reverse=True)
    return recs[:5]


# ═════════════════════════════════════════════════════════════════════════════
# WORKBOOK BUILDER
# ═════════════════════════════════════════════════════════════════════════════
def build_workbook(df, pm, bkd, fuc, ref_date, asset_label, analyses, cost_df=None, mw=None):
    p1, p2, p3, p4, p5, p6, p7 = analyses
    if mw is None: mw = build_month_window(df)
    n_mo = mw['n']
    pl   = mw['period_label']   # e.g. "H1 2026" or "Jan 2025 – Jun 2026"
    date_range = (f"{df['Created On'].min().strftime('%b %Y')}–"
                  f"{df['Created On'].max().strftime('%b %Y')}")
    wb = Workbook()

    # ── SHEET 0: Executive Scorecard ─────────────────────────────────────────
    ws0 = wb.active; ws0.title = "Executive Scorecard"; ws0.sheet_view.showGridLines = False
    title_row(ws0, f"PM EFFECTIVENESS  ·  {asset_label}  ·  EXECUTIVE SCORECARD", 1, "L")
    sub_row(ws0, f"Data: {date_range}  |  {len(df)} WOs  |  Ref date: {ref_date.strftime('%Y-%m-%d')}  |  GM09=Preventive  GM02=Breakdown  GM01/PM01=Follow-up Corrective", 2, "L")
    ws0.row_dimensions[3].height = 10

    # Scorecard header
    for col, txt, ec in [("B","Pillar","D"),("E","RAG",None),("F","Key Metric","J"),("K","Action Required","L")]:
        hdr(ws0, col, 4, txt, ec)
    ws0.row_dimensions[4].height = 18

    scorecard = [
        (PILLARS[0], p1['rag'], p1['key_metric'],
         "Review late PMs; investigate cycle time outliers" if p1['rag'] != "GREEN" else "Maintain current PM execution cadence"),
        (PILLARS[1], p2['rag'], p2['key_metric'],
         f"MTBF trending {p2['trend_dir']} — review PM frequency for top failure modes" if p2['rag'] != "GREEN" else "MTBF trend positive — continue monitoring"),
        (PILLARS[2], p3['rag'], p3['key_metric'],
         f"{p3['gaps']} failure clusters have no PM task — expand PM scope" if p3['gaps'] else "All failure clusters have PM coverage"),
        (PILLARS[3], p4['rag'], p4['key_metric'],
         f"Require technicians to book hours on GM02 BKD WOs — {p4['bkd_zero']} with 0hrs" if p4['rag'] != "GREEN" else "Hour capture is complete"),
        (PILLARS[4], p5['rag'], p5['key_metric'],
         f"Review {p5['age_91p']} WOs aged 90+ days; assess if still valid" if p5['rag'] != "GREEN" else "Backlog age is healthy"),
        (PILLARS[5], p6['rag'], p6['key_metric'],
         (f"BKD avg cost ${p6['avg_bkd_cost']:,.0f} vs PM avg ${p6['avg_pm_cost']:,.0f}. {p6['bkd_zero']} BKD WOs still missing labor hours."
          if p6['has_cost'] else
          f"{p6['bkd_zero']} BKD WOs have zero labor hours — failure cost unknown. Upload cost export for spend analysis.")),
        (PILLARS[6], p7['rag'], p7['key_metric'],
         f"MTBF {p7['mtbf_overall']}d — investigate top failure clusters to extend run intervals" if p7['rag'] != "GREEN" else f"MTBF is healthy at {p7['mtbf_overall']} days"),
    ]

    for i, (pillar, rag, metric, action) in enumerate(scorecard):
        r = i + 5; bg = rag_bg(rag)
        ws0.merge_cells(f"B{r}:D{r}"); ws0.merge_cells(f"F{r}:J{r}"); ws0.merge_cells(f"K{r}:L{r}")
        dc(ws0,"B",r,pillar,bg,bold=True,align="left")
        ws0[f"E{r}"].value = rag_label(rag)
        ws0[f"E{r}"].font = Font(name="Arial",bold=True,size=9,color=rag_fg(rag))
        ws0[f"E{r}"].fill = wfill(bg); ws0[f"E{r}"].alignment = cal(); ws0[f"E{r}"].border = bdr()
        dc(ws0,"F",r,metric,bg,align="left"); dc(ws0,"K",r,action,bg,align="left",wrap=True)
        ws0.row_dimensions[r].height = 26

    ws0.row_dimensions[13].height = 10
    sec(ws0,"OVERALL FINDINGS",13,"L")
    total_bkds = len(bkd)
    red_count  = sum(1 for _,rag,*_ in scorecard if rag=="RED")
    amb_count  = sum(1 for _,rag,*_ in scorecard if rag=="AMBER")
    findings = [
        f"{red_count} pillar(s) RED / {amb_count} AMBER — see detail tabs for corrective actions.",
        f"PM compliance: {p1['teco_pct']:.0f}% TECO rate across {len(pm)} PM WOs. Schedule adherence: {p1['adh_pct']:.0f}%."
        + (f"  Schedule stability: {p1['stable_pct']:.0f}% executed on original planned date." if p1.get('stable_pct') is not None else ""),
        f"MTBF: {p7['mtbf_overall']} days overall ({pl}). Trend: {p2['trend_dir']}."
        + (f"  Elapsed response time: {p7['mttr_elapsed']}h avg ({p7['mttr_elapsed_n']} closed BKDs)." if p7.get('mttr_elapsed') else ""),
        f"{p3['gaps']} of {len(p3['cov_df'])} failure clusters have no corresponding PM task.",
        (f"⚠  {p4['bkd_zero']}/{total_bkds} breakdown WOs carry zero actual hours — failure labor cost is largely invisible."
         if total_bkds > 0 else
         "✓ No breakdown WOs in this period — equipment running without recorded failures."),
        f"Backlog: {p5['total_open']} open WOs, avg age {p5['avg_age']:.0f} days, {p5['age_91p']} aged 90+ days."
        + (f"  ⚠ {p5['overdue_ct']} WOs currently overdue — {p5['due_0_7']} more due within 7 days." if p5.get('overdue_ct', 0) > 0 else ""),
    ]
    for j, txt in enumerate(findings):
        r = 14 + j
        fg = rag_fg("RED") if "⚠" in txt else "1F2D3D"
        bg = RAG_RED_BG if "⚠" in txt else (RAG_RED_BG if j==0 and red_count>0 else LIGHT_ROW)
        ws0.merge_cells(f"B{r}:L{r}")
        dc(ws0,"B",r,txt,bg,align="left",wrap=True); ws0.row_dimensions[r].height = 22

    # ── PRIORITY ACTIONS — dynamic, data-driven, ranked by severity score ────
    recs      = generate_recommendations(p1, p2, p3, p4, p5, p6, p7, pm, bkd, mw)
    r_pa_hdr  = 14 + len(findings) + 1   # one spacer row, then section header
    r_pa_data = r_pa_hdr + 1
    ws0.row_dimensions[r_pa_hdr - 1].height = 10
    sec(ws0, f"PRIORITY ACTIONS  ·  {len(recs)} item(s) ranked by estimated impact", r_pa_hdr, "L")
    sev_bg = {"HIGH": RAG_RED_BG, "MED": ORANGE, "LOW": LIGHT_ROW}
    sev_fc = {"HIGH": RAG_RED,    "MED": "833C11","LOW": "1F2D3D"}
    if not recs:
        ws0.merge_cells(f"B{r_pa_data}:L{r_pa_data}")
        dc(ws0,"B",r_pa_data,"✓ No priority actions — all pillar metrics within acceptable thresholds.",GREEN_FLG,align="left")
        ws0.row_dimensions[r_pa_data].height = 22
    else:
        for col,txt,ec in [("B","#",None),("C","Severity","D"),("E","Pillar","G"),
                            ("H","Finding","J"),("K","Recommended Action","L")]:
            hdr(ws0,col,r_pa_data,txt,ec)
        ws0.row_dimensions[r_pa_data].height = 18
        for i, rec in enumerate(recs):
            r  = r_pa_data + 1 + i
            bg = sev_bg.get(rec['severity'], LIGHT_ROW)
            ws0.merge_cells(f"C{r}:D{r}"); ws0.merge_cells(f"E{r}:G{r}")
            ws0.merge_cells(f"H{r}:J{r}"); ws0.merge_cells(f"K{r}:L{r}")
            dc(ws0,"B",r,str(i+1),bg,bold=True)
            ws0[f"C{r}"].value     = rec['severity']
            ws0[f"C{r}"].font      = Font(name="Arial",bold=True,size=9,color=sev_fc.get(rec['severity'],"1F2D3D"))
            ws0[f"C{r}"].fill      = wfill(bg)
            ws0[f"C{r}"].alignment = cal()
            ws0[f"C{r}"].border    = bdr()
            dc(ws0,"E",r,rec['category'],bg,align="left")
            dc(ws0,"H",r,rec['finding'],bg,align="left",wrap=True)
            dc(ws0,"K",r,rec['action'],bg,align="left",wrap=True)
            ws0.row_dimensions[r].height = 52

    widths(ws0,{"A":2,"B":4,"C":9,"D":4,"E":12,"F":8,"G":6,"H":12,"I":10,"J":10,"K":28,"L":6})

    # ── SHEET 1: Pillar 1 — Compliance & Scheduling ──────────────────────────
    ws1 = wb.create_sheet("1 · Compliance"); ws1.sheet_view.showGridLines = False
    title_row(ws1, f"PILLAR 1: COMPLIANCE & SCHEDULING  ·  {asset_label}", 1, "L")
    sub_row(ws1, "PM TECO rate, schedule adherence by task, and WO cycle time (Created → Start)", 2, "L")
    ws1.row_dimensions[3].height = 10

    # KPI row
    kpis = [
        ("B","PM TECO Rate",f"{p1['teco_pct']:.1f}%",rag_bg("GREEN" if p1['teco_pct']>=90 else "AMBER" if p1['teco_pct']>=75 else "RED")),
        ("F","Schedule Adherence",f"{p1['adh_pct']:.1f}%",rag_bg("GREEN" if p1['adh_pct']>=80 else "AMBER" if p1['adh_pct']>=65 else "RED")),
        ("J","Avg Cycle Time",f"{p1['avg_cycle']:.1f} days",LIGHT_ROW),
    ]
    for col, lbl, val, bg in kpis:
        ec = chr(ord(col)+2)
        ws1.merge_cells(f"{col}4:{ec}4"); ws1[f"{col}4"].value = lbl
        ws1[f"{col}4"].font = Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws1[f"{col}4"].fill = wfill(WHITE); ws1[f"{col}4"].alignment = cal("left")
        ws1.merge_cells(f"{col}5:{ec}5"); ws1[f"{col}5"].value = val
        ws1[f"{col}5"].font = Font(name="Arial",bold=True,size=16,color="1F2D3D")
        ws1[f"{col}5"].fill = wfill(bg); ws1[f"{col}5"].alignment = cal(); ws1[f"{col}5"].border = bdr()
    ws1.row_dimensions[4].height=16; ws1.row_dimensions[5].height=30; ws1.row_dimensions[6].height=10

    sec(ws1,f"SCHEDULE ADHERENCE BY PM TASK  ({pl})",7,"L")
    sub_row(ws1,"On time = gap within ±20% of expected interval (min ±1d). Avg |Dev| = mean absolute deviation — can't be zero when OT% < 100%.",8,"L",height=26)
    for col,txt,ec in [("B","PM Task","D"),("E","Interval / Tol","F"),("G","Count","H"),
                        ("I","Avg Gap","J"),("K","Avg |Dev|","L"),("M","OT %",None)]:
        hdr(ws1,col,9,txt,ec)
    ws1.row_dimensions[9].height = 20
    if len(p1['by_type']):
        sorted_bt = p1['by_type'].sort_values('OT_Pct', ascending=True).reset_index(drop=True)
        for i, row in sorted_bt.iterrows():
            r  = 10 + i
            ot = row['OT_Pct']
            bg = GREEN_FLG if ot==100 else (LIGHT_ROW if ot>=90 else (ORANGE if ot>=67 else RED_FLG))
            abs_var_s = f"±{row['Avg_AbsVar']:.1f}d (max {row['Max_Dev']:.0f}d)"
            tol = max(1, int(row['Expected_Days']) * 0.2) if row['Expected_Days'] else 1
            for mc in [f"B{r}:D{r}",f"E{r}:F{r}",f"G{r}:H{r}",f"I{r}:J{r}",f"K{r}:L{r}"]:
                ws1.merge_cells(mc)
            dc(ws1,"B",r,row['PM_Description'],bg,bold=True,align="left",wrap=True)
            dc(ws1,"E",r,f"{int(row['Expected_Days'])}d / ±{tol:.0f}d",bg)
            dc(ws1,"G",r,int(row['Intervals']),bg)
            dc(ws1,"I",r,f"{row['Avg_Gap']:.1f}d",bg)
            dc(ws1,"K",r,abs_var_s,bg)
            dc(ws1,"M",r,f"{int(ot)}%",bg,bold=True); ws1.row_dimensions[r].height=26

    row_ct = 9 + (len(p1['by_type']) if len(p1['by_type']) else 1) + 1
    ws1.row_dimensions[row_ct].height = 10
    sec(ws1,"WO CYCLE TIME  (Created → Basic Start Date)",row_ct+1,"L")
    for col,txt,ec in [("B","Bucket",None),("D","Count",None),("F","% of PM WOs","H")]:
        hdr(ws1,col,row_ct+2,txt,ec)
    ws1.row_dimensions[row_ct+2].height=18
    ct_total = p1['ct_lt7'] + p1['ct_7_30'] + p1['ct_30p']
    for i,(lbl,ct) in enumerate([("≤ 7 days",p1['ct_lt7']),("8 – 30 days",p1['ct_7_30']),("31+ days",p1['ct_30p'])]):
        r = row_ct+3+i; bg = GREEN_FLG if i==0 else (LIGHT_ROW if i==1 else RED_FLG)
        ws1.merge_cells(f"B{r}:C{r}"); ws1.merge_cells(f"D{r}:E{r}"); ws1.merge_cells(f"F{r}:H{r}")
        dc(ws1,"B",r,lbl,bg,bold=True); dc(ws1,"D",r,ct,bg)
        dc(ws1,"F",r,f"{100*ct/ct_total:.0f}%" if ct_total else "—",bg)
        ws1.row_dimensions[r].height=18

    # ── Actual hours by PM task ──
    r_hrs = row_ct + 7; ws1.row_dimensions[r_hrs].height=10
    sec(ws1,"ACTUAL HOURS ENTERED BY PM TASK",r_hrs+1,"M")
    for col,txt,ec in [("B","PM Task Description","E"),("F","WO Count","G"),
                        ("H","Total Hours","I"),("J","Avg Hrs / WO","K"),
                        ("L","Zero-Hr WOs","M")]:
        hdr(ws1,col,r_hrs+2,txt,ec)
    ws1.row_dimensions[r_hrs+2].height=18
    total_pm_hrs = p1['pm_task_hrs']['Total_Hrs'].sum()
    for i,(_,row) in enumerate(p1['pm_task_hrs'].iterrows()):
        r = r_hrs+3+i
        bg = RED_FLG if row['Zero_Hr']>0 else (GREEN_FLG if row['Total_Hrs']>0 else LIGHT_ROW)
        bg = LIGHT_ROW if i%2==0 and row['Zero_Hr']==0 else bg
        for mc in [f"B{r}:E{r}",f"F{r}:G{r}",f"H{r}:I{r}",f"J{r}:K{r}",f"L{r}:M{r}"]:
            ws1.merge_cells(mc)
        dc(ws1,"B",r,row['Description'],bg,align="left")
        dc(ws1,"F",r,int(row['Count']),bg)
        dc(ws1,"H",r,f"{row['Total_Hrs']:.1f}",bg,bold=True)
        dc(ws1,"J",r,f"{row['Avg_Hrs']:.2f}",bg)
        dc(ws1,"L",r,f"{int(row['Zero_Hr'])} ({int(row['Zero_Pct'])}%)",
           YELLOW if row['Zero_Hr']>0 else GREEN_FLG, bold=(row['Zero_Hr']>0))
        ws1.row_dimensions[r].height=20
    # Total row
    r_tot = r_hrs+3+len(p1['pm_task_hrs'])
    for mc in [f"B{r_tot}:E{r_tot}",f"F{r_tot}:G{r_tot}",f"H{r_tot}:I{r_tot}",
               f"J{r_tot}:K{r_tot}",f"L{r_tot}:M{r_tot}"]:
        ws1.merge_cells(mc)
    pm_zero_total = p1['pm_task_hrs']['Zero_Hr'].sum()
    dc(ws1,"B",r_tot,"Total",MID_BG,bold=True,align="left",color=WHITE)
    dc(ws1,"F",r_tot,int(p1['pm_task_hrs']['Count'].sum()),MID_BG,color=WHITE)
    dc(ws1,"H",r_tot,f"{total_pm_hrs:.1f}",MID_BG,bold=True,color=WHITE)
    dc(ws1,"J",r_tot,f"{total_pm_hrs/p1['pm_task_hrs']['Count'].sum():.2f}" if p1['pm_task_hrs']['Count'].sum() else "—",MID_BG,color=WHITE)
    dc(ws1,"L",r_tot,f"{int(pm_zero_total)} ({int(100*pm_zero_total/p1['pm_task_hrs']['Count'].sum())}%)" if p1['pm_task_hrs']['Count'].sum() else "—",MID_BG,color=WHITE)
    ws1.row_dimensions[r_tot].height=20
    # ── Schedule Stability section (new format only) ──────────────────────────
    if p1.get('stable_pct') is not None:
        r_stab = r_tot + 2; ws1.row_dimensions[r_stab-1].height = 10
        stab_bg = GREEN_FLG if p1['stable_pct'] >= 95 else (ORANGE if p1['stable_pct'] >= 80 else RED_FLG)
        sec(ws1,"SCHEDULE STABILITY  ·  original planned date vs. executed date",r_stab-1,"M",HDR_BG)
        for col,txt,ec in [("B","Metric","E"),("F","Value","H"),("I","Assessment","M")]:
            hdr(ws1,col,r_stab,txt,ec)
        ws1.row_dimensions[r_stab].height = 18
        stab_rows = [
            ("% PMs executed on original planned date",
             f"{p1['stable_pct']:.0f}%",
             "✓ Schedule is stable" if p1['stable_pct']>=95 else "⚠ PMs being rescheduled from original plan",
             stab_bg),
            ("WOs rescheduled at least once",
             str(len(p1['rescheduled_df'])),
             "—" if not len(p1['rescheduled_df']) else f"Avg slip: {p1['avg_slip']:.0f} days",
             LIGHT_ROW if not len(p1['rescheduled_df']) else ORANGE),
            ("WOs rescheduled 2+ times",
             str(p1['multi_reschedule']),
             "Repeated reschedules indicate resource or parts constraints",
             RED_FLG if p1['multi_reschedule'] > 0 else GREEN_FLG),
        ]
        for i,(lbl,val,assess,bg) in enumerate(stab_rows):
            r = r_stab+1+i
            ws1.merge_cells(f"B{r}:E{r}"); ws1.merge_cells(f"F{r}:H{r}"); ws1.merge_cells(f"I{r}:M{r}")
            dc(ws1,"B",r,lbl,bg,align="left"); dc(ws1,"F",r,val,bg,bold=True)
            dc(ws1,"I",r,assess,bg,align="left",wrap=True); ws1.row_dimensions[r].height=22
        # Rescheduled WO detail table (if any)
        if len(p1['rescheduled_df']):
            r_rs = r_stab + 4 + 1; ws1.row_dimensions[r_rs-1].height = 10
            sec(ws1,"RESCHEDULED PMs  ·  sorted by slip days",r_rs-1,"M",ORANGE)
            for col,txt,ec in [("B","PM Task","F"),("G","Planned Date","I"),("J","Slip Days","K"),("L","# Reschedules","M")]:
                hdr(ws1,col,r_rs,txt,ec)
            ws1.row_dimensions[r_rs].height = 18
            for i,(_,row) in enumerate(p1['rescheduled_df'].head(10).iterrows()):
                r = r_rs+1+i; bg = LIGHT_ROW if i%2==0 else WHITE
                ws1.merge_cells(f"B{r}:F{r}"); ws1.merge_cells(f"G{r}:I{r}")
                ws1.merge_cells(f"J{r}:K{r}"); ws1.merge_cells(f"L{r}:M{r}")
                dc(ws1,"B",r,row['Description'],bg,align="left",wrap=True)
                dc(ws1,"G",r,row['Basic Start Date'].strftime('%Y-%m-%d') if pd.notna(row['Basic Start Date']) else '—',bg)
                dc(ws1,"J",r,int(row['Schedule_Slip_Days']),ORANGE if row['Schedule_Slip_Days']>30 else bg,bold=True)
                dc(ws1,"L",r,int(row.get('Schedule_Changes',0)),bg); ws1.row_dimensions[r].height=22

    widths(ws1,{"A":2,"B":18,"C":8,"D":8,"E":10,"F":8,"G":5,"H":10,"I":5,"J":10,"K":5,"L":14,"M":6})

    # ── SHEET 2: Pillar 2 — Failure Prevention ───────────────────────────────
    ws2 = wb.create_sheet("2 · Failure Prevention"); ws2.sheet_view.showGridLines = False
    title_row(ws2,f"PILLAR 2: FAILURE PREVENTION  ·  {asset_label}",1,"L")
    sub_row(ws2,"PM→Breakdown timing, MTBF by month with 3-month rolling average and trend direction",2,"L")
    ws2.row_dimensions[3].height=10

    for col,lbl,val,bg in [
        ("B","Overall MTBF",f"{p2['overall_mtbf']} days",rag_bg(p7['rag'])),
        ("F",f"{pl} Trend",p2['trend_dir'],rag_bg(p2['trend_clr'])),
        ("J","PM→BKD Pairs",str(len(p2['pm_bkd'])),LIGHT_ROW),
    ]:
        ec = chr(ord(col)+2)
        ws2.merge_cells(f"{col}4:{ec}4"); ws2[f"{col}4"].value=lbl
        ws2[f"{col}4"].font=Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws2[f"{col}4"].fill=wfill(WHITE); ws2[f"{col}4"].alignment=cal("left")
        ws2.merge_cells(f"{col}5:{ec}5"); ws2[f"{col}5"].value=val
        ws2[f"{col}5"].font=Font(name="Arial",bold=True,size=14,color="1F2D3D")
        ws2[f"{col}5"].fill=wfill(bg); ws2[f"{col}5"].alignment=cal(); ws2[f"{col}5"].border=bdr()
    ws2.row_dimensions[4].height=16; ws2.row_dimensions[5].height=30; ws2.row_dimensions[6].height=10

    sec(ws2,"MTBF BY MONTH WITH ROLLING 3-MONTH AVERAGE",7,"L")
    for col,txt,ec in [("B","Month",None),("D","BKDs",None),("F","MTBF (30d/BKD)",None),
                        ("H","MTBF Actual",None),("J","Rolling 3mo",None),("L","Trend","L")]:
        hdr(ws2,col,8,txt,ec)
    ws2.row_dimensions[8].height=18
    prior_val = None
    for i,(mo,bkd_c,m30,ma,r3) in enumerate(zip(
            p2['months_str'],p2['bkd_mo'],p2['mtbf_30'],p2['mtbf_actual'],p2['rolling_3'])):
        r = i+9; bg = LIGHT_ROW if i%2==0 else WHITE
        if prior_val and ma:
            delta = ma-prior_val
            trend = f"↑ +{delta:.1f}d" if delta>1 else (f"↓ {delta:.1f}d" if delta<-1 else "→")
            tbg   = GREEN_FLG if delta>1 else (RED_FLG if delta<-1 else LIGHT_ROW)
        else:
            trend = "—"; tbg = bg
        ws2.merge_cells(f"B{r}:C{r}"); ws2.merge_cells(f"D{r}:E{r}"); ws2.merge_cells(f"F{r}:G{r}")
        ws2.merge_cells(f"H{r}:I{r}"); ws2.merge_cells(f"J{r}:K{r}"); ws2.merge_cells(f"L{r}:L{r}")
        dc(ws2,"B",r,mo,bg,bold=True); dc(ws2,"D",r,bkd_c,bg)
        dc(ws2,"F",r,f"{m30}d" if m30 else "—",bg)
        dc(ws2,"H",r,f"{ma}d" if ma else "—",bg)
        dc(ws2,"J",r,f"{r3}d" if r3 else "—",bg)
        dc(ws2,"L",r,trend,tbg,bold=True); ws2.row_dimensions[r].height=18
        if ma: prior_val = ma

    # MTBF chart using hidden columns
    for i,(mo,m30,r3) in enumerate(zip(p2['months_str'],p2['mtbf_30'],p2['rolling_3'])):
        ws2[f"O{9+i}"].value = mo; ws2[f"P{9+i}"].value = m30; ws2[f"Q{9+i}"].value = r3
    ws2["P8"].value = "MTBF (30d)"; ws2["Q8"].value = "Rolling 3mo Avg"

    r_lc2 = 9 + n_mo + 1
    ws2.row_dimensions[r_lc2].height=10
    sec(ws2,f"MTBF TREND  ({pl})",r_lc2,"L",HDR_BG)
    lc = LineChart(); lc.title="MTBF Trend with 3-Month Rolling Avg"; lc.style=10
    lc.width = min(22 + max(0, n_mo-6)*2, 38); lc.height=12
    lc.y_axis.title="Days Between Breakdowns"
    lc.add_data(Reference(ws2,min_col=16,min_row=8,max_row=8+n_mo),titles_from_data=True)
    lc.add_data(Reference(ws2,min_col=17,min_row=8,max_row=8+n_mo),titles_from_data=True)
    lc.set_categories(Reference(ws2,min_col=15,min_row=9,max_row=8+n_mo))
    lc.series[0].graphicalProperties.line.solidFill="2E75B6"; lc.series[0].graphicalProperties.line.width=28575
    lc.series[0].marker.symbol="circle"; lc.series[0].marker.size=7
    lc.series[1].graphicalProperties.line.solidFill="ED7D31"; lc.series[1].graphicalProperties.line.width=19050
    lc.visible_cells_only = False
    ws2.add_chart(lc,f"B{r_lc2+1}")
    for col in ["O","P","Q"]: ws2.column_dimensions[col].hidden = True

    # PM→BKD summary — chart is 12cm ≈ 19 rows; add clearance
    r_pb = max(44, r_lc2 + 22); ws2.row_dimensions[r_pb].height=10
    sec(ws2,"PM → BREAKDOWN TIMING SUMMARY",r_pb+1,"L",HDR_BG)
    cat_clr = {"Same day": GREEN_FLG,"1-3 days": RED_FLG,"4-7 days": ORANGE,"8+ days": LIGHT_ROW}
    if len(p2['pm_bkd']):
        vc2 = p2['pm_bkd']['Category'].value_counts()
        for col,txt,ec in [("B","Timing Bucket",None),("E","Count",None),("G","% of BKDs",None),
                            ("I","Interpretation","L")]:
            hdr(ws2,col,r_pb+2,txt,ec)
        ws2.row_dimensions[r_pb+2].height=18
        buckets = [
            ("Same day","PM likely surfaced failure during inspection"),
            ("1-3 days","Immediate failure after PM — latent defect or component disturbed"),
            ("4-7 days","Short-interval failure — PM may not address root cause"),
            ("8+ days","Failure outside normal PM interval"),
        ]
        for i,(cat,interp) in enumerate(buckets):
            r = r_pb+3+i; ct = vc2.get(cat,0); bg = cat_clr.get(cat,LIGHT_ROW)
            ws2.merge_cells(f"B{r}:D{r}"); ws2.merge_cells(f"E{r}:F{r}"); ws2.merge_cells(f"G{r}:H{r}"); ws2.merge_cells(f"I{r}:L{r}")
            dc(ws2,"B",r,cat,bg,bold=True); dc(ws2,"E",r,ct,bg)
            dc(ws2,"G",r,f"{100*ct/len(p2['pm_bkd']):.0f}%",bg)
            dc(ws2,"I",r,interp,bg,align="left",wrap=True); ws2.row_dimensions[r].height=22
    # PM → Breakdown Correlation
    r_corr = r_pb + 8; ws2.row_dimensions[r_corr].height=10
    sec(ws2,"PM → BREAKDOWN CORRELATION ANALYSIS",r_corr+1,"L",HDR_BG)
    corr_bg = GREEN_FLG if p2['corr_coef'] < -0.3 else (RED_FLG if p2['corr_coef'] > 0.3 else ORANGE)
    ws2.merge_cells(f"B{r_corr+2}:E{r_corr+2}"); ws2.merge_cells(f"F{r_corr+2}:L{r_corr+2}")
    dc(ws2,"B",r_corr+2,"Pearson r (monthly PM count vs BKD count)",HDR_BG,align="left",bold=True)
    dc(ws2,"F",r_corr+2,f"{p2['corr_coef']:+.2f}  —  {p2['corr_interp']}",corr_bg,align="left")
    ws2.row_dimensions[r_corr+2].height=22
    if len(p2['task_corr']):
        hdr_r = r_corr+3
        for col,txt,ec in [("B","PM Task","F"),("G","PM Count","H"),("I","BKDs w/in 7d","J"),("K","BKD Rate %","L")]:
            hdr(ws2,col,hdr_r,txt,ec)
        ws2.row_dimensions[hdr_r].height=18
        for i,(_,row) in enumerate(p2['task_corr'].iterrows()):
            r = hdr_r+1+i
            bg = RED_FLG if row['BKD_Rate_Pct']>=50 else (ORANGE if row['BKD_Rate_Pct']>=25 else LIGHT_ROW)
            ws2.merge_cells(f"B{r}:F{r}"); ws2.merge_cells(f"G{r}:H{r}")
            ws2.merge_cells(f"I{r}:J{r}"); ws2.merge_cells(f"K{r}:L{r}")
            dc(ws2,"B",r,row['PM_Task'],bg,align="left",wrap=True)
            dc(ws2,"G",r,int(row['PM_Count']),bg)
            dc(ws2,"I",r,int(row['BKD_Within_7d']),bg)
            dc(ws2,"K",r,f"{int(row['BKD_Rate_Pct'])}%",bg,bold=True)
            ws2.row_dimensions[r].height=30
    ws2.freeze_panes = "B9"
    widths(ws2,{"A":2,"B":14,"C":4,"D":8,"E":8,"F":4,"G":14,"H":4,"I":28,"J":10,"K":4,"L":14})

    # ── SHEET 3: Pillar 3 — Failure Mode Coverage ─────────────────────────────
    ws3 = wb.create_sheet("3 · Coverage Gap"); ws3.sheet_view.showGridLines = False
    title_row(ws3,f"PILLAR 3: FAILURE MODE COVERAGE  ·  {asset_label}",1,"L")
    sub_row(ws3,"Each breakdown cluster checked against PM task descriptions. GAP = breakdown cluster exists but no PM task addresses it.",2,"L")
    ws3.row_dimensions[3].height=10

    for col,lbl,val,bg in [
        ("B","Coverage Gaps",str(p3['gaps']),rag_bg("GREEN" if p3['gaps']==0 else "RED")),
        ("F","High Priority Gaps",str(p3['high_gaps']),rag_bg("GREEN" if p3['high_gaps']==0 else "RED")),
        ("J","Clusters Analyzed",str(len(p3['cov_df'])),LIGHT_ROW),
    ]:
        ec = chr(ord(col)+2)
        ws3.merge_cells(f"{col}4:{ec}4"); ws3[f"{col}4"].value=lbl
        ws3[f"{col}4"].font=Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws3[f"{col}4"].fill=wfill(WHITE); ws3[f"{col}4"].alignment=cal("left")
        ws3.merge_cells(f"{col}5:{ec}5"); ws3[f"{col}5"].value=val
        ws3[f"{col}5"].font=Font(name="Arial",bold=True,size=16,color="1F2D3D")
        ws3[f"{col}5"].fill=wfill(bg); ws3[f"{col}5"].alignment=cal(); ws3[f"{col}5"].border=bdr()
    ws3.row_dimensions[4].height=16; ws3.row_dimensions[5].height=30; ws3.row_dimensions[6].height=10

    sec(ws3,"FAILURE CLUSTER — PM COVERAGE CHECK",7,"L")
    for col,txt,ec in [("B","Failure Cluster","C"),("D","BKD WOs","E"),("F","PM Covered?","G"),
                        ("H","Status","I"),("J","Priority","K"),("L","Recommended Action","L")]:
        hdr(ws3,col,8,txt,ec)
    ws3.row_dimensions[8].height=18
    cov_sorted = p3['cov_df'].sort_values('BKD_Count',ascending=False)
    for i,(_,row) in enumerate(cov_sorted.iterrows()):
        r = i+9
        bg = RED_FLG if row['Gap'] else (GREEN_FLG if row['BKD_Count']>0 else LIGHT_ROW)
        action = (f"Add PM task targeting {row['Cluster'].lower()} components" if row['Gap']
                  else ("PM coverage confirmed" if row['BKD_Count']>0 else "No breakdowns — continue monitoring"))
        for mc in [f"B{r}:C{r}",f"D{r}:E{r}",f"F{r}:G{r}",f"H{r}:I{r}",f"J{r}:K{r}"]:
            ws3.merge_cells(mc)
        dc(ws3,"B",r,row['Cluster'],bg,bold=True,align="left")
        dc(ws3,"D",r,row['BKD_Count'],bg)
        dc(ws3,"F",r,"✓ Yes" if row['PM_Covered'] else "✗ No",
           GREEN_FLG if row['PM_Covered'] else RED_FLG,bold=True)
        dc(ws3,"H",r,row['Status'],bg)
        dc(ws3,"J",r,row['Priority'],RED_FLG if row['Priority']=="HIGH" else (ORANGE if row['Priority']=="MED" else GREEN_FLG),bold=True)
        dc(ws3,"L",r,action,bg,align="left",wrap=True); ws3.row_dimensions[r].height=26
    ws3.freeze_panes = "B9"
    widths(ws3,{"A":2,"B":8,"C":20,"D":8,"E":6,"F":8,"G":6,"H":18,"I":4,"J":8,"K":6,"L":30})

    # ── SHEET 4: Pillar 4 — Work Order Quality ─────────────────────────────
    ws4 = wb.create_sheet("4 · WO Quality"); ws4.sheet_view.showGridLines = False
    title_row(ws4,f"PILLAR 4: WORK ORDER QUALITY  ·  {asset_label}",1,"L")
    sub_row(ws4,"Zero-hour analysis by WO type. Planned vs. actual hours variance (if planned hours available in export).",2,"L")
    ws4.row_dimensions[3].height=10

    for col,lbl,val,bg in [
        ("B","PM Zero-Hour %",f"{p4['pm_zero_pct']:.0f}%",rag_bg("GREEN" if p4['pm_zero_pct']==0 else "AMBER")),
        ("F","BKD Zero-Hour %",f"{p4['bkd_zero_pct']:.0f}%",rag_bg("GREEN" if p4['bkd_zero_pct']<20 else "AMBER" if p4['bkd_zero_pct']<60 else "RED")),
        ("J","Follow-up Corr Zero-Hr %",f"{p4['fuc_zero_pct']:.0f}%",ORANGE),
    ]:
        ec = chr(ord(col)+2)
        ws4.merge_cells(f"{col}4:{ec}4"); ws4[f"{col}4"].value=lbl
        ws4[f"{col}4"].font=Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws4[f"{col}4"].fill=wfill(WHITE); ws4[f"{col}4"].alignment=cal("left")
        ws4.merge_cells(f"{col}5:{ec}5"); ws4[f"{col}5"].value=val
        ws4[f"{col}5"].font=Font(name="Arial",bold=True,size=16,color="1F2D3D")
        ws4[f"{col}5"].fill=wfill(bg); ws4[f"{col}5"].alignment=cal(); ws4[f"{col}5"].border=bdr()
    ws4.row_dimensions[4].height=16; ws4.row_dimensions[5].height=30; ws4.row_dimensions[6].height=10

    sec(ws4,"ZERO-HOUR WO SUMMARY BY TYPE",7,"L")
    for col,txt,ec in [("B","WO Type","D"),("E","Total WOs","F"),("G","Zero-Hour","H"),("I","Zero-Hr %","J"),
                        ("K","Assessment","L")]:
        hdr(ws4,col,8,txt,ec)
    ws4.row_dimensions[8].height=18
    zh_rows = [
        ("PM (GM09)",      len(pm),  p4['pm_zero'],  p4['pm_zero_pct'],
         "✓ Good — all PMs have hours" if p4['pm_zero']==0 else f"⚠ {p4['pm_zero']} PMs missing hours",
         GREEN_FLG if p4['pm_zero']==0 else YELLOW),
        ("Breakdown (GM02)",len(bkd), p4['bkd_zero'], p4['bkd_zero_pct'],
         "⚠ Labor not captured — true failure cost unknown", YELLOW),
        ("Follow-up (GM01/PM01)",len(fuc),p4['fuc_zero'],p4['fuc_zero_pct'],
         "⚠ Separate parts-only from labor WOs; audit non-parts zeros", ORANGE),
    ]
    for i,(wt,tot,zh,zhp,assess,bg) in enumerate(zh_rows):
        r = i+9
        for mc in [f"B{r}:D{r}",f"E{r}:F{r}",f"G{r}:H{r}",f"I{r}:J{r}",f"K{r}:L{r}"]:
            ws4.merge_cells(mc)
        dc(ws4,"B",r,wt,bg,bold=True,align="left"); dc(ws4,"E",r,tot,bg)
        dc(ws4,"G",r,zh,bg); dc(ws4,"I",r,f"{zhp:.1f}%",bg,bold=True)
        dc(ws4,"K",r,assess,bg,align="left",wrap=True); ws4.row_dimensions[r].height=32

    if p4['has_planned']:
        ws4.row_dimensions[13].height=10
        sec(ws4,"PLANNED vs. ACTUAL HOURS — PM WOs",14,"L")
        for col,txt,ec in [("B","Metric","D"),("E","Value","G")]:
            hdr(ws4,col,15,txt,ec)
        ws4.row_dimensions[15].height=18
        pva_rows = [
            ("PM WOs compared",       len(p4['pva']),  LIGHT_ROW),
            ("Avg variance %",         f"{p4['avg_var_pct']:+.1f}%", ORANGE if abs(p4['avg_var_pct'])>20 else LIGHT_ROW),
            ("Over estimate (actual > planned)", p4['over_cnt'],  RED_FLG if p4['over_cnt']>3 else LIGHT_ROW),
            ("Under estimate (actual < planned)",p4['under_cnt'], LIGHT_ROW),
            ("On estimate (actual = planned)",   p4['on_cnt'],    GREEN_FLG),
        ]
        for i,(lbl,val,bg) in enumerate(pva_rows):
            r = 16+i
            ws4.merge_cells(f"B{r}:D{r}"); ws4.merge_cells(f"E{r}:G{r}")
            dc(ws4,"B",r,lbl,bg,align="left"); dc(ws4,"E",r,str(val),bg,bold=True)
            ws4.row_dimensions[r].height=18
    else:
        ws4.row_dimensions[13].height=10
        sec(ws4,"PLANNED vs. ACTUAL HOURS",14,"L",HDR_BG)
        ws4.merge_cells("B15:L15")
        dc(ws4,"B",15,
           "Planned Hours column not found in Celonis export. Add 'Planned Hours' to the export to enable this analysis.",
           YELLOW,align="left",wrap=True); ws4.row_dimensions[15].height=28
    ws4.freeze_panes = "B9"
    widths(ws4,{"A":2,"B":8,"C":8,"D":22,"E":10,"F":8,"G":10,"H":4,"I":10,"J":4,"K":32,"L":2})

    # ── SHEET 5: Pillar 5 — Backlog Health ───────────────────────────────────
    ws5 = wb.create_sheet("5 · Backlog Health"); ws5.sheet_view.showGridLines = False
    title_row(ws5,f"PILLAR 5: BACKLOG HEALTH  ·  {asset_label}",1,"L")
    sub_row(ws5,f"Open WOs (no TECO/CLSD) aged from Created On to ref date {ref_date.strftime('%Y-%m-%d')}. Monthly new vs. closed trend.",2,"L")
    ws5.row_dimensions[3].height=10

    for col,lbl,val,bg in [
        ("B","Total Open WOs",str(p5['total_open']),rag_bg("GREEN" if p5['total_open']<15 else "AMBER" if p5['total_open']<30 else "RED")),
        ("F","Aged 90+ Days",str(p5['age_91p']),rag_bg("GREEN" if p5['age_91p']==0 else "AMBER" if p5['age_91p']<=5 else "RED")),
        ("J","Avg Age",f"{p5['avg_age']:.0f} days",LIGHT_ROW),
    ]:
        ec = chr(ord(col)+2)
        ws5.merge_cells(f"{col}4:{ec}4"); ws5[f"{col}4"].value=lbl
        ws5[f"{col}4"].font=Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws5[f"{col}4"].fill=wfill(WHITE); ws5[f"{col}4"].alignment=cal("left")
        ws5.merge_cells(f"{col}5:{ec}5"); ws5[f"{col}5"].value=val
        ws5[f"{col}5"].font=Font(name="Arial",bold=True,size=16,color="1F2D3D")
        ws5[f"{col}5"].fill=wfill(bg); ws5[f"{col}5"].alignment=cal(); ws5[f"{col}5"].border=bdr()
    ws5.row_dimensions[4].height=16; ws5.row_dimensions[5].height=30; ws5.row_dimensions[6].height=10

    sec(ws5,"BACKLOG AGING BUCKETS",7,"L")
    for col,txt,ec in [("B","Age Bucket","D"),("E","Count","F"),("G","% of Open","H"),("I","Assessment","L")]:
        hdr(ws5,col,8,txt,ec)
    ws5.row_dimensions[8].height=18
    aging = [
        ("0 – 30 days",  p5['age_0_30'],  GREEN_FLG, "Current — monitor"),
        ("31 – 90 days", p5['age_31_90'], ORANGE,    "Aging — review priority and assign completion date"),
        ("91+ days",     p5['age_91p'],   RED_FLG,   "Stale — assess validity; close or escalate"),
    ]
    for i,(lbl,ct,bg,assess) in enumerate(aging):
        r = i+9
        ws5.merge_cells(f"B{r}:D{r}"); ws5.merge_cells(f"E{r}:F{r}"); ws5.merge_cells(f"G{r}:H{r}"); ws5.merge_cells(f"I{r}:L{r}")
        dc(ws5,"B",r,lbl,bg,bold=True); dc(ws5,"E",r,ct,bg)
        dc(ws5,"G",r,f"{100*ct/p5['total_open']:.0f}%" if p5['total_open'] else "—",bg)
        dc(ws5,"I",r,assess,bg,align="left",wrap=True); ws5.row_dimensions[r].height=22

    sec(ws5,"OPEN WOs BY TYPE",13,"L",HDR_BG)
    for col,txt,ec in [("B","Order Type","D"),("E","Count","F"),("G","Avg Age (d)","H"),("I","Max Age (d)","J")]:
        hdr(ws5,col,14,txt,ec)
    ws5.row_dimensions[14].height=18
    for i,(_,row) in enumerate(p5['by_type'].sort_values('Count',ascending=False).iterrows()):
        r = 15+i; bg = LIGHT_ROW if i%2==0 else WHITE
        ws5.merge_cells(f"B{r}:D{r}"); ws5.merge_cells(f"E{r}:F{r}"); ws5.merge_cells(f"G{r}:H{r}"); ws5.merge_cells(f"I{r}:J{r}")
        dc(ws5,"B",r,row['OrderType'],bg,bold=True); dc(ws5,"E",r,int(row['Count']),bg)
        dc(ws5,"G",r,f"{row['Avg_Age']:.0f}",bg); dc(ws5,"I",r,int(row['Max_Age']),bg)
        ws5.row_dimensions[r].height=18

    r_ch = 15 + len(p5['by_type']) + 2
    ws5.row_dimensions[r_ch].height=10
    sec(ws5,f"MONTHLY NEW vs. CLOSED WOs ({pl})",r_ch+1,"L",HDR_BG)
    for i,(mo,nw,cl,nt) in enumerate(zip(p5['months_str'],p5['new_mo'],p5['closed_mo'],p5['net_mo'])):
        ws5[f"O{r_ch+2+i}"].value = mo
        ws5[f"P{r_ch+2+i}"].value = nw
        ws5[f"Q{r_ch+2+i}"].value = cl
    ws5[f"O{r_ch+1}"].value = "Month"
    ws5[f"P{r_ch+1}"].value = "New WOs"
    ws5[f"Q{r_ch+1}"].value = "Closed (TECO)"
    bc5 = BarChart(); bc5.type="col"; bc5.grouping="clustered"
    bc5.title=f"Monthly New vs. TECO'd WOs ({pl})"; bc5.style=10
    bc5.width = min(22 + max(0, n_mo-6)*2, 38); bc5.height=12
    bc5.add_data(Reference(ws5,min_col=16,min_row=r_ch+1,max_row=r_ch+1+n_mo),titles_from_data=True)
    bc5.add_data(Reference(ws5,min_col=17,min_row=r_ch+1,max_row=r_ch+1+n_mo),titles_from_data=True)
    bc5.set_categories(Reference(ws5,min_col=15,min_row=r_ch+2,max_row=r_ch+1+n_mo))
    bc5.series[0].graphicalProperties.solidFill="2E75B6"
    bc5.series[1].graphicalProperties.solidFill="375623"
    bc5.visible_cells_only = False
    from openpyxl.chart.legend import Legend as _Legend
    bc5.legend = _Legend(); bc5.legend.legendPos = 'b'
    ws5.add_chart(bc5,f"B{r_ch+2}")
    for col in ["O","P","Q"]: ws5.column_dimensions[col].hidden = True

    # ── Due Date Urgency section (new format only) — placed below chart ───────
    if p5.get('overdue_ct', 0) > 0 or p5.get('due_0_7', 0) > 0 or p5.get('due_8_30', 0) > 0:
        r_due = r_ch + 25; ws5.row_dimensions[r_due-1].height = 10
        urg_hdr_bg = RAG_RED if p5['overdue_ct'] > 0 else MID_BG
        sec(ws5,f"DUE DATE URGENCY  ·  {p5['overdue_ct']} WOs overdue",r_due-1,"L",urg_hdr_bg)
        for col,txt,ec in [("B","Status","D"),("E","Count","F"),("G","Assessment","L")]:
            hdr(ws5,col,r_due,txt,ec)
        ws5.row_dimensions[r_due].height = 18
        urgency_rows = [
            ("Overdue (past due date)",   p5['overdue_ct'], RED_FLG,   "Immediate action required — review and close or escalate"),
            ("Due within 7 days",         p5['due_0_7'],    ORANGE,    "Assign resources now to avoid becoming overdue"),
            ("Due in 8–30 days",          p5['due_8_30'],   LIGHT_ROW, "Plan execution — confirm parts and resource availability"),
        ]
        for i,(lbl,ct,bg,assess) in enumerate(urgency_rows):
            r = r_due+1+i
            ws5.merge_cells(f"B{r}:D{r}"); ws5.merge_cells(f"E{r}:F{r}"); ws5.merge_cells(f"G{r}:L{r}")
            dc(ws5,"B",r,lbl,bg,bold=True,align="left")
            dc(ws5,"E",r,ct,bg,bold=True)
            dc(ws5,"G",r,assess,bg,align="left",wrap=True); ws5.row_dimensions[r].height=22
        # Overdue WO detail
        if len(p5.get('overdue_df', pd.DataFrame())):
            r_ov = r_due + 5; ws5.row_dimensions[r_ov-1].height = 10
            sec(ws5,"OVERDUE WOs  ·  most overdue first",r_ov-1,"L",RAG_RED)
            cols5 = list(p5['overdue_df'].columns)
            status_col = cols5[3] if len(cols5) > 3 else None
            for col,txt,ec in [("B","WO#","C"),("D","Description","H"),("I","Days Overdue","J"),("K","Status","L")]:
                hdr(ws5,col,r_ov,txt,ec)
            ws5.row_dimensions[r_ov].height = 18
            for i,(_,row) in enumerate(p5['overdue_df'].iterrows()):
                r = r_ov+1+i; bg = RED_FLG if i < 5 else LIGHT_ROW
                ws5.merge_cells(f"B{r}:C{r}"); ws5.merge_cells(f"D{r}:H{r}")
                ws5.merge_cells(f"I{r}:J{r}"); ws5.merge_cells(f"K{r}:L{r}")
                dc(ws5,"B",r,str(row.iloc[0]),bg)
                dc(ws5,"D",r,str(row.iloc[1]),bg,align="left",wrap=True)
                dc(ws5,"I",r,f"{abs(int(row.iloc[2]))}d",bg,bold=True)
                dc(ws5,"K",r,str(row.iloc[3]) if len(row) > 3 else "—",bg,align="left")
                ws5.row_dimensions[r].height=22

    ws5.freeze_panes = "B8"
    widths(ws5,{"A":2,"B":10,"C":8,"D":20,"E":8,"F":6,"G":12,"H":4,"I":12,"J":4,"K":10,"L":2})

    # ── SHEET 6: Pillar 6 — Cost Visibility ──────────────────────────────────
    ws6 = wb.create_sheet("6 · Cost Visibility"); ws6.sheet_view.showGridLines = False
    title_row(ws6,f"PILLAR 6: COST VISIBILITY  ·  {asset_label}",1,"L")
    cost_sub = ("Actual cost by WO type. Cost variance vs. plan. ⚠ Labor hours flagged separately — required for MTTR and productivity analysis."
                if p6['has_cost'] else
                "No cost export provided. Upload cost file to enable cost analysis. Labor hour capture shown below.")
    sub_row(ws6, cost_sub, 2, "L")
    ws6.row_dimensions[3].height=10

    if p6['has_cost']:
        kpi3 = [
            ("B", "Total Spend",       f"${p6['total_cost']:,.0f}",   LIGHT_ROW),
            ("F", "BKD vs PM Avg Cost",f"${p6['avg_bkd_cost']:,.0f} vs ${p6['avg_pm_cost']:,.0f}",
             rag_bg("AMBER")),
            ("J", "WOs Over Plan",     f"{p6['exceeds']} ({p6['exceed_pct']:.0f}%)",
             rag_bg("GREEN" if p6['exceed_pct']<15 else "AMBER" if p6['exceed_pct']<30 else "RED")),
        ]
    else:
        kpi3 = [
            ("B", "Cost File",         "Not uploaded",     YELLOW),
            ("F", "BKD Zero-Hour %",   f"{p6['bkd_zero_pct']:.0f}%",
             rag_bg("RED" if p6['bkd_zero_pct']>50 else "AMBER")),
            ("J", "PM Zero-Hour %",    f"{p6['pm_zero']} WOs",
             rag_bg("GREEN" if p6['pm_zero']==0 else "AMBER")),
        ]
    for col,lbl,val,bg in kpi3:
        ec = chr(ord(col)+2)
        ws6.merge_cells(f"{col}4:{ec}4"); ws6[f"{col}4"].value=lbl
        ws6[f"{col}4"].font=Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws6[f"{col}4"].fill=wfill(WHITE); ws6[f"{col}4"].alignment=cal("left")
        ws6.merge_cells(f"{col}5:{ec}5"); ws6[f"{col}5"].value=val
        ws6[f"{col}5"].font=Font(name="Arial",bold=True,size=14,color="1F2D3D")
        ws6[f"{col}5"].fill=wfill(bg); ws6[f"{col}5"].alignment=cal(); ws6[f"{col}5"].border=bdr()
    ws6.row_dimensions[4].height=16; ws6.row_dimensions[5].height=30; ws6.row_dimensions[6].height=10

    if p6['has_cost']:
        # ── Cost by WO type
        sec(ws6,"COST BY WORK ORDER TYPE",7,"L")
        for col,txt,ec in [("B","WO Type","D"),("E","Count","F"),("G","Total Cost","H"),
                            ("I","Avg Cost / WO","J"),("K","% of Spend","L")]:
            hdr(ws6,col,8,txt,ec)
        ws6.row_dimensions[8].height=18
        type_rows = [
            ("PM (GM09)",             len(pm),  p6['pm_cost'],  p6['avg_pm_cost'],  GREEN_FLG),
            ("Breakdown (GM02)",      len(bkd), p6['bkd_cost'], p6['avg_bkd_cost'], RED_FLG),
            ("Follow-up (GM01/PM01)", len(fuc), p6['fuc_cost'], p6['fuc_cost']/len(fuc) if len(fuc) else 0, ORANGE),
        ]
        for i,(wt,cnt,tc,ac,bg) in enumerate(type_rows):
            r = i+9
            for mc in [f"B{r}:D{r}",f"E{r}:F{r}",f"G{r}:H{r}",f"I{r}:J{r}",f"K{r}:L{r}"]:
                ws6.merge_cells(mc)
            dc(ws6,"B",r,wt,bg,bold=True,align="left"); dc(ws6,"E",r,cnt,bg)
            dc(ws6,"G",r,f"${tc:,.0f}",bg,bold=True)
            dc(ws6,"I",r,f"${ac:,.0f}",bg)
            dc(ws6,"K",r,f"{100*tc/p6['total_cost']:.0f}%" if p6['total_cost'] else "—",bg)
            ws6.row_dimensions[r].height=22
        # Total row
        r=12
        for mc in [f"B{r}:D{r}",f"E{r}:F{r}",f"G{r}:H{r}",f"I{r}:J{r}",f"K{r}:L{r}"]:
            ws6.merge_cells(mc)
        dc(ws6,"B",r,"Total",MID_BG,bold=True,align="left",color=WHITE)
        dc(ws6,"E",r,len(pm)+len(bkd)+len(fuc),MID_BG,color=WHITE)
        dc(ws6,"G",r,f"${p6['total_cost']:,.0f}",MID_BG,bold=True,color=WHITE)
        dc(ws6,"I",r,f"${p6['total_cost']/(len(pm)+len(bkd)+len(fuc)):,.0f}" if (len(pm)+len(bkd)+len(fuc)) else "—",MID_BG,color=WHITE)
        dc(ws6,"K",r,"100%",MID_BG,bold=True,color=WHITE); ws6.row_dimensions[r].height=22

        # ── Cost variance
        ws6.row_dimensions[13].height=10
        sec(ws6,"COST VARIANCE vs. PLAN",14,"L",HDR_BG)
        for col,txt,ec in [("B","Classification","F"),("G","Count","H"),("I","% of WOs","J"),("K","Assessment","L")]:
            hdr(ws6,col,15,txt,ec)
        ws6.row_dimensions[15].height=18
        var_rows = [
            ("Within Boundaries",                p6['within'], GREEN_FLG, "On track"),
            ("Actual Cost is Lower than Planned", p6['below'],  GREEN_FLG, "Under plan — verify scope was completed"),
            ("Actual Cost Exceeds Planned Cost",  p6['exceeds'],RED_FLG,  "Review WOs — scope creep or poor estimation"),
        ]
        tot_wos = p6['within'] + p6['below'] + p6['exceeds']
        for i,(lbl,ct,bg,assess) in enumerate(var_rows):
            r = 16+i
            for mc in [f"B{r}:F{r}",f"G{r}:H{r}",f"I{r}:J{r}",f"K{r}:L{r}"]:
                ws6.merge_cells(mc)
            dc(ws6,"B",r,lbl,bg,bold=True,align="left"); dc(ws6,"G",r,ct,bg)
            dc(ws6,"I",r,f"{100*ct/tot_wos:.0f}%" if tot_wos else "—",bg)
            dc(ws6,"K",r,assess,bg,align="left",wrap=True); ws6.row_dimensions[r].height=22

        # ── Monthly cost chart data + chart
        ws6.row_dimensions[20].height=10
        sec(ws6,f"MONTHLY SPEND TREND  ({pl})",21,"L",HDR_BG)
        for i,(mo,cost) in enumerate(zip(p6['months_str'],p6['cost_mo'])):
            ws6[f"O{22+i}"].value = mo; ws6[f"P{22+i}"].value = int(cost)
        ws6["P21"].value = "Cost ($)"
        bc6 = BarChart(); bc6.type="col"; bc6.title=f"Monthly Spend ({pl})"
        bc6.y_axis.title="Cost ($)"; bc6.style=10
        bc6.width = min(22 + max(0, n_mo-6)*2, 38); bc6.height=11
        bc6.add_data(Reference(ws6,min_col=16,min_row=21,max_row=21+n_mo),titles_from_data=True)
        bc6.set_categories(Reference(ws6,min_col=15,min_row=22,max_row=21+n_mo))
        _BAR_PALETTE = ["2E75B6","4472C4","5B9BD5","70AD47","ED7D31","C00000","A9D18E","F4B942","833C11","0070C0","843C0C","595959"]
        for _idx in range(n_mo):
            _pt = DataPoint(idx=_idx); _pt.graphicalProperties.solidFill = _BAR_PALETTE[_idx % len(_BAR_PALETTE)]
            bc6.series[0].dPt.append(_pt)
        bc6.visible_cells_only = False
        ws6.add_chart(bc6,"B22")
        for col in ["O","P"]: ws6.column_dimensions[col].hidden = True

        # ── Top cost WOs
        r_top = 36; ws6.row_dimensions[r_top].height=10
        sec(ws6,f"TOP {min(10,len(p6['top_wos']))} WOs BY COST",r_top+1,"L",HDR_BG)
        for col,txt,ec in [("B","WO Number","C"),("D","Description","H"),
                            ("I","Cost","J"),("K","vs. Plan","L")]:
            hdr(ws6,col,r_top+2,txt,ec)
        ws6.row_dimensions[r_top+2].height=18
        for i,(_,row) in enumerate(p6['top_wos'].iterrows()):
            r = r_top+3+i
            classification = str(row['Actual Cost Classification']).strip()
            bg = RED_FLG if 'Exceeds' in classification else LIGHT_ROW
            for mc in [f"B{r}:C{r}",f"D{r}:H{r}",f"I{r}:J{r}",f"K{r}:L{r}"]:
                ws6.merge_cells(mc)
            dc(ws6,"B",r,str(row['Work Order Number']),bg)
            dc(ws6,"D",r,row['Description'],bg,align="left")
            dc(ws6,"I",r,f"${int(row['Cost per Work Order']):,}",bg,bold=True)
            dc(ws6,"K",r,classification,bg,align="left",wrap=True)
            ws6.row_dimensions[r].height=18

        next_r = r_top + 3 + len(p6['top_wos']) + 1
    else:
        next_r = 8

    # ── LABOR HOUR FLAG (always shown) ────────────────────────────────────────
    ws6.row_dimensions[next_r].height=10
    sec(ws6,"⚠  LABOR HOUR CAPTURE — ACTION REQUIRED",next_r+1,"L",RAG_RED)
    ws6.merge_cells(f"B{next_r+2}:L{next_r+2}")
    flag_msg = (f"Even with cost data available, labor HOURS are not being recorded on "
                f"{p6['bkd_zero']}/{len(bkd)} Breakdown WOs ({p6['bkd_zero_pct']:.0f}%). "
                f"Hours are required for MTTR calculation, technician productivity analysis, "
                f"and scheduling. Cost alone does not show how long repairs took."
                if p6['has_cost'] else
                f"{p6['bkd_zero']}/{len(bkd)} Breakdown WOs ({p6['bkd_zero_pct']:.0f}%) have zero actual hours. "
                f"Labor cost of failures cannot be assessed. MTTR cannot be calculated. "
                f"Upload cost export to see dollar impact. Require technicians to book time on all GM02 WOs.")
    dc(ws6,"B",next_r+2,flag_msg,RED_FLG,align="left",wrap=True,bold=True)
    ws6.row_dimensions[next_r+2].height=60

    for col,txt,ec in [("B","WO Type","D"),("E","Zero-Hr WOs","F"),("G","Zero-Hr %","H"),("I","Impact","L")]:
        hdr(ws6,col,next_r+3,txt,ec)
    ws6.row_dimensions[next_r+3].height=18
    labor_rows = [
        ("PM (GM09)",             p6['pm_zero'],   100*p6['pm_zero']/len(pm)   if len(pm)  else 0,
         "✓ Good — PMs have hours" if p6['pm_zero']==0 else "Investigate missing PM hours", GREEN_FLG if p6['pm_zero']==0 else YELLOW),
        ("Breakdown (GM02)",      p6['bkd_zero'],  p6['bkd_zero_pct'],
         "MTTR unknown. Repair duration not tracked. Technician time invisible.", YELLOW),
        ("Follow-up (GM01/PM01)", p6['fuc_zero'],  p6['fuc_zero_pct'],
         "Separate parts-only WOs from labor WOs; audit remaining zero-hour entries.", ORANGE),
    ]
    for i,(wt,zh,zhp,impact,bg) in enumerate(labor_rows):
        r = next_r+4+i
        for mc in [f"B{r}:D{r}",f"E{r}:F{r}",f"G{r}:H{r}",f"I{r}:L{r}"]:
            ws6.merge_cells(mc)
        dc(ws6,"B",r,wt,bg,bold=True,align="left"); dc(ws6,"E",r,zh,bg)
        dc(ws6,"G",r,f"{zhp:.0f}%",bg,bold=True)
        dc(ws6,"I",r,impact,bg,align="left",wrap=True); ws6.row_dimensions[r].height=28

    ws6.freeze_panes = "B8"
    widths(ws6,{"A":2,"B":8,"C":8,"D":26,"E":8,"F":6,"G":8,"H":4,"I":12,"J":8,"K":22,"L":2})

    # ── SHEET 7: Pillar 7 — Equipment Reliability ─────────────────────────────
    ws7 = wb.create_sheet("7 · Reliability"); ws7.sheet_view.showGridLines = False
    title_row(ws7,f"PILLAR 7: EQUIPMENT RELIABILITY  ·  {asset_label}",1,"L")
    sub_row(ws7,f"MTBF (Mean Time Between Failures) and MTTR (Mean Time to Repair) by month. {pl}.",2,"L")
    ws7.row_dimensions[3].height=10

    for col,lbl,val,bg in [
        ("B","MTBF (Overall)",f"{p7['mtbf_overall']} days",rag_bg("GREEN" if p7['mtbf_overall']>7 else "AMBER" if p7['mtbf_overall']>3 else "RED")),
        ("F","MTTR",f"{p7['mttr']} hrs" if p7['mttr'] else "N/A",rag_bg("AMBER") if not p7['mttr'] else LIGHT_ROW),
        ("J",f"Total Breakdowns ({pl})",str(sum(p7['bkd_mo'])),LIGHT_ROW),
    ]:
        ec = chr(ord(col)+2)
        ws7.merge_cells(f"{col}4:{ec}4"); ws7[f"{col}4"].value=lbl
        ws7[f"{col}4"].font=Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws7[f"{col}4"].fill=wfill(WHITE); ws7[f"{col}4"].alignment=cal("left")
        ws7.merge_cells(f"{col}5:{ec}5"); ws7[f"{col}5"].value=val
        ws7[f"{col}5"].font=Font(name="Arial",bold=True,size=16,color="1F2D3D")
        ws7[f"{col}5"].fill=wfill(bg); ws7[f"{col}5"].alignment=cal(); ws7[f"{col}5"].border=bdr()
    ws7.row_dimensions[4].height=16; ws7.row_dimensions[5].height=30; ws7.row_dimensions[6].height=10

    sec(ws7,f"MONTHLY RELIABILITY METRICS  ({pl})",7,"L")
    for col,txt,ec in [("B","Month","C"),("D","BKDs","E"),("F","MTBF (30d/BKD)","G"),
                        ("H","MTTR (hrs)","I"),("J","Best/Worst","L")]:
        hdr(ws7,col,8,txt,ec)
    ws7.row_dimensions[8].height=18
    for i,(mo,bkd_c,m30,mt) in enumerate(zip(p7['months_str'],p7['bkd_mo'],
                                               [round(30/c,1) if c>0 else 0 for c in p7['bkd_mo']],
                                               p7['mttr_mo'])):
        r = i+9; bg = LIGHT_ROW if i%2==0 else WHITE
        flag = ""; fbg = bg
        _nonzero = [c for c in p7['bkd_mo'] if c > 0]
        if _nonzero and bkd_c == min(_nonzero): flag = "Best month ↑"; fbg = GREEN_FLG
        elif _nonzero and bkd_c == max(_nonzero): flag = "Worst month ↓"; fbg = RED_FLG
        for mc in [f"B{r}:C{r}",f"D{r}:E{r}",f"F{r}:G{r}",f"H{r}:I{r}",f"J{r}:L{r}"]:
            ws7.merge_cells(mc)
        dc(ws7,"B",r,mo,bg,bold=True); dc(ws7,"D",r,bkd_c,bg)
        dc(ws7,"F",r,f"{m30}d" if m30 else "—",bg)
        dc(ws7,"H",r,f"{mt}h" if mt else "—",bg)
        dc(ws7,"J",r,flag,fbg,align="left"); ws7.row_dimensions[r].height=18

    r_note7 = 9 + n_mo + 1
    ws7.row_dimensions[r_note7].height=10
    sec(ws7,f"MTTR NOTE: {p7['mttr_note']}",r_note7,"L",ORANGE if not p7['mttr'] else HDR_BG)

    # Elapsed MTTR from completion records (new format only)
    if p7.get('mttr_elapsed') is not None:
        r_el = r_note7 + 2; ws7.row_dimensions[r_el-1].height = 8
        ws7.merge_cells(f"B{r_el}:L{r_el}")
        el_txt = (f"Elapsed response time (ActualFinish − Basic Start Date): "
                  f"{p7['mttr_elapsed']}h avg across {p7['mttr_elapsed_n']} closed BKD WOs. "
                  f"Measures total time from WO creation to close (includes waiting, parts, approvals). "
                  f"Compare to booked labor hours above to see non-wrench time.")
        dc(ws7,"B",r_el,el_txt,LIGHT_ROW,align="left",wrap=True)
        ws7.row_dimensions[r_el].height = 36

    # MTBF chart (hidden cols)
    for i,(mo,m30) in enumerate(zip(p7['months_str'],[round(30/c,1) if c>0 else 0 for c in p7['bkd_mo']])):
        ws7[f"O{9+i}"].value = mo; ws7[f"P{9+i}"].value = m30
    ws7["P8"].value = "MTBF (days)"
    # If elapsed MTTR note was written at r_note7+2, bump chart section down to avoid collision
    r_lc7 = (r_note7 + 4) if p7.get('mttr_elapsed') is not None else (r_note7 + 2)
    ws7.row_dimensions[r_lc7].height=10
    sec(ws7,f"MONTHLY MTBF CHART  ({pl})",r_lc7,"L",HDR_BG)
    lc7 = LineChart(); lc7.title=f"MTBF by Month ({pl})"; lc7.style=10
    lc7.width = min(22 + max(0, n_mo-6)*2, 38); lc7.height=12
    lc7.y_axis.title="Days Between Failures"
    lc7.add_data(Reference(ws7,min_col=16,min_row=8,max_row=8+n_mo),titles_from_data=True)
    lc7.set_categories(Reference(ws7,min_col=15,min_row=9,max_row=8+n_mo))
    lc7.series[0].graphicalProperties.line.solidFill="2E75B6"; lc7.series[0].graphicalProperties.line.width=28575
    lc7.series[0].marker.symbol="circle"; lc7.series[0].marker.size=8
    lc7.visible_cells_only = False
    ws7.add_chart(lc7,f"B{r_lc7+1}")
    for col in ["O","P"]: ws7.column_dimensions[col].hidden = True
    ws7.freeze_panes = "B9"
    widths(ws7,{"A":2,"B":10,"C":8,"D":10,"E":4,"F":14,"G":4,"H":14,"I":4,"J":16,"K":8,"L":4})

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ═════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ═════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="PM Effectiveness App 1.2", page_icon="🔧", layout="centered")

st.title("PM Effectiveness App 1.2")
st.caption("Supports new Celonis export format (Maintenance Activity Type + Priority). Adds schedule stability, due date urgency, and elapsed MTTR. Also handles legacy format.")

asset_label  = st.text_input("Asset label", value="NEW MILFORD  ·  MBF01 | CARTONER")

col_a, col_b = st.columns(2)
with col_a:
    uploaded = st.file_uploader("Celonis WO Export (.xlsx) — required", type=["xlsx"])
with col_b:
    cost_uploaded = st.file_uploader("Celonis Cost Export (.xlsx) — optional", type=["xlsx"],
                                     help="Enables Pillar 6 cost analysis. Join key: Work Order Number.")

if uploaded:
    with st.spinner("Running analysis across 7 pillars…"):
        try:
            df, pm, bkd, fuc, ref_date = prep_data(uploaded)
            mw = build_month_window(df)

            cost_df = None
            if cost_uploaded:
                cost_df = pd.read_excel(cost_uploaded)

            p1 = pillar1_compliance(pm, bkd, fuc, df, mw)
            p2 = pillar2_failure_prevention(pm, bkd, mw)
            p3 = pillar3_coverage(pm, bkd)
            p4 = pillar4_wo_quality(pm, bkd, fuc)
            p5 = pillar5_backlog(df, ref_date, mw)
            p6 = pillar6_cost(pm, bkd, fuc, df, cost_df, mw)
            p7 = pillar7_reliability(bkd, df, ref_date, mw)

            st.success(f"Analysis complete — {mw['period_label']}." + (" Cost data included." if cost_df is not None else " No cost file — upload cost export for Pillar 6."))
            cols = st.columns(7)
            for col_ui, pname, p_data in zip(cols, PILLARS, [p1,p2,p3,p4,p5,p6,p7]):
                rag   = p_data['rag']
                emoji = "🟢" if rag=="GREEN" else ("🟡" if rag=="AMBER" else "🔴")
                col_ui.metric(pname.split(". ",1)[1], emoji)

            output = build_workbook(df, pm, bkd, fuc, ref_date, asset_label,
                                    [p1,p2,p3,p4,p5,p6,p7], cost_df=cost_df, mw=mw)
            from datetime import date
            fname = f"PM_Effectiveness_{date.today().strftime('%Y%m%d')}.xlsx"
            st.download_button(
                label="⬇️ Download Analysis Workbook",
                data=output, file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Analysis failed: {e}")
            import traceback; st.code(traceback.format_exc())

with st.expander("Expected file formats"):
    st.markdown("""
**WO Export (required)**
- `WorkOrderNumber`, `Description`, `OrderType`, `System Status`
- `Work order classification for preventive and corrective processing metrics`
- `Created On`, `Basic Start Date`, `Actual Hours`, `SystemCondition`
- `Planned Hours` *(optional — enables planned vs. actual variance)*

**Cost Export (optional)**
- `Work Order Number`, `Description`, `Cost per Work Order`
- `Actual Cost Classification`, `Created On`
""")
