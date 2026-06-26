import io, re
import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# ── palette ──
DARK_BG = "1F2D3D"; MID_BG = "2E4057"; HDR_BG = "3A5068"
LIGHT_ROW = "EAF1F8"; WHITE = "FFFFFF"; YELLOW = "FFF2CC"
RED_FLG = "FCE4D6"; GREEN_FLG = "E2EFDA"; ORANGE = "FCE9D6"

def wfill(h):       return PatternFill("solid", start_color=h, fgColor=h)
def cal(h="center", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def title_row(ws, text, row, end_col="J"):
    ws.merge_cells(f"B{row}:{end_col}{row}")
    c = ws[f"B{row}"]; c.value = text
    c.font = Font(name="Arial", bold=True, color=WHITE, size=14)
    c.fill = wfill(DARK_BG); c.alignment = cal(); ws.row_dimensions[row].height = 28

def sub_row(ws, text, row, end_col="J"):
    ws.merge_cells(f"B{row}:{end_col}{row}")
    c = ws[f"B{row}"]; c.value = text
    c.font = Font(name="Arial", italic=True, color="AAAAAA", size=9)
    c.fill = wfill(DARK_BG); c.alignment = cal(wrap=True); ws.row_dimensions[row].height = 18

def sec(ws, text, row, end_col="J", bg=MID_BG):
    ws.merge_cells(f"B{row}:{end_col}{row}")
    c = ws[f"B{row}"]; c.value = text
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = wfill(bg); c.alignment = cal("left"); ws.row_dimensions[row].height = 18

def hdr(ws, col, row, text, end_col=None):
    ref = f"{col}{row}"
    if end_col: ws.merge_cells(f"{ref}:{end_col}{row}")
    ws[ref].value = text
    ws[ref].font = Font(name="Arial", bold=True, color=WHITE, size=9)
    ws[ref].fill = wfill(HDR_BG); ws[ref].alignment = cal(wrap=True); ws[ref].border = bdr()

def dc(ws, col, row, val, bg=WHITE, bold=False, align="center", wrap=False):
    ws[f"{col}{row}"].value = val
    ws[f"{col}{row}"].font = Font(name="Arial", size=9, color="1F2D3D", bold=bold)
    ws[f"{col}{row}"].fill = wfill(bg)
    ws[f"{col}{row}"].alignment = cal(align, wrap=wrap)
    ws[f"{col}{row}"].border = bdr()

def widths(ws, m):
    for c, w in m.items(): ws.column_dimensions[c].width = w

def pfreq(d):
    if pd.isna(d): return None
    m = re.search(r'(\d+)\s*WK', d, re.IGNORECASE)
    if m: return int(m.group(1)) * 7
    if re.search(r'5W', d, re.IGNORECASE): return 35
    return None

def hf(s, f): return f in str(s).split() if pd.notna(s) else False


def run_analysis(file_obj, asset_label):
    df = pd.read_excel(file_obj)
    df.rename(columns={
        'Work order classification for preventive and corrective processing metrics': 'WO_Classification'
    }, inplace=True)
    df['Basic Start Date'] = pd.to_datetime(df['Basic Start Date'], errors='coerce')
    df['Created On']       = pd.to_datetime(df['Created On'],       errors='coerce')
    df['is_TECO'] = df['System Status'].apply(lambda x: hf(x, 'TECO'))
    df['is_CLSD'] = df['System Status'].apply(lambda x: hf(x, 'CLSD'))

    pm  = df[df['OrderType'] == 'GM09'].copy().sort_values('Basic Start Date')
    bkd = df[df['OrderType'] == 'GM02'].copy().sort_values('Basic Start Date')
    fuc = df[df['OrderType'].isin(['GM01', 'PM01'])].copy().sort_values('Basic Start Date')

    # ── derived metrics ──
    pm_teco_pct  = f"{pm['is_TECO'].mean()*100:.1f}%"
    bkd_teco_pct = f"{bkd['is_TECO'].mean()*100:.1f}%" if len(bkd) else "N/A"
    fuc_teco_pct = f"{fuc['is_TECO'].mean()*100:.1f}%" if len(fuc) else "N/A"
    pm_zero      = (pm['Actual Hours'] == 0).sum()
    bkd_zero     = (bkd['Actual Hours'] == 0).sum()
    fuc_zero     = (fuc['Actual Hours'] == 0).sum()
    pm_on_run    = (pm['SystemCondition'] == '1 - On the run').sum()
    pm_mdown     = (pm['SystemCondition'] == '2 - Machine must be down').sum()
    pm_hrs       = pm['Actual Hours'].sum()
    pm_avg_hrs   = f"{pm['Actual Hours'].mean():.2f}"
    open_wos     = (~df['is_TECO'] & ~df['is_CLSD']).sum()
    pm_bkd_ratio = f"{len(pm)/len(bkd):.2f}:1" if len(bkd) else "N/A"
    bkd_zero_pct = f"{100*bkd_zero/len(bkd):.1f}%" if len(bkd) else "N/A"
    fuc_zero_pct = f"{100*fuc_zero/len(fuc):.1f}%" if len(fuc) else "N/A"
    pm_on_run_pct= f"{100*pm_on_run/len(pm):.1f}%" if len(pm) else "N/A"
    pm_mdown_pct = f"{100*pm_mdown/len(pm):.1f}%" if len(pm) else "N/A"

    # MTBF
    bkd_sorted = bkd.dropna(subset=['Basic Start Date']).sort_values('Basic Start Date')
    bd_dates   = bkd_sorted['Basic Start Date'].tolist()
    bd_gaps    = [(bd_dates[i]-bd_dates[i-1]).days for i in range(1, len(bd_dates))]
    mtbf_overall = f"{np.mean(bd_gaps):.1f} days" if bd_gaps else "N/A"

    # Schedule adherence (H1 2026)
    pm_h1 = pm[pm['Basic Start Date'].between('2026-01-01', '2026-06-30')].copy()
    sched_rows = []
    for desc, grp in pm_h1.groupby('Description'):
        grp = grp.sort_values('Basic Start Date'); interval = pfreq(desc)
        if interval is None or len(grp) < 2: continue
        dates = grp['Basic Start Date'].tolist()
        for i in range(1, len(dates)):
            actual = (dates[i]-dates[i-1]).days; var = actual - interval
            status = ('On time' if abs(var) <= max(1, interval*0.2)
                      else 'Early' if var < 0
                      else 'Late (≤50%)' if var <= interval*0.5
                      else 'Significantly late')
            sched_rows.append({'PM_Description': desc, 'Expected_Days': interval,
                'From': dates[i-1].strftime('%Y-%m-%d'), 'To': dates[i].strftime('%Y-%m-%d'),
                'Actual_Gap': actual, 'Variance': var, 'Status': status})
    sched = pd.DataFrame(sched_rows) if sched_rows else pd.DataFrame()
    if len(sched):
        vc = sched['Status'].value_counts()
        on_time  = vc.get('On time', 0); early = vc.get('Early', 0)
        late_ct  = vc.get('Late (≤50%)', 0) + vc.get('Significantly late', 0)
        total_i  = len(sched)
        adh_pct  = f"{100*on_time/total_i:.1f}%"
        by_type  = sched.groupby(['PM_Description','Expected_Days']).agg(
            Intervals=('Actual_Gap','count'), Avg_Gap=('Actual_Gap','mean'),
            Avg_Var=('Variance','mean'), Max_Late=('Variance','max'),
            On_Time=('Status', lambda x: (x=='On time').sum())).reset_index()
        by_type['OT_Pct'] = (by_type['On_Time']/by_type['Intervals']*100).round(0).astype(int)
    else:
        on_time = early = late_ct = total_i = 0; adh_pct = "N/A"
        by_type = pd.DataFrame()

    # PM→Breakdown pairs
    pm_bkd_rows = []
    for _, bd in bkd.iterrows():
        prior = pm[pm['Basic Start Date'] <= bd['Basic Start Date']]
        if not len(prior): continue
        last = prior.iloc[-1]; days = (bd['Basic Start Date'] - last['Basic Start Date']).days
        cat = ("Same day (PM revealed)" if days == 0
               else "1-3 days (immediate failure)" if days <= 3
               else "4-7 days (short-interval)" if days <= 7
               else "8+ days")
        pm_bkd_rows.append({'PM_WO': last['WorkOrderNumber'], 'PM_Desc': last['Description'],
            'PM_Date': last['Basic Start Date'].strftime('%Y-%m-%d'),
            'BKD_WO': bd['WorkOrderNumber'], 'BKD_Desc': bd['Description'],
            'BKD_Date': bd['Basic Start Date'].strftime('%Y-%m-%d'), 'Days': days, 'Category': cat})
    pm_bkd = pd.DataFrame(pm_bkd_rows) if pm_bkd_rows else pd.DataFrame()

    # Monthly counts (H1 2026)
    months_str = ['Jan 2026','Feb 2026','Mar 2026','Apr 2026','May 2026','Jun 2026']
    month_map  = {'2026-01': 0,'2026-02': 1,'2026-03': 2,'2026-04': 3,'2026-05': 4,'2026-06': 5}
    pm_mo  = [0]*6; bkd_mo = [0]*6; fuc_mo = [0]*6
    for frame, arr in [(pm, pm_mo), (bkd, bkd_mo), (fuc, fuc_mo)]:
        frame2 = frame.dropna(subset=['Basic Start Date'])
        for _, row in frame2.iterrows():
            key = row['Basic Start Date'].strftime('%Y-%m')
            if key in month_map: arr[month_map[key]] += 1

    # MTBF by month
    bkd_h1 = bkd[bkd['Basic Start Date'].between('2026-01-01','2026-06-30')].dropna(subset=['Basic Start Date'])
    bkd_h1 = bkd_h1.copy(); bkd_h1['Month'] = bkd_h1['Basic Start Date'].dt.to_period('M')
    mtbf_30 = [round(30/c, 1) if c > 0 else 0 for c in bkd_mo]
    mtbf_actual = []
    for mo_key in ['2026-01','2026-02','2026-03','2026-04','2026-05','2026-06']:
        grp = bkd_h1[bkd_h1['Month'].astype(str) == mo_key].sort_values('Basic Start Date')
        d = grp['Basic Start Date'].tolist()
        if len(d) >= 2:
            gaps = [(d[i]-d[i-1]).days for i in range(1, len(d))]
            mtbf_actual.append(round(np.mean(gaps), 1))
        else:
            mtbf_actual.append(0)

    # Failure clusters (keyword-based)
    kw_clusters = [
        ("Glue System",          ['glue','nozzle','nordson','versa blue','nozzels']),
        ("Gearbox / Shaft",      ['gearbox','gear box','shaft']),
        ("Pusher / Movement",    ['pusher','layer pusher']),
        ("Fuses / Electrical",   ['fuse','vfd','not moving','electrical']),
        ("Belts / Chains / Lugs",['belt','chain','lug']),
        ("Carton Handling",      ['carton','blow off']),
        ("Drive / Other",        ['drive shaft','resolver','wadding']),
    ]
    cluster_data = []
    for name, keywords in kw_clusters:
        mask = bkd['Description'].str.lower().str.contains('|'.join(keywords), na=False)
        matched = bkd[mask]
        if len(matched):
            cluster_data.append((name, len(matched), ', '.join(matched['Description'].tolist())))

    # Breakdown by proximity bucket
    if len(pm_bkd):
        b0  = (pm_bkd['Category'] == 'Same day (PM revealed)').sum()
        b13 = (pm_bkd['Category'] == '1-3 days (immediate failure)').sum()
        b47 = (pm_bkd['Category'] == '4-7 days (short-interval)').sum()
        b8p = (pm_bkd['Category'] == '8+ days').sum()
        total_bkd_pairs = len(pm_bkd)
    else:
        b0 = b13 = b47 = b8p = total_bkd_pairs = 0

    bkd_within7 = b0 + b13 + b47
    bkd_w7_pct  = f"{100*bkd_within7/len(bkd):.0f}%" if len(bkd) else "N/A"

    # Date range label
    min_date = df['Created On'].min()
    max_date = df['Created On'].max()
    date_range = f"{min_date.strftime('%b %Y')}–{max_date.strftime('%b %Y')}" if pd.notna(min_date) else "Unknown"

    cat_clr = {"Same day (PM revealed)": GREEN_FLG, "1-3 days (immediate failure)": RED_FLG,
                "4-7 days (short-interval)": ORANGE, "8+ days": LIGHT_ROW}

    # ══════════════════════════════════════════════
    # BUILD WORKBOOK
    # ══════════════════════════════════════════════
    wb = Workbook()

    # ── S1: Summary Dashboard ──
    ws1 = wb.active; ws1.title = "Summary Dashboard"; ws1.sheet_view.showGridLines = False
    title_row(ws1, f"PM EFFECTIVENESS ANALYSIS  ·  {asset_label}", 1, "H")
    sub_row(ws1, f"Celonis Export  |  {date_range}  |  235 WOs  |  GM09=Preventive  GM02=Breakdown  GM01/PM01=Follow-up Corrective", 2, "H")
    ws1.row_dimensions[3].height = 8
    for col, txt in [("B","Metric"),("C","Value"),("E","Calculation"),("H","Time Period")]:
        hdr(ws1, col, 4, txt)
    ws1.merge_cells("E4:G4"); ws1.row_dimensions[4].height = 18

    summary_rows = [
        ("Total Work Orders",                   len(df),                "COUNT all WOs",                                              date_range,      WHITE),
        ("PM Work Orders (GM09)",               len(pm),                "COUNT where OrderType = 'GM09'",                             date_range,      LIGHT_ROW),
        ("Breakdown WOs (GM02)",                len(bkd),               "COUNT where OrderType = 'GM02' — unplanned",                 date_range,      RED_FLG),
        ("Follow-up Corrective (GM01/PM01)",    len(fuc),               "COUNT where OrderType IN ('GM01','PM01')",                   date_range,      ORANGE),
        ("PM : Breakdown Ratio",                pm_bkd_ratio,           "PM WO count / Breakdown count",                              date_range,      LIGHT_ROW),
        ("PM Compliance Rate (TECO %)",         pm_teco_pct,            "PM WOs TECO'd / Total PM WOs",                              date_range,      GREEN_FLG),
        ("PM Schedule Adherence",               adh_pct,                "On-time intervals / Total intervals (H1 2026)",              "H1 2026",       GREEN_FLG),
        ("PM Zero-Hour WOs",                    f"{pm_zero} ({100*pm_zero/len(pm):.0f}%)" if len(pm) else "N/A",
                                                                         "✓ PM WOs with 0 actual hours — ideally 0%",                 date_range,      GREEN_FLG if pm_zero == 0 else YELLOW),
        ("⚠ Breakdown Zero-Hour WOs",          f"{bkd_zero}/{len(bkd)} ({bkd_zero_pct})" if len(bkd) else "N/A",
                                                                         "Labor not captured on GM02 WOs — understates failure cost", date_range,      YELLOW),
        ("⚠ Follow-up Corrective Zero-Hr",     f"{fuc_zero}/{len(fuc)} ({fuc_zero_pct})" if len(fuc) else "N/A",
                                                                         "Many are parts-only; audit non-parts zero-hour entries",    date_range,      YELLOW),
        ("PM On-the-Run Rate",                  f"{pm_on_run}/{len(pm)} ({pm_on_run_pct})" if len(pm) else "N/A",
                                                                         "PM WOs with SystemCond='On the run' / Total PM",            date_range,      ORANGE),
        ("MTBF (avg days between BKDs)",        mtbf_overall,           "Avg days between consecutive GM02 events",                   date_range,      LIGHT_ROW),
        (f"BKDs within 7 days of prior PM",    f"{bkd_within7}/{len(bkd)} ({bkd_w7_pct})" if len(bkd) else "N/A",
                                                                         "GM02 BKDs occurring ≤7 days after prior PM — see PM→BKD sheet", "H1 2026",  RED_FLG),
        ("Open WOs (no TECO/CLSD)",             open_wos,               "WOs without TECO or CLSD flag",                             "As of export",  LIGHT_ROW),
    ]
    for i, (metric, val, calc, period, bg) in enumerate(summary_rows):
        r = i+5; ws1.merge_cells(f"E{r}:G{r}")
        dc(ws1,"B",r,metric,bg,bold=True,align="left"); dc(ws1,"C",r,str(val),bg)
        dc(ws1,"E",r,calc,bg,align="left",wrap=True)
        dc(ws1,"H",r,period,bg,align="left",wrap=True); ws1.row_dimensions[r].height = 22

    r_k = len(summary_rows)+6; ws1.row_dimensions[r_k-1].height = 10
    sec(ws1,"KEY FINDINGS",r_k,"H")
    findings = [
        (f"PM compliance: {pm_teco_pct} TECO rate across {len(pm)} PM WOs.", GREEN_FLG),
        (f"Schedule adherence: {adh_pct} of measured intervals on time (H1 2026).", GREEN_FLG if adh_pct not in ("N/A","0.0%") else LIGHT_ROW),
        (f"{bkd_within7}/{len(bkd)} breakdowns occurred within 7 days of a prior PM — see PM→Breakdown sheet for detail.", RED_FLG),
        (f"MTBF: {mtbf_overall} average between breakdown events.", LIGHT_ROW),
        (f"⚠ {bkd_zero}/{len(bkd)} GM02 Breakdown WOs carry zero actual hours — true failure cost cannot be assessed until resolved.", YELLOW),
        (f"Glue system, gearbox/shaft, and pusher components are the top repeat failure clusters — see Repeat Failure Modes sheet.", ORANGE),
    ]
    for j, (txt, bg) in enumerate(findings):
        r_f = r_k+1+j; ws1.merge_cells(f"B{r_f}:H{r_f}")
        dc(ws1,"B",r_f,txt,bg,align="left",wrap=True); ws1.row_dimensions[r_f].height = 24
    widths(ws1,{"A":2,"B":33,"C":14,"D":2,"E":38,"F":8,"G":8,"H":28,"I":2})

    # ── S2: Monthly Trend ──
    ws2 = wb.create_sheet("Monthly Trend"); ws2.sheet_view.showGridLines = False
    title_row(ws2, f"MONTHLY WORK ORDER TREND  ·  {asset_label}  ·  Jan–Jun 2026", 1, "K")
    sub_row(ws2, "PM=GM09  |  Breakdown=GM02  |  Follow-up Corrective=GM01+PM01  |  Grouped by Basic Start Date", 2, "K")
    ws2.row_dimensions[3].height = 8
    for j, h in enumerate(["Month","PM (GM09)","Breakdown (GM02)","Follow-up Corr","Total","PM %","BKD %","FUC %","PM:BKD"]):
        hdr(ws2, get_column_letter(j+2), 4, h)
    ws2.row_dimensions[4].height = 18
    for i, (mo, pm_c, bkd_c, fuc_c) in enumerate(zip(months_str, pm_mo, bkd_mo, fuc_mo)):
        r = i+5; tot = pm_c+bkd_c+fuc_c; bg = LIGHT_ROW if i%2 == 0 else WHITE
        vals = [mo, pm_c, bkd_c, fuc_c, tot,
                f"{100*pm_c/tot:.0f}%" if tot else "—",
                f"{100*bkd_c/tot:.0f}%" if tot else "—",
                f"{100*fuc_c/tot:.0f}%" if tot else "—",
                f"{pm_c/bkd_c:.1f}:1" if bkd_c else "—"]
        for j, v in enumerate(vals):
            dc(ws2, get_column_letter(j+2), r, v, bg, bold=(j==0))
        ws2.row_dimensions[r].height = 18
    tp,tb,tf = sum(pm_mo),sum(bkd_mo),sum(fuc_mo); tt = tp+tb+tf
    for j, v in enumerate(["H1 2026 Total",tp,tb,tf,tt,
                            f"{100*tp/tt:.0f}%" if tt else "—",
                            f"{100*tb/tt:.0f}%" if tt else "—",
                            f"{100*tf/tt:.0f}%" if tt else "—",
                            f"{tp/tb:.1f}:1" if tb else "—"]):
        c = ws2[f"{get_column_letter(j+2)}11"]
        c.value = v; c.font = Font(name="Arial",bold=True,size=9,color=WHITE)
        c.fill = wfill(MID_BG); c.alignment = cal(); c.border = bdr()
    ws2.row_dimensions[11].height = 20
    ws2.row_dimensions[13].height = 8
    sec(ws2,"MONTHLY WO MIX — PM / BREAKDOWN / FOLLOW-UP CORRECTIVE",13,"K",HDR_BG)
    cats = Reference(ws2, min_col=2, min_row=5, max_row=10)
    bar = BarChart(); bar.type="col"; bar.grouping="stacked"; bar.overlap=100
    bar.title = "Monthly WO Count by Type"; bar.y_axis.title = "Work Orders"
    bar.style = 10; bar.width = 22; bar.height = 12
    for ci, clr in [(3,"2E75B6"),(4,"C00000"),(5,"ED7D31")]:
        bar.add_data(Reference(ws2,min_col=ci,min_row=4,max_row=10),titles_from_data=True)
    bar.set_categories(cats)
    for idx, clr in enumerate(["2E75B6","C00000","ED7D31"]): bar.series[idx].graphicalProperties.solidFill = clr
    ws2.add_chart(bar,"B14")
    ws2.row_dimensions[30].height = 8
    sec(ws2,"MONTHLY BREAKDOWN TREND (GM02)",30,"K",HDR_BG)
    lc = LineChart(); lc.title = "Monthly Breakdown WOs (GM02)"; lc.style = 10; lc.width = 22; lc.height = 10
    lc.add_data(Reference(ws2,min_col=4,min_row=4,max_row=10),titles_from_data=True)
    lc.set_categories(cats); lc.series[0].graphicalProperties.line.solidFill = "C00000"
    lc.series[0].graphicalProperties.line.width = 28575
    lc.series[0].marker.symbol = "circle"; lc.series[0].marker.size = 7
    ws2.add_chart(lc,"B31")
    widths(ws2,{"A":2,"B":16,"C":14,"D":18,"E":18,"F":10,"G":9,"H":9,"I":9,"J":12,"K":2})

    # ── S3: Zero Hour Analysis ──
    ws3 = wb.create_sheet("Zero Hour WO Analysis"); ws3.sheet_view.showGridLines = False
    title_row(ws3,f"ZERO HOUR WORK ORDER ANALYSIS  ·  {asset_label}",1,"J")
    sub_row(ws3,"Actual Hours = 0 by WO type. PM (GM09) should be 0%. GM02 and GM01 gaps indicate labor not captured.",2,"J")
    ws3.row_dimensions[3].height = 8
    sec(ws3,"ZERO HOUR SUMMARY BY WO TYPE",4,"J")
    for col,txt,ec in [("B","WO Type",None),("C","Total WOs",None),("D","Zero-Hour WOs",None),
                        ("E","Zero-Hour %",None),("F","Assessment","H"),("I","Action Required",None)]:
        hdr(ws3,col,5,txt,ec)
    ws3.row_dimensions[5].height = 18
    zh_sum = [
        ("PM (GM09)", len(pm), pm_zero,
         f"{100*pm_zero/len(pm):.1f}%" if len(pm) else "N/A",
         "✓ All PMs have hours — no data gap" if pm_zero==0 else f"⚠ {pm_zero} PMs missing hours",
         GREEN_FLG if pm_zero==0 else YELLOW,
         "None — maintain current practice" if pm_zero==0 else "Investigate missing PM hours"),
        ("Breakdown (GM02)", len(bkd), bkd_zero,
         bkd_zero_pct,
         "⚠ Critical: BKD labor not captured on WO",
         YELLOW,
         "Validate — hours may be on parent PM WO or genuinely missing. Required for true failure cost visibility."),
        ("Follow-up Corr (GM01/PM01)", len(fuc), fuc_zero,
         fuc_zero_pct,
         "⚠ Moderate: many are parts orders; some may be missing hours",
         ORANGE,
         "Separate parts-only from labor WOs; audit non-parts zero-hour entries."),
    ]
    for i,(wtype,tot,zero,pct,assess,bg,action) in enumerate(zh_sum):
        r = i+6; ws3.merge_cells(f"F{r}:H{r}")
        dc(ws3,"B",r,wtype,bg,bold=True); dc(ws3,"C",r,tot,bg); dc(ws3,"D",r,zero,bg)
        dc(ws3,"E",r,pct,bg); dc(ws3,"F",r,assess,bg,align="left",wrap=True)
        dc(ws3,"I",r,action,bg,align="left",wrap=True); ws3.row_dimensions[r].height = 36
    # BKD zero-hour detail
    ws3.row_dimensions[10].height = 8
    sec(ws3,f"BREAKDOWN (GM02) ZERO-HOUR WO DETAIL — {bkd_zero} records",10,"J",HDR_BG)
    for col,txt,ec in [("B","WO Number",None),("C","Description","D"),("E","BKD Date",None),
                        ("F","Actual Hrs",None),("G","TECO?",None)]:
        hdr(ws3,col,11,txt,ec)
    ws3.row_dimensions[11].height = 18
    bkd_zero_df = bkd[bkd['Actual Hours']==0].sort_values('Basic Start Date')
    for i,(_,row) in enumerate(bkd_zero_df.iterrows()):
        r = 12+i; bg = LIGHT_ROW if i%2==0 else WHITE
        ws3.merge_cells(f"C{r}:D{r}")
        dc(ws3,"B",r,row['WorkOrderNumber'],bg)
        dc(ws3,"C",r,row['Description'],bg,align="left")
        dc(ws3,"E",r,row['Basic Start Date'].strftime('%Y-%m-%d') if pd.notna(row['Basic Start Date']) else "—",bg)
        dc(ws3,"F",r,0,bg)
        dc(ws3,"G",r,"Yes" if row['is_TECO'] else "No", GREEN_FLG if row['is_TECO'] else RED_FLG)
        ws3.row_dimensions[r].height = 16
    widths(ws3,{"A":2,"B":16,"C":36,"D":8,"E":14,"F":12,"G":10,"H":10,"I":36,"J":2})

    # ── S4: PM → Breakdown ──
    ws4 = wb.create_sheet("PM → Breakdown"); ws4.sheet_view.showGridLines = False
    title_row(ws4,f"PM → BREAKDOWN ANALYSIS  ·  {asset_label}",1,"L")
    sub_row(ws4,"For each GM02 Breakdown, the most recent prior GM09 PM is identified and days between are calculated.",2,"L")
    ws4.row_dimensions[3].height = 8
    sec(ws4,"DAYS FROM PM TO SUBSEQUENT BREAKDOWN",4,"L")
    for col,txt,ec in [("B","Timing Bucket",None),("C","# Breakdowns",None),
                        ("D","% of Total",None),("E","Interpretation","L")]:
        hdr(ws4,col,5,txt,ec)
    ws4.row_dimensions[5].height = 18
    buckets = [
        ("0 — Same day",  b0,  f"{100*b0/total_bkd_pairs:.1f}%" if total_bkd_pairs else "—",
         "PM likely surfaced the failure during inspection — corrective action concurrent", GREEN_FLG),
        ("1 – 3 days",    b13, f"{100*b13/total_bkd_pairs:.1f}%" if total_bkd_pairs else "—",
         "Immediate failure after PM — latent defect surfaced or component disturbed", RED_FLG),
        ("4 – 7 days",    b47, f"{100*b47/total_bkd_pairs:.1f}%" if total_bkd_pairs else "—",
         "Short-interval failure — PM may not address root cause", ORANGE),
        ("8+ days",       b8p, f"{100*b8p/total_bkd_pairs:.1f}%" if total_bkd_pairs else "—",
         "Breakdowns occurring more than 7 days after prior PM", LIGHT_ROW),
    ]
    for i,(bkt,ct,pct,interp,bg) in enumerate(buckets):
        r = i+6; ws4.merge_cells(f"E{r}:L{r}")
        dc(ws4,"B",r,bkt,bg,bold=True); dc(ws4,"C",r,ct,bg)
        dc(ws4,"D",r,pct,bg); dc(ws4,"E",r,interp,bg,align="left",wrap=True)
        ws4.row_dimensions[r].height = 22
    ws4.row_dimensions[11].height = 8
    sec(ws4,"BREAKDOWN TIMING AFTER PRIOR PM",11,"L",HDR_BG)
    bc = BarChart(); bc.type="col"; bc.title=f"Breakdown Timing After Prior PM (n={total_bkd_pairs})"
    bc.y_axis.title = "# Breakdowns"; bc.style = 10; bc.width = 18; bc.height = 11
    bc.add_data(Reference(ws4,min_col=3,min_row=5,max_row=9),titles_from_data=True)
    bc.set_categories(Reference(ws4,min_col=2,min_row=6,max_row=9))
    bc.series[0].graphicalProperties.solidFill = "C00000"
    ws4.add_chart(bc,"B12")
    r_d = 26; ws4.row_dimensions[r_d-1].height = 8
    sec(ws4,f"DETAIL: ALL PM → BREAKDOWN PAIRS ({total_bkd_pairs} records)",r_d,"L")
    for col,txt,ec in [("B","PM WO #",None),("C","PM Description","D"),("E","PM Date",None),
                        ("F","BKD WO #",None),("G","Breakdown Description","H"),
                        ("I","BKD Date",None),("J","Days",None),("K","Category","L")]:
        hdr(ws4,col,r_d+1,txt,ec)
    ws4.row_dimensions[r_d+1].height = 18
    for i, row in pm_bkd.iterrows():
        r = r_d+2+i; bg = cat_clr.get(row['Category'], LIGHT_ROW)
        ws4.merge_cells(f"C{r}:D{r}"); ws4.merge_cells(f"G{r}:H{r}"); ws4.merge_cells(f"K{r}:L{r}")
        dc(ws4,"B",r,row['PM_WO'],bg); dc(ws4,"C",r,row['PM_Desc'],bg,align="left")
        dc(ws4,"E",r,row['PM_Date'],bg); dc(ws4,"F",r,row['BKD_WO'],bg)
        dc(ws4,"G",r,row['BKD_Desc'],bg,align="left"); dc(ws4,"I",r,row['BKD_Date'],bg)
        dc(ws4,"J",r,row['Days'],bg); dc(ws4,"K",r,row['Category'],bg,align="left")
        ws4.row_dimensions[r].height = 16
    widths(ws4,{"A":2,"B":13,"C":36,"D":4,"E":12,"F":13,"G":34,"H":4,"I":12,"J":8,"K":28,"L":2})

    # ── S5: Repeat Failure Modes ──
    ws5 = wb.create_sheet("Repeat Failure Modes"); ws5.sheet_view.showGridLines = False
    title_row(ws5,f"REPEAT FAILURE MODE ANALYSIS  ·  {asset_label}",1,"J")
    sub_row(ws5,"GM02 Breakdown WOs grouped by failure theme. Identifies components with recurring failures for PM scope review.",2,"J")
    ws5.row_dimensions[3].height = 8
    sec(ws5,"FAILURE CLUSTER SUMMARY",4,"J")
    for col,txt,ec in [("B","Failure Cluster",None),("C","# BKD WOs",None),
                        ("D","% of Total",None),("E","Work Order Descriptions","H"),("I","Priority",None)]:
        hdr(ws5,col,5,txt,ec)
    ws5.row_dimensions[5].height = 18
    priority_map = {0: "HIGH", 1: "HIGH", 2: "MED", 3: "MED", 4: "MED", 5: "LOW", 6: "LOW"}
    clr_map = {"HIGH": RED_FLG, "MED": ORANGE, "LOW": LIGHT_ROW}
    cluster_data_sorted = sorted(cluster_data, key=lambda x: -x[1])
    for i,(name,ct,descs) in enumerate(cluster_data_sorted):
        r = i+6; priority = priority_map.get(i,"LOW"); bg = clr_map.get(priority,LIGHT_ROW)
        ws5.merge_cells(f"E{r}:H{r}")
        dc(ws5,"B",r,name,bg,bold=True,align="left"); dc(ws5,"C",r,ct,bg)
        dc(ws5,"D",r,f"{100*ct/len(bkd):.0f}%" if len(bkd) else "—",bg)
        dc(ws5,"E",r,descs,bg,align="left",wrap=True); dc(ws5,"I",r,priority,bg)
        ws5.row_dimensions[r].height = 28
    r_ch = len(cluster_data_sorted)+8; ws5.row_dimensions[r_ch-1].height = 8
    sec(ws5,"BREAKDOWN WOs BY FAILURE CLUSTER",r_ch,"J",HDR_BG)
    cb = BarChart(); cb.type="bar"; cb.title="GM02 Breakdown WOs by Failure Cluster"
    cb.x_axis.title = "# Breakdown WOs"; cb.style = 10; cb.width = 20; cb.height = 12
    cb.add_data(Reference(ws5,min_col=3,min_row=5,max_row=5+len(cluster_data_sorted)),titles_from_data=True)
    cb.set_categories(Reference(ws5,min_col=2,min_row=6,max_row=5+len(cluster_data_sorted)))
    cb.series[0].graphicalProperties.solidFill = "C00000"
    ws5.add_chart(cb,f"B{r_ch+1}")
    widths(ws5,{"A":2,"B":26,"C":12,"D":12,"E":30,"F":8,"G":8,"H":8,"I":12,"J":2})

    # ── S6: PM Schedule Adherence ──
    ws6 = wb.create_sheet("PM Schedule Adherence"); ws6.sheet_view.showGridLines = False
    title_row(ws6,f"PM SCHEDULE ADHERENCE  ·  {asset_label}  ·  H1 2026",1,"P")
    sub_row(ws6,"Expected interval parsed from PM description (e.g. '1WK'=7d). Adherence = actual gap vs. expected ±20%. H1 2026 only.",2,"P")
    ws6.row_dimensions[3].height = 8
    for col,label,val,bg in [("B","Intervals Analyzed",str(total_i),LIGHT_ROW),
                               ("E","On Time",f"{on_time}  ({100*on_time/total_i:.0f}%)" if total_i else "N/A",GREEN_FLG),
                               ("H","Late",f"{late_ct}  ({100*late_ct/total_i:.0f}%)" if total_i else "N/A",RED_FLG if late_ct>3 else LIGHT_ROW),
                               ("K","Early",f"{early}  ({100*early/total_i:.0f}%)" if total_i else "N/A",LIGHT_ROW),
                               ("N","Overall Adherence",adh_pct,GREEN_FLG)]:
        ec = chr(ord(col)+2)
        ws6.merge_cells(f"{col}4:{ec}4"); ws6[f"{col}4"].value = label
        ws6[f"{col}4"].font = Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws6[f"{col}4"].fill = wfill(WHITE); ws6[f"{col}4"].alignment = cal("left")
        ws6.merge_cells(f"{col}5:{ec}5"); ws6[f"{col}5"].value = val
        ws6[f"{col}5"].font = Font(name="Arial",bold=True,size=14,color="1F2D3D")
        ws6[f"{col}5"].fill = wfill(bg); ws6[f"{col}5"].alignment = cal(); ws6[f"{col}5"].border = bdr()
    ws6.row_dimensions[4].height = 16; ws6.row_dimensions[5].height = 28; ws6.row_dimensions[6].height = 10
    sec(ws6,"ADHERENCE BY PM TASK",7,"P")
    for col,txt,ec in [("B","PM Task Description","D"),("E","Expected","F"),("G","Intervals","H"),
                        ("I","Avg Gap","J"),("K","Avg Variance","L"),("M","Max Late","N"),("O","On-Time %","P")]:
        hdr(ws6,col,8,txt,ec)
    ws6.row_dimensions[8].height = 18
    if len(by_type):
        for i,row in by_type.sort_values('OT_Pct',ascending=True).iterrows():
            r = 9+list(by_type.sort_values('OT_Pct',ascending=True).index).index(i)
            ot = row['OT_Pct']
            bg = GREEN_FLG if ot==100 else (LIGHT_ROW if ot>=90 else (ORANGE if ot>=67 else RED_FLG))
            var_s = f"+{row['Avg_Var']:.1f}" if row['Avg_Var']>0 else f"{row['Avg_Var']:.1f}"
            late_s = f"+{int(row['Max_Late'])}d" if row['Max_Late']>0 else "—"
            for mc in [f"B{r}:D{r}",f"E{r}:F{r}",f"G{r}:H{r}",f"I{r}:J{r}",f"K{r}:L{r}",f"M{r}:N{r}",f"O{r}:P{r}"]:
                ws6.merge_cells(mc)
            dc(ws6,"B",r,row['PM_Description'],bg,bold=True,align="left")
            dc(ws6,"E",r,f"{row['Expected_Days']}d ({int(row['Expected_Days']//7)}WK)",bg)
            dc(ws6,"G",r,int(row['Intervals']),bg); dc(ws6,"I",r,f"{row['Avg_Gap']:.1f}d",bg)
            dc(ws6,"K",r,var_s,bg); dc(ws6,"M",r,late_s,bg); dc(ws6,"O",r,f"{int(ot)}%",bg,bold=True)
            ws6.row_dimensions[r].height = 18
    if len(sched):
        r_at = 9+len(by_type); ws6.row_dimensions[r_at].height = 10
        sec(ws6,"LATE / EARLY INTERVALS — DETAIL",r_at+1,"P",HDR_BG)
        late_rows = sched[sched['Status']!='On time'].sort_values('Variance',ascending=False)
        for col,txt,ec in [("B","PM Task","D"),("E","From","F"),("G","To","H"),
                            ("I","Expected","J"),("K","Actual","L"),("M","Variance","N"),("O","Status","P")]:
            hdr(ws6,col,r_at+2,txt,ec)
        ws6.row_dimensions[r_at+2].height = 18
        for i,(_,row) in enumerate(late_rows.iterrows()):
            r = r_at+3+i; bg = RED_FLG if 'late' in row['Status'].lower() else LIGHT_ROW
            var_s = f"+{int(row['Variance'])}d" if row['Variance']>0 else f"{int(row['Variance'])}d"
            for mc in [f"B{r}:D{r}",f"E{r}:F{r}",f"G{r}:H{r}",f"I{r}:J{r}",f"K{r}:L{r}",f"M{r}:N{r}",f"O{r}:P{r}"]:
                ws6.merge_cells(mc)
            dc(ws6,"B",r,row['PM_Description'],bg,bold=True,align="left")
            dc(ws6,"E",r,row['From'],bg); dc(ws6,"G",r,row['To'],bg)
            dc(ws6,"I",r,f"{int(row['Expected_Days'])}d",bg); dc(ws6,"K",r,f"{int(row['Actual_Gap'])}d",bg)
            dc(ws6,"M",r,var_s,bg,bold=True); dc(ws6,"O",r,row['Status'],bg)
            ws6.row_dimensions[r].height = 16
    widths(ws6,{"A":2,"B":8,"C":8,"D":24,"E":8,"F":6,"G":12,"H":4,"I":12,"J":4,"K":12,"L":4,"M":10,"N":4,"O":14,"P":2})

    # ── S7: MTBF Trend ──
    ws7 = wb.create_sheet("MTBF Trend"); ws7.sheet_view.showGridLines = False
    title_row(ws7,f"MEAN TIME BETWEEN FAILURES  ·  {asset_label}  ·  H1 2026",1,"J")
    sub_row(ws7,"MTBF = avg days between consecutive GM02 Breakdown events. Higher = better. H1 2026 only.",2,"J")
    ws7.row_dimensions[3].height = 8
    best_mo  = months_str[mtbf_30.index(max(mtbf_30))] if mtbf_30 else "N/A"
    worst_mo = months_str[mtbf_30.index(min(m for m in mtbf_30 if m > 0))] if any(m>0 for m in mtbf_30) else "N/A"
    best_val  = f"{max(mtbf_30):.1f} days" if mtbf_30 else "N/A"
    worst_val = f"{min(m for m in mtbf_30 if m > 0):.1f} days" if any(m>0 for m in mtbf_30) else "N/A"
    for col,label,val,bg in [("B","H1 2026 Overall MTBF",mtbf_overall,LIGHT_ROW),
                               ("E",f"Best Month ({best_mo})",best_val,GREEN_FLG),
                               ("H",f"Worst Month ({worst_mo})",worst_val,RED_FLG)]:
        ec = chr(ord(col)+1)
        ws7.merge_cells(f"{col}4:{ec}4"); ws7[f"{col}4"].value = label
        ws7[f"{col}4"].font = Font(name="Arial",bold=True,size=8,color="5A5A5A")
        ws7[f"{col}4"].fill = wfill(WHITE); ws7[f"{col}4"].alignment = cal("left")
        ws7.merge_cells(f"{col}5:{ec}5"); ws7[f"{col}5"].value = val
        ws7[f"{col}5"].font = Font(name="Arial",bold=True,size=16,color="1F2D3D")
        ws7[f"{col}5"].fill = wfill(bg); ws7[f"{col}5"].alignment = cal(); ws7[f"{col}5"].border = bdr()
    ws7.row_dimensions[4].height = 16; ws7.row_dimensions[5].height = 30; ws7.row_dimensions[6].height = 10
    sec(ws7,"MONTHLY MTBF TABLE",7,"J")
    for col,txt,ec in [("B","Month",None),("C","Breakdowns",None),
                        ("E","MTBF — 30-Day Window","F"),("G","MTBF — Actual Inter-Event","H"),("I","Trend","J")]:
        hdr(ws7,col,8,txt,ec)
    ws7.row_dimensions[8].height = 18
    prior = None
    for i,(mo,bkd_c,mw,ma) in enumerate(zip(months_str,bkd_mo,mtbf_30,mtbf_actual)):
        r = i+9; bg = LIGHT_ROW if i%2==0 else WHITE
        if ma and prior:
            delta = ma-prior
            trend = f"↑ +{delta:.1f}d" if delta>1 else (f"↓ {delta:.1f}d" if delta<-1 else "→ Stable")
            tbg = GREEN_FLG if delta>1 else (RED_FLG if delta<-1 else LIGHT_ROW)
        else:
            trend = "—"; tbg = bg
        ws7.merge_cells(f"E{r}:F{r}"); ws7.merge_cells(f"G{r}:H{r}"); ws7.merge_cells(f"I{r}:J{r}")
        dc(ws7,"B",r,mo,bg,bold=True); dc(ws7,"C",r,bkd_c,bg); dc(ws7,"D",r,"",bg)
        dc(ws7,"E",r,f"{mw} days" if mw else "—",bg)
        dc(ws7,"G",r,f"{ma} days" if ma else "—",bg)
        dc(ws7,"I",r,trend,tbg,align="left"); ws7.row_dimensions[r].height = 18; prior = ma
    r_s = 15
    ws7.merge_cells(f"E{r_s}:F{r_s}"); ws7.merge_cells(f"G{r_s}:H{r_s}"); ws7.merge_cells(f"I{r_s}:J{r_s}")
    for col,val in [("B","H1 2026 Avg"),("C",sum(bkd_mo)),("D",""),
                     ("E",f"{np.mean([m for m in mtbf_30 if m]):.1f} days" if any(mtbf_30) else "—"),
                     ("G",mtbf_overall),("I","See trend column")]:
        ws7[f"{col}{r_s}"].value = val
        ws7[f"{col}{r_s}"].font = Font(name="Arial",bold=True,size=9,color=WHITE)
        ws7[f"{col}{r_s}"].fill = wfill(MID_BG); ws7[f"{col}{r_s}"].alignment = cal(); ws7[f"{col}{r_s}"].border = bdr()
    ws7.row_dimensions[r_s].height = 20
    ws7.row_dimensions[17].height = 8; sec(ws7,"MTBF TREND  —  30-DAY WINDOW  (higher = better)",17,"J",HDR_BG)
    for i,(mo,mt) in enumerate(zip(months_str,mtbf_30)):
        ws7[f"L{8+i}"].value = mo; ws7[f"M{8+i}"].value = mt
    ws7["L7"].value = "Month"; ws7["M7"].value = "MTBF (days)"
    lc2 = LineChart(); lc2.title = "MTBF Trend — 30-Day Window"
    lc2.y_axis.title = "Days Between Breakdowns"; lc2.style = 10; lc2.width = 22; lc2.height = 12
    lc2.add_data(Reference(ws7,min_col=13,min_row=7,max_row=13),titles_from_data=True)
    lc2.set_categories(Reference(ws7,min_col=12,min_row=8,max_row=13))
    lc2.series[0].graphicalProperties.line.solidFill = "2E75B6"
    lc2.series[0].graphicalProperties.line.width = 28575
    lc2.series[0].marker.symbol = "circle"; lc2.series[0].marker.size = 8
    ws7.add_chart(lc2,"B18")
    for col in ["L","M"]: ws7.column_dimensions[col].hidden = True
    widths(ws7,{"A":2,"B":14,"C":16,"D":4,"E":16,"F":6,"G":18,"H":4,"I":20,"J":2})

    # ── S8: PM Compliance Detail ──
    ws8 = wb.create_sheet("PM Compliance Detail"); ws8.sheet_view.showGridLines = False
    title_row(ws8,f"PM COMPLIANCE DETAIL  ·  {asset_label}",1,"H")
    sub_row(ws8,f"Full dataset: {date_range}  |  Celonis export: Jun 25, 2026",2,"H")
    ws8.row_dimensions[3].height = 8
    for col,txt in [("B","Metric"),("C","Value"),("E","Calculation"),("H","Time Period")]:
        hdr(ws8,col,4,txt)
    ws8.merge_cells("E4:G4"); ws8.row_dimensions[4].height = 18
    comp = [
        ("Total PM WOs (GM09)",             len(pm),          "COUNT where OrderType='GM09'",                          date_range,  WHITE),
        ("PM WOs TECO'd",                   pm['is_TECO'].sum(), "COUNT PM WOs with TECO flag",                        date_range,  LIGHT_ROW),
        ("PM Compliance Rate (TECO %)",     pm_teco_pct,      "TECO'd PM / Total PM WOs",                              date_range,  GREEN_FLG),
        ("PM WOs Not TECO'd",               int((~pm['is_TECO']).sum()), "Without TECO — open or pending",             "As of export", LIGHT_ROW),
        ("PM Actual Hours",                 pm_hrs,           "SUM Actual Hours, GM09",                                date_range,  WHITE),
        ("PM Zero-Hour WOs",                f"{pm_zero} ({100*pm_zero/len(pm):.0f}%)" if len(pm) else "N/A",
                                                              "✓ PM WOs with 0 actual hours",                          date_range,  GREEN_FLG if pm_zero==0 else YELLOW),
        ("Avg PM Hours per WO",             pm_avg_hrs,       f"{pm_hrs:.1f} hrs / {len(pm)} WOs",                    date_range,  LIGHT_ROW),
        ("PM Schedule Adherence",           adh_pct,          "On-time intervals / Total intervals",                   "H1 2026",   GREEN_FLG),
        ("PM On-the-Run Rate",              f"{pm_on_run}/{len(pm)} ({pm_on_run_pct})" if len(pm) else "N/A",
                                                              "PM WOs w/ SystemCond='On the run' / Total PM",          date_range,  ORANGE),
        ("PM Machine-Down Rate",            f"{pm_mdown}/{len(pm)} ({pm_mdown_pct})" if len(pm) else "N/A",
                                                              "PM WOs w/ SystemCond='Machine must be down' / Total PM", date_range, ORANGE),
        ("Breakdown WOs (GM02)",            len(bkd),         "COUNT where OrderType='GM02'",                          date_range,  RED_FLG),
        ("Breakdown TECO %",                bkd_teco_pct,     "GM02 WOs with TECO / Total GM02",                       date_range,  LIGHT_ROW),
        ("⚠ Breakdown Zero-Hour %",        f"{bkd_zero}/{len(bkd)} ({bkd_zero_pct})" if len(bkd) else "N/A",
                                                              "GM02 WOs with 0 actual hours — labor not captured",     date_range,  YELLOW),
        ("Follow-up Corrective (GM01/PM01)",len(fuc),         "COUNT where OrderType IN ('GM01','PM01')",              date_range,  WHITE),
        ("⚠ Follow-up Corr Zero-Hr",       f"{fuc_zero}/{len(fuc)} ({fuc_zero_pct})" if len(fuc) else "N/A",
                                                              "Many are parts-only; audit non-parts zero-hour entries", date_range, YELLOW),
        ("Open WOs (no TECO/CLSD)",         open_wos,         "WOs without TECO or CLSD flag",                        "As of export", LIGHT_ROW),
    ]
    for i,(metric,val,calc,period,bg) in enumerate(comp):
        r = i+5; ws8.merge_cells(f"E{r}:G{r}")
        dc(ws8,"B",r,metric,bg,bold=True,align="left"); dc(ws8,"C",r,str(val),bg)
        dc(ws8,"E",r,calc,bg,align="left",wrap=True)
        dc(ws8,"H",r,period,bg,align="left",wrap=True); ws8.row_dimensions[r].height = 22
    widths(ws8,{"A":2,"B":34,"C":14,"D":2,"E":38,"F":8,"G":8,"H":28,"I":2})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════
# STREAMLIT UI
# ══════════════════════════════════════════════════
st.set_page_config(page_title="PM Effectiveness Analysis", page_icon="🔧", layout="centered")

st.title("PM Effectiveness Analysis")
st.caption("Upload a Celonis export to generate the full analysis workbook.")

asset_label = st.text_input("Asset label (used in report headers)", value="NEW MILFORD  ·  MBF01 | CARTONER")

uploaded = st.file_uploader("Upload Celonis Export (.xlsx)", type=["xlsx"])

if uploaded:
    with st.spinner("Running analysis…"):
        try:
            output = run_analysis(uploaded, asset_label)
            st.success("Analysis complete.")
            fname = uploaded.name.replace("Celonis_Export", "PM_Effectiveness").replace(".xlsx","") + "_Analysis.xlsx"
            st.download_button(
                label="⬇️ Download Analysis Workbook",
                data=output,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Something went wrong: {e}")
            st.info("Make sure the uploaded file is a Celonis export with the standard column structure.")

with st.expander("Expected file format"):
    st.markdown("""
The uploaded file should be a Celonis export with these columns:
- `Work Order`
- `Plant`
- `AssetName`
- `WorkOrderNumber`
- `Description`
- `OrderType` (GM09=PM, GM02=Breakdown, GM01/PM01=Follow-up Corrective)
- `Work order classification for preventive and corrective processing metrics`
- `System Status`
- `Created On`
- `Basic Start Date`
- `Actual Hours`
- `SystemCondition`
""")
